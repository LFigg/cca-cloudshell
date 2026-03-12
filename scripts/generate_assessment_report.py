#!/usr/bin/env python3
"""
Comprehensive Assessment Report Generator

Generates a multi-tab Excel report for Cohesity sizing and TCO analysis.
Combines inventory data from AWS, Azure, GCP, and M365 collectors with
cost data to provide complete environment visibility.

Tabs:
1. Executive Summary - Environment overview, sizing summary, protection status
2. Sizing Inputs - Workload inventory by type for Cohesity sizing calculator
3. Regional Distribution - Resources by region for cluster placement planning
4. Protection Analysis - Coverage percentages and snapshot analysis
5. Unprotected Resources - Prioritized list for protection planning
6. TCO Inputs - Current backup costs and Cohesity TCO calculator inputs
7. Account Detail - Multi-account/subscription breakdown
8. Raw Data - Full resource inventory for reference

Note: For M365 data, use the dedicated M365 report generator (generate_m365_report.py)
"""

import argparse
import glob
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Add lib directory to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from lib.change_rate import load_change_rate_files
from lib.constants import WORKLOAD_CATEGORIES

# =============================================================================
# CONSTANTS AND STYLING
# =============================================================================

# Excel styling
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Status colors
STATUS_COLORS = {
    'protected': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'partial': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    'unprotected': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'info': PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
}

# Provider colors
PROVIDER_COLORS = {
    'AWS': PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid"),
    'Azure': PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid"),
    'GCP': PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid"),
    'M365': PatternFill(start_color="D83B01", end_color="D83B01", fill_type="solid"),
}


# =============================================================================
# DATA LOADING AND VALIDATION
# =============================================================================

def load_json_file(filepath: str) -> Optional[Dict[str, Any]]:
    """Load and parse a JSON file."""
    try:
        with open(filepath, 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"Warning: Could not load {filepath}: {e}")
        return None


def find_latest_files(directory: str, pattern: str) -> Optional[str]:
    """Find the most recent file matching a pattern in a directory."""
    matches = glob.glob(os.path.join(directory, pattern))
    if not matches:
        return None
    # Sort by modification time, newest first
    return max(matches, key=os.path.getmtime)


def load_inventory_files(paths: List[str]) -> Tuple[List[Dict], Dict[str, Any]]:
    """
    Load and merge multiple inventory files.

    Returns:
        Tuple of (merged_resources, metadata_dict)
    """
    all_resources = []
    metadata = {
        'run_ids': [],
        'timestamps': [],
        'providers': set(),
        'accounts': set(),
        'orgs': set(),  # Orgs (AWS) or Tenants (Azure) based on parent directory
    }

    for path in paths:
        data = load_json_file(path)
        if not data:
            continue

        # Track org/tenant from parent directory name
        from pathlib import Path
        parent_dir = Path(path).parent.name
        if parent_dir:
            metadata['orgs'].add(parent_dir)

        # Extract resources
        resources = data.get('resources', [])
        all_resources.extend(resources)

        # Collect metadata
        if data.get('run_id'):
            metadata['run_ids'].append(data['run_id'])
        if data.get('timestamp'):
            metadata['timestamps'].append(data['timestamp'])
        if data.get('provider'):
            metadata['providers'].add(data['provider'])

        # Extract accounts from resources (account_id for AWS, subscription_id for Azure)
        for r in resources:
            if r.get('account_id'):
                metadata['accounts'].add(r['account_id'])
            elif r.get('subscription_id'):
                metadata['accounts'].add(r['subscription_id'])

    metadata['providers'] = list(metadata['providers'])
    metadata['accounts'] = list(metadata['accounts'])
    metadata['orgs'] = list(metadata['orgs'])

    return all_resources, metadata


def load_summary_files(paths: List[str]) -> Dict[str, Any]:
    """Load and merge summary files."""
    merged = {
        'total_resources': 0,
        'total_size_gb': 0,
        'by_provider': {},
        'by_region': {},
        'by_type': {},
    }

    for path in paths:
        data = load_json_file(path)
        if not data:
            continue

        summary = data.get('summary', {})
        merged['total_resources'] += summary.get('total_resources', 0)
        merged['total_size_gb'] += summary.get('total_size_gb', 0)

        # Merge by_region
        for region, stats in summary.get('by_region', {}).items():
            if region not in merged['by_region']:
                merged['by_region'][region] = {'count': 0, 'size_gb': 0}
            merged['by_region'][region]['count'] += stats.get('count', 0)
            merged['by_region'][region]['size_gb'] += stats.get('size_gb', 0)

    return merged


def load_cost_files(paths: List[str]) -> Dict[str, Any]:
    """Load and merge cost data files."""
    merged = {
        'total_cost': 0,
        'currency': 'USD',
        'by_provider': {},
        'by_category': {},
        'records': [],
    }

    for path in paths:
        data = load_json_file(path)
        if not data:
            continue

        # Handle top-level total_cost (actual cost_collect.py format)
        if 'total_cost' in data and isinstance(data['total_cost'], (int, float)):
            merged['total_cost'] += data['total_cost']

        # Handle summaries array format (actual cost_collect.py format)
        if 'summaries' in data:
            for summary in data['summaries']:
                provider = summary.get('provider', 'unknown')
                category = summary.get('category', 'other')
                cost = summary.get('total_cost', 0)

                if provider not in merged['by_provider']:
                    merged['by_provider'][provider] = {'total': 0, 'categories': {}}
                merged['by_provider'][provider]['total'] += cost

                if category not in merged['by_provider'][provider]['categories']:
                    merged['by_provider'][provider]['categories'][category] = 0
                merged['by_provider'][provider]['categories'][category] += cost

                # Also track service breakdown
                for service, svc_cost in summary.get('service_breakdown', {}).items():
                    svc_key = f"{category}:{service}"
                    if svc_key not in merged['by_category']:
                        merged['by_category'][svc_key] = 0
                    merged['by_category'][svc_key] += svc_cost

        # Handle older summary.providers format (test fixtures)
        if 'summary' in data:
            summary = data['summary']
            merged['total_cost'] += summary.get('total_cost', 0)

            for provider, pdata in summary.get('providers', {}).items():
                if provider not in merged['by_provider']:
                    merged['by_provider'][provider] = {'total': 0, 'categories': {}}
                merged['by_provider'][provider]['total'] += pdata.get('total_cost', 0)

                for cat, cost in pdata.get('categories', {}).items():
                    if cat not in merged['by_provider'][provider]['categories']:
                        merged['by_provider'][provider]['categories'][cat] = 0
                    merged['by_provider'][provider]['categories'][cat] += cost

        if 'records' in data:
            merged['records'].extend(data['records'])

    return merged


# =============================================================================
# DATA ANALYSIS HELPERS
# =============================================================================

def get_provider(resource_type: str) -> str:
    """Extract provider from resource type."""
    if resource_type.startswith('aws:'):
        return 'AWS'
    elif resource_type.startswith('azure:'):
        return 'Azure'
    elif resource_type.startswith('gcp:'):
        return 'GCP'
    elif resource_type.startswith('m365:'):
        return 'M365'
    elif resource_type.startswith('k8s:'):
        return 'Kubernetes'
    return 'Unknown'


def get_workload_category(resource_type: str) -> str:
    """Map resource type to workload category."""
    for _category, config in WORKLOAD_CATEGORIES.items():
        if resource_type in config['types']:
            return config['label']

    # Fallback categorization
    if 'snapshot' in resource_type:
        return 'Snapshots'
    if 'backup' in resource_type:
        return 'Backup Services'

    return 'Other'


def _is_kubernetes_node(resource: Dict) -> bool:
    """
    Check if an EC2/VM instance is a Kubernetes worker node.

    Detects EKS, AKS, GKE nodes via common tag patterns:
    - kubernetes.io/cluster/*
    - eks:cluster-name / aws:eks:cluster-name
    - aks-managed-* tags
    - gke-* labels
    """
    tags = resource.get('tags', {}) or {}

    # Check for Kubernetes-related tags
    k8s_tag_patterns = [
        'kubernetes.io/cluster/',
        'eks:cluster-name',
        'aws:eks:cluster-name',
        'alpha.eksctl.io/cluster-name',
        'eksctl.cluster.k8s.io/',
        'KubernetesCluster',
        'k8s.io/cluster-autoscaler/',
        'aks-managed-',
        'kubernetes.azure.com/',
        'gke-',
        'cloud.google.com/gke-',
    ]

    for tag_key in tags.keys():
        for pattern in k8s_tag_patterns:
            if pattern in tag_key:
                return True

    # Also check instance name patterns
    name = (resource.get('name', '') or '').lower()
    if any(p in name for p in ['eks-node', 'k8s-node', 'kubernetes-node', '-worker', '-node-']):
        # Additional check: must have k8s-related metadata or be near a cluster
        if any('kube' in k.lower() or 'eks' in k.lower() for k in tags.keys()):
            return True

    return False


def categorize_resources(resources: List[Dict]) -> Dict[str, Dict[str, Any]]:
    """
    Categorize resources by workload type.

    Note:
    - EC2 instances that are K8s worker nodes are categorized under 'Kubernetes/Containers'
    - EBS volumes attached to EC2 instances are counted under the VM's category
    - Unattached volumes remain under 'Block Storage (Unattached)'
    """
    categories: Dict[str, Dict[str, Any]] = {}

    # Build lookup of volumes by ID for efficient access
    volume_by_id: Dict[str, Dict] = {}
    for r in resources:
        if r.get('resource_type') == 'aws:ec2:volume':
            volume_by_id[r.get('resource_id', '')] = r

    # Track which volumes are attached to instances (to avoid double-counting)
    attached_volume_ids: set = set()

    # First pass: Process EC2 instances and calculate their storage
    for r in resources:
        rtype = r.get('resource_type', '')

        if rtype == 'aws:ec2:instance':
            # Get attached volumes
            meta = r.get('metadata', {}) or {}
            attached_vols = meta.get('attached_volumes', [])

            # Calculate total storage from attached volumes
            vm_storage_gb = 0
            for vol_id in attached_vols:
                if vol_id in volume_by_id:
                    vm_storage_gb += volume_by_id[vol_id].get('size_gb', 0) or 0
                    attached_volume_ids.add(vol_id)

            # Check if this is a K8s node
            if _is_kubernetes_node(r):
                category = 'Kubernetes/Containers'
            else:
                category = 'Virtual Machines'

            if category not in categories:
                categories[category] = {'count': 0, 'size_gb': 0, 'resources': []}
            categories[category]['count'] += 1
            categories[category]['size_gb'] += vm_storage_gb
            categories[category]['resources'].append(r)

        elif rtype == 'azure:vm':
            # Azure VMs - similar logic for attached disks
            meta = r.get('metadata', {}) or {}
            attached_disks = meta.get('attached_disks', [])
            vm_storage_gb = r.get('size_gb', 0) or 0
            for disk_id in attached_disks:
                if disk_id in volume_by_id:
                    vm_storage_gb += volume_by_id[disk_id].get('size_gb', 0) or 0
                    attached_volume_ids.add(disk_id)

            # Check if this is a K8s node
            if _is_kubernetes_node(r):
                category = 'Kubernetes/Containers'
            else:
                category = 'Virtual Machines'

            if category not in categories:
                categories[category] = {'count': 0, 'size_gb': 0, 'resources': []}
            categories[category]['count'] += 1
            categories[category]['size_gb'] += vm_storage_gb
            categories[category]['resources'].append(r)

        elif rtype == 'gcp:compute:instance':
            # GCP instances
            meta = r.get('metadata', {}) or {}
            attached_disks = meta.get('attached_disks', []) or meta.get('disks', [])
            vm_storage_gb = r.get('size_gb', 0) or 0
            for disk_id in attached_disks:
                if disk_id in volume_by_id:
                    vm_storage_gb += volume_by_id[disk_id].get('size_gb', 0) or 0
                    attached_volume_ids.add(disk_id)

            # Check if this is a K8s node
            if _is_kubernetes_node(r):
                category = 'Kubernetes/Containers'
            else:
                category = 'Virtual Machines'

            if category not in categories:
                categories[category] = {'count': 0, 'size_gb': 0, 'resources': []}
            categories[category]['count'] += 1
            categories[category]['size_gb'] += vm_storage_gb
            categories[category]['resources'].append(r)

    # Second pass: Process remaining resources
    for r in resources:
        rtype = r.get('resource_type', '')

        # Skip EC2/VM instances (already processed)
        if rtype in ['aws:ec2:instance', 'azure:vm', 'gcp:compute:instance']:
            continue

        # Skip database read replicas - they replicate from primary, don't backup separately
        meta = r.get('metadata', {}) or {}
        if rtype in ['aws:rds:instance', 'azure:sql:database', 'gcp:sql:instance']:
            if meta.get('is_read_replica'):
                # Track separately for reporting purposes
                if 'Read Replicas (excluded)' not in categories:
                    categories['Read Replicas (excluded)'] = {'count': 0, 'size_gb': 0, 'resources': []}
                categories['Read Replicas (excluded)']['count'] += 1
                categories['Read Replicas (excluded)']['size_gb'] += r.get('size_gb', 0) or 0
                categories['Read Replicas (excluded)']['resources'].append(r)
                continue

        # For block storage, only count unattached volumes
        if rtype in ['aws:ec2:volume', 'azure:disk', 'gcp:compute:disk']:
            if r.get('resource_id', '') in attached_volume_ids:
                continue  # Already counted under VMs
            category = 'Block Storage (Unattached)'
        else:
            category = get_workload_category(rtype)

        if category not in categories:
            categories[category] = {'count': 0, 'size_gb': 0, 'resources': []}
        categories[category]['count'] += 1
        categories[category]['size_gb'] += r.get('size_gb', 0) or 0
        categories[category]['resources'].append(r)

    return categories


def analyze_protection_status(resources: List[Dict]) -> Dict[str, Any]:
    """
    Analyze protection coverage across resources.

    Protection is determined by:
    - AWS Backup: backup_plan metadata or aws:backup:source-resource tag on snapshots
    - Snapshots: Existence of recent snapshots (within 30 days) for volumes
    - DLM: aws:dlm:lifecycle-policy-id tag on snapshots
    - RDS: Automated backup retention > 0
    - Azure: Protected item resources (azure:backup:protecteditem) with source_resource_id
    """
    # Identify protectable resources (exclude snapshots, backup plans, etc.)
    protectable_types = []
    for config in WORKLOAD_CATEGORIES.values():
        protectable_types.extend(config['types'])

    # Filter protectable resources, excluding read replicas (they replicate from primary)
    def is_protectable(r: Dict) -> bool:
        rtype = r.get('resource_type')
        if rtype not in protectable_types:
            return False
        # Exclude database read replicas across all cloud providers
        if rtype in ['aws:rds:instance', 'azure:sql:database', 'gcp:sql:instance']:
            meta = r.get('metadata', {}) or {}
            if meta.get('is_read_replica'):
                return False
        return True

    protectable = [r for r in resources if is_protectable(r)]

    # Build snapshot lookup: volume_id -> list of snapshots
    # Also track which resources have AWS Backup or DLM protection
    from datetime import datetime, timedelta
    now = datetime.now().astimezone()
    thirty_days_ago = now - timedelta(days=30)

    aws_backup_protected_vols: set = set()
    dlm_protected_vols: set = set()
    recent_snapshot_vols: set = set()  # Volumes with snapshots in last 30 days
    rds_automated_backup_dbs: set = set()  # RDS databases with automated backups
    azure_protected_resources: set = set()  # Azure resources protected by Recovery Services vault
    azure_protected_names: set = set()  # Fallback: extracted resource names from protected items
    azure_protected_sub_name: set = set()  # Fallback: (subscription, resource_name) pairs

    # Build Azure protected resource index from protected items
    for r in resources:
        if r.get('resource_type') == 'azure:backup:protecteditem':
            meta = r.get('metadata', {}) or {}
            source_id = meta.get('source_resource_id', '')
            if source_id:
                # Normalize to lowercase for case-insensitive matching
                azure_protected_resources.add(source_id.lower())
                # Also extract resource name for fallback matching (last path component)
                # e.g., /subscriptions/.../virtualMachines/my-vm -> my-vm
                parts = source_id.lower().split('/')
                if len(parts) > 1:
                    azure_protected_names.add(parts[-1])
                    # Also track subscription+name pair (handles resource group hash inconsistency)
                    if 'subscriptions' in parts:
                        sub_idx = parts.index('subscriptions') + 1
                        if sub_idx < len(parts):
                            azure_protected_sub_name.add((parts[sub_idx], parts[-1]))

    for r in resources:
        rtype = r.get('resource_type', '')
        if rtype == 'aws:ec2:snapshot':
            tags = r.get('tags', {}) or {}
            meta = r.get('metadata', {}) or {}
            vol_id = meta.get('volume_id') or r.get('parent_resource_id')

            if vol_id:
                # Check for AWS Backup
                if tags.get('aws:backup:source-resource'):
                    aws_backup_protected_vols.add(vol_id)

                # Check for DLM
                if tags.get('aws:dlm:lifecycle-policy-id'):
                    dlm_protected_vols.add(vol_id)

                # Check if recent
                start_time = meta.get('start_time', '')
                if start_time:
                    try:
                        dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                        if dt > thirty_days_ago:
                            recent_snapshot_vols.add(vol_id)
                    except Exception:
                        pass

        elif rtype in ['aws:rds:snapshot', 'aws:rds:cluster-snapshot']:
            tags = r.get('tags', {}) or {}
            meta = r.get('metadata', {}) or {}
            db_id = meta.get('db_instance_id') or meta.get('db_cluster_id') or r.get('parent_resource_id')
            if db_id:
                # Check for AWS Backup
                if tags.get('aws:backup:source-resource'):
                    aws_backup_protected_vols.add(db_id)
                # Check for automated RDS snapshots (indicates backup retention is enabled)
                if meta.get('snapshot_type') == 'automated':
                    rds_automated_backup_dbs.add(db_id)

    # Build volume-to-instance mapping for EC2
    instance_volumes: Dict[str, List[str]] = {}  # instance_id -> [vol_ids]
    volume_to_instance: Dict[str, str] = {}  # vol_id -> instance_id

    for r in resources:
        if r.get('resource_type') == 'aws:ec2:instance':
            inst_id = r.get('resource_id')
            meta = r.get('metadata', {}) or {}
            vols = meta.get('attached_volumes', [])
            if inst_id and vols:
                instance_volumes[inst_id] = vols
                for v in vols:
                    volume_to_instance[v] = inst_id

    # Calculate protection status
    protected = []
    unprotected = []

    for r in protectable:
        rtype = r.get('resource_type', '')
        metadata = r.get('metadata', {}) or {}
        tags = r.get('tags', {}) or {}
        rid = r.get('resource_id', '')

        is_protected = False

        # Check direct protection indicators
        if metadata.get('backup_plan'):
            is_protected = True
        elif metadata.get('recovery_vault'):
            is_protected = True
        elif tags.get('aws:backup:source-resource'):
            is_protected = True
        elif metadata.get('protected_by'):
            is_protected = True

        # Check Azure protected items index (cross-reference by resource_id or name)
        elif rtype.startswith('azure:') and rid:
            if rid.lower() in azure_protected_resources:
                is_protected = True
            else:
                # Fallback: match by subscription + resource name (handles resource group hash inconsistency)
                rid_parts = rid.lower().split('/')
                if len(rid_parts) > 1:
                    # Try subscription+name matching first (more precise than name-only)
                    if 'subscriptions' in rid_parts:
                        sub_idx = rid_parts.index('subscriptions') + 1
                        if sub_idx < len(rid_parts):
                            sub_name_key = (rid_parts[sub_idx], rid_parts[-1])
                            if sub_name_key in azure_protected_sub_name:
                                is_protected = True
                    # Last resort: name-only matching
                    if not is_protected and rid_parts[-1] in azure_protected_names:
                        is_protected = True

        # For EC2 volumes, check snapshot coverage
        elif rtype == 'aws:ec2:volume':
            if rid in aws_backup_protected_vols:
                is_protected = True
            elif rid in dlm_protected_vols:
                is_protected = True
            elif rid in recent_snapshot_vols:
                is_protected = True

        # For EC2 instances, check if ALL attached volumes are protected
        elif rtype == 'aws:ec2:instance':
            attached = metadata.get('attached_volumes', [])
            if attached:
                protected_vols = sum(1 for v in attached if
                                    v in aws_backup_protected_vols or
                                    v in dlm_protected_vols or
                                    v in recent_snapshot_vols)
                if protected_vols == len(attached):
                    is_protected = True
                elif protected_vols > 0:
                    # Partially protected - still count as protected but note it
                    is_protected = True
                    f'Partial ({protected_vols}/{len(attached)} volumes)'

        # For RDS, check backup retention or presence of automated snapshots
        elif rtype in ['aws:rds:instance', 'aws:rds:cluster']:
            # Get the DB identifier (could be resource_id or name)
            db_identifier = rid or r.get('name', '')

            retention = metadata.get('backup_retention_period', 0)
            if retention and retention > 0:
                is_protected = True
            elif rid in aws_backup_protected_vols:
                is_protected = True
            elif db_identifier in rds_automated_backup_dbs:
                is_protected = True
            else:
                # Check by name match (snapshot db_instance_id might match name not resource_id)
                name = r.get('name', '')
                if name and name in rds_automated_backup_dbs:
                    is_protected = True

        if is_protected:
            protected.append(r)
        else:
            unprotected.append(r)

    # Calculate totals
    total_protectable_size = sum(r.get('size_gb', 0) or 0 for r in protectable)
    protected_size = sum(r.get('size_gb', 0) or 0 for r in protected)
    unprotected_size = sum(r.get('size_gb', 0) or 0 for r in unprotected)

    return {
        'total_protectable': len(protectable),
        'protected_count': len(protected),
        'unprotected_count': len(unprotected),
        'protected_resources': protected,
        'unprotected_resources': unprotected,
        'total_size_gb': total_protectable_size,
        'protected_size_gb': protected_size,
        'unprotected_size_gb': unprotected_size,
        'coverage_percent': (len(protected) / len(protectable) * 100) if protectable else 0,
    }


def analyze_regions(resources: List[Dict]) -> Dict[str, Dict[str, Any]]:
    """Analyze resource distribution by region."""
    regions: Dict[str, Dict[str, Any]] = {}

    for r in resources:
        region = r.get('region', 'unknown')
        provider = get_provider(r.get('resource_type', ''))

        if region not in regions:
            regions[region] = {
                'count': 0,
                'size_gb': 0,
                'providers': set(),
                'types': {}
            }

        regions[region]['count'] += 1
        regions[region]['size_gb'] += r.get('size_gb', 0) or 0
        regions[region]['providers'].add(provider)

        rtype = r.get('resource_type', 'unknown')
        if rtype not in regions[region]['types']:
            regions[region]['types'][rtype] = 0
        regions[region]['types'][rtype] += 1

    # Convert sets to lists for JSON compatibility
    for region in regions:
        regions[region]['providers'] = list(regions[region]['providers'])

    return regions


def analyze_accounts(resources: List[Dict]) -> Dict[str, Dict[str, Any]]:
    """Analyze resources by account/subscription."""
    accounts: Dict[str, Dict[str, Any]] = {}

    for r in resources:
        # Use account_id (AWS) or subscription_id (Azure) or 'unknown'
        account = r.get('account_id') or r.get('subscription_id') or 'unknown'
        provider = get_provider(r.get('resource_type', ''))
        rtype = r.get('resource_type', 'unknown')
        size = r.get('size_gb', 0) or 0

        if account not in accounts:
            accounts[account] = {
                'count': 0,
                'size_gb': 0,
                'provider': '',
                'regions': set(),
                'types': {},
                'type_sizes': {}  # Track size per type
            }

        accounts[account]['count'] += 1
        accounts[account]['size_gb'] += size
        accounts[account]['provider'] = provider
        accounts[account]['regions'].add(r.get('region', 'unknown'))

        if rtype not in accounts[account]['types']:
            accounts[account]['types'][rtype] = 0
            accounts[account]['type_sizes'][rtype] = 0
        accounts[account]['types'][rtype] += 1
        accounts[account]['type_sizes'][rtype] += size

    # Convert sets to lists
    for account in accounts:
        accounts[account]['regions'] = list(accounts[account]['regions'])

    return accounts


def get_snapshots(resources: List[Dict]) -> List[Dict]:
    """Extract snapshot resources."""
    snapshot_types = [
        'aws:ec2:snapshot', 'aws:rds:snapshot', 'aws:rds:cluster-snapshot',
        'azure:snapshot', 'azure:disk:snapshot',
        'gcp:compute:snapshot',
    ]
    return [r for r in resources if r.get('resource_type') in snapshot_types]


def analyze_snapshots(resources: List[Dict]) -> Dict[str, Any]:
    """Analyze snapshot inventory."""
    snapshots = get_snapshots(resources)

    total_size = sum(s.get('size_gb', 0) or 0 for s in snapshots)

    # Group by type
    by_type = defaultdict(lambda: {'count': 0, 'size_gb': 0})
    for s in snapshots:
        rtype = s.get('resource_type', 'unknown')
        by_type[rtype]['count'] += 1
        by_type[rtype]['size_gb'] += s.get('size_gb', 0) or 0

    return {
        'total_count': len(snapshots),
        'total_size_gb': total_size,
        'by_type': dict(by_type),
        'snapshots': snapshots,
    }


def analyze_snapshot_patterns(resources: List[Dict]) -> Dict[str, Any]:
    """
    Analyze snapshot patterns to identify automated backups outside AWS Backup.

    Detects:
    - AWS Backup managed snapshots
    - DLM (Data Lifecycle Manager) managed snapshots
    - Script-based automated backups (daily/weekly/monthly patterns)
    - AMI artifacts
    - Cross-region/account copies
    - Manual/ad-hoc snapshots

    Returns detailed analysis for reporting.
    """
    ebs_snapshots = [r for r in resources if r.get('resource_type') == 'aws:ec2:snapshot']
    rds_snapshots = [r for r in resources if r.get('resource_type') in
                    ['aws:rds:snapshot', 'aws:rds:cluster-snapshot']]

    if not ebs_snapshots and not rds_snapshots:
        return {'has_data': False}

    # Categorize EBS snapshots by source
    categories = {
        'aws_backup': [],
        'dlm': [],
        'script_daily': [],
        'script_weekly': [],
        'script_monthly': [],
        'script_other': [],
        'ami_artifact': [],
        'cross_copy': [],
        'manual': [],
    }

    # Track description patterns for automated detection
    desc_patterns: Dict[str, int] = {}
    dlm_policies: Dict[str, Dict[str, Any]] = {}

    for s in ebs_snapshots:
        desc = (s.get('metadata', {}) or {}).get('description', '') or ''
        tags = s.get('tags', {}) or {}
        desc_lower = desc.lower()

        # Categorize by source
        if 'aws backup' in desc_lower or tags.get('aws:backup:source-resource'):
            categories['aws_backup'].append(s)
        elif tags.get('aws:dlm:lifecycle-policy-id'):
            categories['dlm'].append(s)
            # Track DLM policy details
            policy_id = tags.get('aws:dlm:lifecycle-policy-id')
            if policy_id:
                if policy_id not in dlm_policies:
                    dlm_policies[policy_id] = {
                        'count': 0,
                        'size_gb': 0,
                        'schedule_name': tags.get('aws:dlm:lifecycle-schedule-name', 'Unknown'),
                    }
                dlm_policies[policy_id]['count'] += 1
                dlm_policies[policy_id]['size_gb'] += s.get('size_gb', 0) or 0
        elif 'createimage' in desc_lower or 'for ami-' in desc_lower or 'destinationami' in desc_lower:
            categories['ami_artifact'].append(s)
        elif 'copied' in desc_lower:
            categories['cross_copy'].append(s)
        elif 'daily' in desc_lower:
            categories['script_daily'].append(s)
            _track_pattern(desc, desc_patterns)
        elif 'weekly' in desc_lower:
            categories['script_weekly'].append(s)
            _track_pattern(desc, desc_patterns)
        elif 'monthly' in desc_lower:
            categories['script_monthly'].append(s)
            _track_pattern(desc, desc_patterns)
        elif 'hourly' in desc_lower or 'backup' in desc_lower or 'snapshot' in desc_lower:
            # Likely automated but not explicit schedule
            categories['script_other'].append(s)
            _track_pattern(desc, desc_patterns)
        elif desc:
            # Has description but doesn't match patterns - could be manual or custom
            categories['manual'].append(s)
            _track_pattern(desc, desc_patterns)
        else:
            categories['manual'].append(s)

    # Build list of non-AWS-Backup/AMI snapshots for schedule analysis
    # Use IDs from categories instead of expensive list membership check
    non_backup = (categories['dlm'] + categories['script_daily'] + categories['script_weekly'] +
                  categories['script_monthly'] + categories['script_other'] +
                  categories['cross_copy'] + categories['manual'])
    schedule_analysis = _analyze_schedule_times(non_backup)

    # Analyze retention/age distribution
    retention_analysis = _analyze_retention(ebs_snapshots)

    # Analyze RDS snapshots
    rds_analysis = _analyze_rds_snapshots(rds_snapshots)

    # Calculate totals
    total_script_based = (len(categories['script_daily']) + len(categories['script_weekly']) +
                          len(categories['script_monthly']) + len(categories['script_other']))

    return {
        'has_data': True,
        'total_ebs_snapshots': len(ebs_snapshots),
        'total_rds_snapshots': len(rds_snapshots),
        'categories': {k: len(v) for k, v in categories.items()},
        'category_sizes': {k: sum(s.get('size_gb', 0) or 0 for s in v) for k, v in categories.items()},
        'total_script_based': total_script_based,
        'dlm_policies': dlm_policies,
        'desc_patterns': desc_patterns,
        'schedule_analysis': schedule_analysis,
        'retention_analysis': retention_analysis,
        'rds_analysis': rds_analysis,
    }


def _track_pattern(desc: str, patterns: Dict[str, int]) -> None:
    """Track description pattern, normalizing IDs."""
    # Normalize resource IDs for pattern matching
    normalized = re.sub(r'vol-[a-f0-9]+', 'vol-xxx', desc)
    normalized = re.sub(r'i-[a-f0-9]+', 'i-xxx', normalized)
    normalized = re.sub(r'snap-[a-f0-9]+', 'snap-xxx', normalized)
    normalized = re.sub(r'ami-[a-f0-9]+', 'ami-xxx', normalized)
    # Truncate long descriptions
    if len(normalized) > 100:
        normalized = normalized[:97] + '...'
    patterns[normalized] = patterns.get(normalized, 0) + 1


def _analyze_schedule_times(snapshots: List[Dict]) -> Dict[str, Any]:
    """Analyze creation times to detect scheduling patterns."""
    from collections import Counter

    hour_dist = Counter()
    dow_dist = Counter()

    for s in snapshots:
        meta = s.get('metadata', {}) or {}
        start_time = meta.get('start_time', '')
        if start_time:
            try:
                dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                hour_dist[dt.hour] += 1
                dow_dist[dt.strftime('%A')] += 1
            except Exception:
                pass

    # Find peak hours (likely scheduled times)
    total_with_time = sum(hour_dist.values())

    # Calculate percentages for peak hours
    peak_hours = []
    for hour, count in hour_dist.most_common(5):
        pct = (count / total_with_time * 100) if total_with_time else 0
        peak_hours.append((hour, pct))

    # Calculate percentages for day of week
    dow_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    dow_distribution = []
    for day in dow_order:
        count = dow_dist.get(day, 0)
        pct = (count / total_with_time * 100) if total_with_time else 0
        if count > 0:
            dow_distribution.append((day, pct))

    # Detect if there's a clear schedule (>30% in one hour = automated)
    likely_scheduled = False
    schedule_time = None
    if peak_hours and total_with_time > 0:
        top_hour, top_pct = peak_hours[0]
        if top_pct > 30:
            likely_scheduled = True
            schedule_time = f"{top_hour:02d}:00 UTC"

    return {
        'total_analyzed': total_with_time,
        'hour_distribution': dict(hour_dist),
        'day_distribution': dict(dow_dist),
        'peak_hours': peak_hours,
        'dow_distribution': dow_distribution,
        'likely_scheduled': likely_scheduled,
        'detected_schedule_time': schedule_time,
    }


def _analyze_retention(snapshots: List[Dict]) -> Dict[str, Any]:
    """Analyze snapshot ages to understand retention patterns."""
    now = datetime.now().astimezone()
    ages = []

    for s in snapshots:
        meta = s.get('metadata', {}) or {}
        start_time = meta.get('start_time', '')
        if start_time:
            try:
                dt = datetime.fromisoformat(start_time.replace('Z', '+00:00'))
                age = (now - dt).days
                if age >= 0:
                    ages.append(age)
            except Exception:
                pass

    if not ages:
        return {}

    # Age buckets matching what generate_snapshot_analysis expects
    under_7_days = sum(1 for a in ages if a < 7)
    _7_to_14_days = sum(1 for a in ages if 7 <= a < 14)
    _14_to_30_days = sum(1 for a in ages if 14 <= a < 30)
    _30_to_90_days = sum(1 for a in ages if 30 <= a < 90)
    _90_to_365_days = sum(1 for a in ages if 90 <= a < 365)
    over_365_days = sum(1 for a in ages if a >= 365)

    # Infer retention policy from distribution
    inferred_policies = []
    total = len(ages)
    if total > 0:
        if _7_to_14_days / total > 0.15:
            inferred_policies.append("~7 day retention (daily backups)")
        if _14_to_30_days / total > 0.15:
            inferred_policies.append("~14-30 day retention (weekly)")
        if _30_to_90_days / total > 0.10:
            inferred_policies.append("~30-90 day retention (monthly)")
        if over_365_days / total > 0.10:
            inferred_policies.append("Long-term/no deletion (>1 year old snapshots)")

    return {
        'under_7_days': under_7_days,
        '7_to_14_days': _7_to_14_days,
        '14_to_30_days': _14_to_30_days,
        '30_to_90_days': _30_to_90_days,
        '90_to_365_days': _90_to_365_days,
        'over_365_days': over_365_days,
        'oldest_days': max(ages),
        'newest_days': min(ages),
        'average_age': sum(ages) / len(ages),
        'inferred_policies': inferred_policies,
    }


def _analyze_rds_snapshots(snapshots: List[Dict]) -> Dict[str, Any]:
    """Analyze RDS snapshot patterns."""
    if not snapshots:
        return {'has_data': False}

    # Categorize by snapshot type
    automated = []
    manual = []
    aws_backup = []

    for s in snapshots:
        meta = s.get('metadata', {}) or {}
        snap_type = meta.get('snapshot_type', '')

        if snap_type == 'automated':
            automated.append(s)
        elif snap_type == 'awsbackup':
            aws_backup.append(s)
        else:
            manual.append(s)

    return {
        'has_data': True,
        'total': len(snapshots),
        'automated': len(automated),
        'automated_size_gb': sum(s.get('size_gb', 0) or 0 for s in automated),
        'manual': len(manual),
        'manual_size_gb': sum(s.get('size_gb', 0) or 0 for s in manual),
        'aws_backup': len(aws_backup),
        'aws_backup_size_gb': sum(s.get('size_gb', 0) or 0 for s in aws_backup),
    }


# =============================================================================
# EXCEL GENERATION HELPERS
# =============================================================================

def set_cell(ws, row: int, col: int, value: Any,
             font: Optional[Font] = None, fill: Optional[PatternFill] = None,
             border: Optional[Border] = None, alignment: Optional[Alignment] = None) -> None:
    """Set cell value with optional styling."""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment


def write_header_row(ws, row: int, headers: List[str], start_col: int = 1) -> None:
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER


def write_data_row(ws, row: int, data: List[Any], start_col: int = 1,
                   fill: Optional[PatternFill] = None) -> None:
    """Write a data row with optional styling."""
    for col_idx, value in enumerate(data, start=start_col):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def write_section_header(ws, row: int, title: str,
                         subtitle: Optional[str] = None) -> int:
    """Write a section header, return next row."""
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    if subtitle:
        ws.cell(row=row + 1, column=1, value=subtitle)
        return row + 3
    return row + 2


def set_column_widths(ws, widths: Dict[str, int]) -> None:
    """Set column widths from a dict mapping column letters to widths."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def format_number(value: float, decimals: int = 1) -> str:
    """Format number with thousands separator."""
    if value is None:
        return ''
    if decimals == 0:
        return f"{int(value):,}"
    return f"{value:,.{decimals}f}"


def format_percent(value: float) -> str:
    """Format percentage."""
    if value is None:
        return ''
    return f"{value:.1f}%"


def format_currency(value: float, currency: str = 'USD') -> str:
    """Format currency value."""
    if value is None:
        return ''
    if currency == 'USD':
        return f"${value:,.2f}"
    return f"{value:,.2f} {currency}"


# =============================================================================
# TAB GENERATORS
# =============================================================================

def generate_executive_summary(wb: Workbook, resources: List[Dict],
                                cost_data: Dict, metadata: Dict) -> None:
    """Generate Executive Summary tab."""
    ws = wb.active
    assert ws is not None, "Workbook must have an active sheet"
    ws.title = "Executive Summary"

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Cohesity Assessment Report").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    row += 3

    # === Environment Overview Section ===
    row = write_section_header(ws, row, "Environment Overview")

    # Count by provider
    provider_counts = defaultdict(int)
    provider_sizes = defaultdict(float)
    for r in resources:
        provider = get_provider(r.get('resource_type', ''))
        provider_counts[provider] += 1
        provider_sizes[provider] += r.get('size_gb', 0) or 0

    overview_data = [
        ("Total Resources", len(resources)),
        ("Total Orgs/Tenants", len(metadata.get('orgs', []))),
        ("Total Accounts/Subscriptions", len(metadata.get('accounts', []))),
        ("Cloud Providers", ', '.join(sorted(provider_counts.keys()))),
    ]

    for label, value in overview_data:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Provider breakdown
    write_header_row(ws, row, ["Provider", "Resources", "Size (GB)"])
    row += 1

    for provider in sorted(provider_counts.keys()):
        write_data_row(ws, row, [
            provider,
            provider_counts[provider],
            round(provider_sizes[provider], 1)
        ])
        # Color code provider
        if provider in PROVIDER_COLORS:
            ws.cell(row=row, column=1).fill = PROVIDER_COLORS[provider]
            ws.cell(row=row, column=1).font = Font(color="FFFFFF", bold=True)
        row += 1

    row += 2

    # === Sizing Summary Section ===
    row = write_section_header(ws, row, "Sizing Summary",
                                "(For Cohesity sizing calculator)")

    categories = categorize_resources(resources)

    write_header_row(ws, row, ["Workload Category", "Count", "Size (GB)", "Size (TB)"])
    row += 1

    total_size = 0
    for category in sorted(categories.keys()):
        data = categories[category]
        if data['count'] > 0 and category not in ['Snapshots', 'Backup Services', 'Other']:
            size_gb = data['size_gb']
            total_size += size_gb
            write_data_row(ws, row, [
                category,
                data['count'],
                round(size_gb, 1),
                round(size_gb / 1024, 2)
            ])
            row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(total_size, 1)).font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_size / 1024, 2)).font = Font(bold=True)
    row += 2

    # === Protection Status Section ===
    row = write_section_header(ws, row, "Protection Status")

    protection = analyze_protection_status(resources)

    prot_data = [
        ("Total Protectable Resources", protection['total_protectable']),
        ("Protected", protection['protected_count']),
        ("Unprotected", protection['unprotected_count']),
        ("Coverage", format_percent(protection['coverage_percent'])),
    ]

    for label, value in prot_data:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        # Color code protection status
        if label == "Protected":
            cell.fill = STATUS_COLORS['protected']
        elif label == "Unprotected":
            cell.fill = STATUS_COLORS['unprotected']
        row += 1

    row += 1

    # Size breakdown
    write_header_row(ws, row, ["Status", "Size (GB)", "Size (TB)"])
    row += 1

    for status, size_key, fill in [
        ("Protected", 'protected_size_gb', STATUS_COLORS['protected']),
        ("Unprotected", 'unprotected_size_gb', STATUS_COLORS['unprotected']),
    ]:
        size_gb = protection[size_key]
        write_data_row(ws, row, [status, round(size_gb, 1), round(size_gb / 1024, 2)])
        ws.cell(row=row, column=1).fill = fill
        row += 1

    row += 2

    # === Current Backup Cost Section ===
    if cost_data.get('total_cost', 0) > 0:
        row = write_section_header(ws, row, "Current Backup/Storage Costs",
                                    "(Monthly snapshot and backup storage costs)")

        ws.cell(row=row, column=1, value="Total Monthly Cost")
        ws.cell(row=row, column=2, value=format_currency(cost_data['total_cost']))
        row += 2

        if cost_data.get('by_provider'):
            write_header_row(ws, row, ["Provider", "Monthly Cost"])
            row += 1

            for provider, pdata in sorted(cost_data['by_provider'].items()):
                write_data_row(ws, row, [
                    provider.upper(),
                    format_currency(pdata.get('total', 0))
                ])
                row += 1

    # Set column widths
    set_column_widths(ws, {'A': 30, 'B': 20, 'C': 15, 'D': 15})


def generate_sizing_inputs(wb: Workbook, resources: List[Dict],
                          change_rate_data: Optional[Dict[str, Any]] = None) -> None:
    """Generate Sizing Inputs tab with complete coverage, protected-only, regional, and encryption breakdown."""
    ws = wb.create_sheet(title="Sizing Inputs")

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Cohesity Sizing Calculator Inputs").font = TITLE_FONT
    row += 2

    # Categorize all resources
    categorize_resources(resources)

    # Get protection status for "apples to apples" sizing
    protection = analyze_protection_status(resources)
    protected_resources = protection.get('protected_resources', [])

    # Default change rates by category (including database-specific rates)
    change_rates = {
        'Virtual Machines': 3.0,
        'Block Storage': 2.0,
        'Block Storage (Unattached)': 2.0,
        'File Storage': 1.0,
        'Object Storage': 0.5,
        'Kubernetes/Containers': 3.0,
        'Cache/In-Memory': 2.0,
        # Database-specific change rates (transaction log generation)
        'DB: MySQL/MariaDB': 5.0,
        'DB: PostgreSQL': 5.0,
        'DB: SQL Server': 5.0,
        'DB: Oracle': 5.0,
        'DB: Cosmos DB': 3.0,
        'DB: DocumentDB': 4.0,
        'DB: Neptune': 4.0,
        'DB: DynamoDB': 2.0,
        'DB: Redshift': 3.0,
        'DB: Synapse': 3.0,
        'DB: BigTable': 2.0,
        'DB: Spanner': 3.0,
        'Databases': 5.0,  # Fallback for unrecognized
    }

    priority_order = ['Virtual Machines', 'Block Storage', 'Block Storage (Unattached)',
                      'DB: MySQL/MariaDB', 'DB: PostgreSQL', 'DB: SQL Server', 'DB: Oracle',
                      'DB: Cosmos DB', 'DB: DocumentDB', 'DB: Neptune', 'DB: DynamoDB',
                      'DB: Redshift', 'DB: Synapse', 'DB: BigTable', 'DB: Spanner',
                      'File Storage', 'Object Storage', 'Kubernetes/Containers', 'Cache/In-Memory']

    # ==========================================================================
    # SECTION 1: Complete Coverage (Full Environment)
    # ==========================================================================
    row = write_section_header(ws, row, "Option 1: Complete Coverage",
                                "(Full environment protection - input into Cohesity sizing calculator)")

    # Re-categorize with database breakdown
    categories_with_db: Dict[str, Dict[str, Any]] = {}

    # Build volume lookup for VM storage calculation
    volume_by_id: Dict[str, Dict] = {}
    for r in resources:
        if r.get('resource_type') == 'aws:ec2:volume':
            volume_by_id[r.get('resource_id', '')] = r

    attached_volume_ids: set = set()

    for r in resources:
        rtype = r.get('resource_type', '')
        meta = r.get('metadata', {}) or {}
        size_gb = r.get('size_gb', 0) or 0

        # Skip replicas and snapshots
        if meta.get('is_read_replica'):
            continue
        if 'snapshot' in rtype.lower():
            continue

        # Check if database - get specific engine
        db_engine = _get_db_engine_group(r)
        if db_engine:
            category = db_engine
        elif rtype == 'aws:ec2:instance':
            # Calculate storage from attached volumes
            attached_vols = meta.get('attached_volumes', [])
            size_gb = 0
            for vol_id in attached_vols:
                if vol_id in volume_by_id:
                    size_gb += volume_by_id[vol_id].get('size_gb', 0) or 0
                    attached_volume_ids.add(vol_id)
            if _is_kubernetes_node(r):
                category = 'Kubernetes/Containers'
            else:
                category = 'Virtual Machines'
        elif rtype in ['azure:compute:vm', 'gcp:compute:instance']:
            if _is_kubernetes_node(r):
                category = 'Kubernetes/Containers'
            else:
                category = 'Virtual Machines'
        elif rtype == 'aws:ec2:volume':
            if r.get('resource_id', '') in attached_volume_ids:
                continue  # Already counted with VM
            attached = meta.get('attached_to')
            if attached:
                continue
            category = 'Block Storage (Unattached)'
        elif rtype in ['azure:compute:disk', 'gcp:compute:disk']:
            attached = meta.get('attached_to')
            if attached:
                continue
            category = 'Block Storage (Unattached)'
        elif rtype in ['aws:efs:filesystem', 'aws:fsx:filesystem', 'azure:storage:fileshare',
                       'gcp:filestore:instance']:
            category = 'File Storage'
        elif rtype in ['aws:s3:bucket', 'azure:storage:blob', 'gcp:storage:bucket']:
            category = 'Object Storage'
        elif rtype in ['aws:eks:cluster', 'azure:aks:cluster', 'gcp:container:cluster']:
            category = 'Kubernetes/Containers'
        elif rtype in ['aws:elasticache:cluster', 'azure:cache:redis', 'gcp:memorystore:instance']:
            category = 'Cache/In-Memory'
        else:
            cat = get_workload_category(rtype)
            if cat in ['Snapshots', 'Backup Services', 'Other']:
                continue
            category = cat

        if category not in categories_with_db:
            categories_with_db[category] = {'count': 0, 'size_gb': 0}
        categories_with_db[category]['count'] += 1
        categories_with_db[category]['size_gb'] += size_gb

    write_header_row(ws, row, [
        "Workload Type", "Count", "Size (GB)", "Size (TB)",
        "Daily Change Rate (%)", "Est. Daily Change (GB)"
    ])
    row += 1

    total_size = 0
    total_change = 0

    sorted_categories = sorted(
        categories_with_db.keys(),
        key=lambda x: priority_order.index(x) if x in priority_order else 100
    )

    for category in sorted_categories:
        data = categories_with_db[category]
        if category in ['Snapshots', 'Backup Services', 'Other']:
            continue
        if data['count'] == 0:
            continue

        size_gb = data['size_gb']
        total_size += size_gb

        change_rate = change_rates.get(category, 2.0)
        daily_change = size_gb * (change_rate / 100)
        total_change += daily_change

        write_data_row(ws, row, [
            category,
            data['count'],
            round(size_gb, 1),
            round(size_gb / 1024, 2),
            change_rate,
            round(daily_change, 1)
        ])
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(total_size, 1)).font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_size / 1024, 2)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(total_change, 1)).font = Font(bold=True)
    row += 3

    # ==========================================================================
    # SECTION 2: Currently Protected Only (Apples to Apples)
    # ==========================================================================
    row = write_section_header(ws, row, "Option 2: Currently Protected Only (Apples-to-Apples)",
                                "(Migrate existing backup coverage to Cohesity - same scope)")

    # Categorize only protected resources (reuse volume_by_id from above)
    protected_categories: Dict[str, Dict] = {}
    protected_volume_ids: set = set()

    for r in protected_resources:
        rtype = r.get('resource_type', '')
        meta = r.get('metadata', {}) or {}

        # Check if database - get specific engine
        db_engine = _get_db_engine_group(r)
        if db_engine:
            category = db_engine
            size_gb = r.get('size_gb', 0) or 0

        elif rtype == 'aws:ec2:instance':
            if _is_kubernetes_node(r):
                category = 'Kubernetes/Containers'
            else:
                category = 'Virtual Machines'

            # Calculate storage from attached volumes
            attached_vols = meta.get('attached_volumes', [])
            size_gb = 0
            for vol_id in attached_vols:
                if vol_id in volume_by_id:
                    size_gb += volume_by_id[vol_id].get('size_gb', 0) or 0
                    protected_volume_ids.add(vol_id)

        elif rtype == 'aws:ec2:volume':
            # Check if already counted via an instance
            if r.get('resource_id', '') in protected_volume_ids:
                continue
            category = 'Block Storage (Unattached)'
            size_gb = r.get('size_gb', 0) or 0

        elif rtype in ['aws:efs:filesystem', 'aws:fsx:filesystem']:
            category = 'File Storage'
            size_gb = r.get('size_gb', 0) or 0

        elif rtype == 'aws:s3:bucket':
            category = 'Object Storage'
            size_gb = r.get('size_gb', 0) or 0

        elif rtype in ['aws:eks:cluster', 'azure:aks:cluster', 'gcp:container:cluster']:
            category = 'Kubernetes/Containers'
            size_gb = r.get('size_gb', 0) or 0

        else:
            # Generic handling for other types
            category = get_workload_category(rtype)
            if category in ['Snapshots', 'Backup Services', 'Other']:
                continue
            size_gb = r.get('size_gb', 0) or 0

        if category not in protected_categories:
            protected_categories[category] = {'count': 0, 'size_gb': 0}
        protected_categories[category]['count'] += 1
        protected_categories[category]['size_gb'] += size_gb

    write_header_row(ws, row, [
        "Workload Type", "Count", "Size (GB)", "Size (TB)",
        "Daily Change Rate (%)", "Est. Daily Change (GB)"
    ])
    row += 1

    protected_total_size = 0
    protected_total_change = 0

    sorted_protected = sorted(
        protected_categories.keys(),
        key=lambda x: priority_order.index(x) if x in priority_order else 100
    )

    for category in sorted_protected:
        data = protected_categories[category]
        if data['count'] == 0:
            continue

        size_gb = data['size_gb']
        protected_total_size += size_gb

        change_rate = change_rates.get(category, 2.0)
        daily_change = size_gb * (change_rate / 100)
        protected_total_change += daily_change

        write_data_row(ws, row, [
            category,
            data['count'],
            round(size_gb, 1),
            round(size_gb / 1024, 2),
            change_rate,
            round(daily_change, 1)
        ])
        row += 1

    # Total row for protected
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(protected_total_size, 1)).font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(protected_total_size / 1024, 2)).font = Font(bold=True)
    ws.cell(row=row, column=6, value=round(protected_total_change, 1)).font = Font(bold=True)
    row += 2

    # Coverage comparison
    if total_size > 0:
        coverage_pct = (protected_total_size / total_size) * 100
        ws.cell(row=row, column=1, value="Currently Protected Coverage:")
        ws.cell(row=row, column=2, value=f"{coverage_pct:.1f}% of total environment")
        row += 1

        gap_size = total_size - protected_total_size
        ws.cell(row=row, column=1, value="Unprotected Gap:")
        ws.cell(row=row, column=2, value=f"{gap_size/1024:.1f} TB ({100-coverage_pct:.1f}%)")
    row += 3

    # ==========================================================================
    # SECTION 3: Regional Breakdown for CE Cluster Planning
    # ==========================================================================
    row = write_section_header(ws, row, "Option 3: Regional Breakdown (Per-Region CE Clusters)",
                                "(Each region group typically requires a dedicated Cohesity Cloud Edition cluster)")

    # Build volume lookup for VM storage calculation (reuse if available)
    regional_volume_by_id: Dict[str, Dict] = {}
    for r in resources:
        if r.get('resource_type') == 'aws:ec2:volume':
            regional_volume_by_id[r.get('resource_id', '')] = r

    # Track volumes attached to VMs (to avoid double-counting)
    regional_attached_vol_ids: set = set()

    # Build regional breakdown with encryption status
    regional_data: Dict[str, Dict[str, Dict[str, Any]]] = {}  # region_group -> category -> {count, size, encrypted_size, ...}

    for r in resources:
        rtype = r.get('resource_type', '')
        meta = r.get('metadata', {}) or {}
        region = r.get('region', 'unknown')

        # Skip replicas and snapshots
        if meta.get('is_read_replica'):
            continue
        if 'snapshot' in rtype.lower():
            continue

        # Get region group
        region_group, _ = _get_region_group(region)

        # Check if database - get specific engine first
        db_engine = _get_db_engine_group(r)
        if db_engine:
            category = db_engine
            size_gb = r.get('size_gb', 0) or 0
        elif rtype == 'aws:ec2:instance':
            if _is_kubernetes_node(r):
                category = 'Kubernetes/Containers'
            else:
                category = 'Virtual Machines'
            # Calculate storage from attached volumes
            size_gb = 0
            attached_vols = meta.get('attached_volumes', [])
            for vol_id in attached_vols:
                if vol_id in regional_volume_by_id:
                    size_gb += regional_volume_by_id[vol_id].get('size_gb', 0) or 0
                    regional_attached_vol_ids.add(vol_id)
        elif rtype == 'aws:ec2:volume':
            # Skip volumes attached to instances (already counted)
            vol_id = r.get('resource_id', '')
            if vol_id in regional_attached_vol_ids:
                continue
            attached = meta.get('attached_instance')
            if attached:
                continue  # Skip attached volumes
            category = 'Block Storage (Unattached)'
            size_gb = r.get('size_gb', 0) or 0
        elif rtype in ['azure:compute:vm', 'gcp:compute:instance']:
            if _is_kubernetes_node(r):
                category = 'Kubernetes/Containers'
            else:
                category = 'Virtual Machines'
            # Azure/GCP may have size_gb directly or in attached disks
            size_gb = r.get('size_gb', 0) or 0
        elif rtype in ['azure:compute:disk', 'gcp:compute:disk']:
            attached = meta.get('attached_to')
            if attached:
                continue
            category = 'Block Storage (Unattached)'
            size_gb = r.get('size_gb', 0) or 0
        elif rtype in ['aws:efs:filesystem', 'aws:fsx:filesystem', 'azure:storage:fileshare',
                       'gcp:filestore:instance']:
            category = 'File Storage'
            size_gb = r.get('size_gb', 0) or 0
        elif rtype in ['aws:s3:bucket', 'azure:storage:blob', 'gcp:storage:bucket']:
            category = 'Object Storage'
            size_gb = r.get('size_gb', 0) or 0
        elif rtype in ['aws:eks:cluster', 'azure:aks:cluster', 'gcp:container:cluster']:
            category = 'Kubernetes/Containers'
            size_gb = r.get('size_gb', 0) or 0
        else:
            cat = get_workload_category(rtype)
            if cat in ['Snapshots', 'Backup Services', 'Other']:
                continue
            category = cat
            size_gb = r.get('size_gb', 0) or 0

        # Check encryption status - TDE and guest-level encryption affect dedupe/compression
        # Server-side encryption (Azure EncryptionAtRestWithPlatformKey, AWS KMS, GCP default)
        # is transparent and does NOT affect Cohesity dedupe
        is_encrypted = False

        # Check for TDE on database resources
        # Azure SQL has TDE enabled by default since 2017
        if rtype in ['azure:sql:database', 'azure:sql:managedinstance']:
            is_encrypted = meta.get('tde_enabled', True)  # Default True for Azure SQL
        elif meta.get('tde_enabled', False):
            is_encrypted = True
        elif rtype in ['aws:ec2:volume', 'aws:rds:instance', 'aws:rds:cluster',
                     'aws:efs:filesystem', 'aws:fsx:filesystem']:
            # AWS: Check for guest-level encryption indicators
            # Server-side KMS encryption (encrypted=True) does NOT affect dedupe
            # Only guest-level (LUKS, BitLocker) would affect dedupe - not detectable from API
            is_encrypted = False  # AWS server-side encryption is transparent
        elif rtype.startswith('azure:'):
            # Azure: Server-side encryption (EncryptionAtRestWithPlatformKey/CustomerKey) is transparent
            # Only Azure Disk Encryption (ADE) with BitLocker/dm-crypt affects dedupe
            enc_type = meta.get('encryption_type', '')
            # ADE would show different encryption_type, platform keys are server-side
            is_encrypted = enc_type and enc_type not in [
                'EncryptionAtRestWithPlatformKey',
                'EncryptionAtRestWithCustomerKey',
                'EncryptionAtRestWithPlatformAndCustomerKeys',
                ''
            ]
        elif rtype.startswith('gcp:'):
            # GCP: Default encryption is server-side and transparent
            # Only CSEK (Customer-Supplied Encryption Keys) or guest-level affects dedupe
            is_encrypted = meta.get('guest_os_encrypted', False)

        # Initialize region group if needed
        if region_group not in regional_data:
            regional_data[region_group] = {}
        if category not in regional_data[region_group]:
            regional_data[region_group][category] = {
                'count': 0, 'size_gb': 0,
                'encrypted_count': 0, 'encrypted_size_gb': 0,
                'unencrypted_count': 0, 'unencrypted_size_gb': 0
            }

        # Update stats
        regional_data[region_group][category]['count'] += 1
        regional_data[region_group][category]['size_gb'] += size_gb
        if is_encrypted:
            regional_data[region_group][category]['encrypted_count'] += 1
            regional_data[region_group][category]['encrypted_size_gb'] += size_gb
        else:
            regional_data[region_group][category]['unencrypted_count'] += 1
            regional_data[region_group][category]['unencrypted_size_gb'] += size_gb

    # Sort region groups by total size
    sorted_regions = sorted(
        regional_data.keys(),
        key=lambda rg: sum(d['size_gb'] for d in regional_data[rg].values()),
        reverse=True
    )

    # Output each region group as a sub-section
    for region_group in sorted_regions:
        region_total_size = sum(d['size_gb'] for d in regional_data[region_group].values())
        if region_total_size < 1:  # Skip negligible regions
            continue

        # Region header
        ws.cell(row=row, column=1, value=f"Region: {region_group}").font = Font(bold=True, color="0000AA")
        ws.cell(row=row, column=2, value=f"Total: {region_total_size/1024:.2f} TB")
        row += 1

        write_header_row(ws, row, [
            "Workload Type", "Count", "Size (TB)",
            "Encrypted (TB)", "Unencrypted (TB)", "Daily Change Rate (%)"
        ])
        row += 1

        region_categories = regional_data[region_group]
        sorted_cats = sorted(
            region_categories.keys(),
            key=lambda x: priority_order.index(x) if x in priority_order else 100
        )

        region_total = 0
        for category in sorted_cats:
            data = region_categories[category]
            if data['count'] == 0:
                continue

            size_tb = data['size_gb'] / 1024
            enc_tb = data['encrypted_size_gb'] / 1024
            unenc_tb = data['unencrypted_size_gb'] / 1024
            cr = change_rates.get(category, 2.0)
            region_total += data['size_gb']

            write_data_row(ws, row, [
                category,
                data['count'],
                round(size_tb, 2),
                round(enc_tb, 2),
                round(unenc_tb, 2),
                cr
            ])
            row += 1

        # Region total
        ws.cell(row=row, column=1, value="Region Total").font = Font(bold=True)
        ws.cell(row=row, column=3, value=round(region_total / 1024, 2)).font = Font(bold=True)
        row += 2

    row += 1

    # ==========================================================================
    # SECTION 4: Workloads by Encryption Status
    # ==========================================================================
    row = write_section_header(ws, row, "Option 4: Workloads by Encryption Status",
                                "(Encrypted workloads achieve 30-50% less deduplication)")

    # Build volume lookup for VM storage calculation (same as Section 3)
    enc_volume_by_id: Dict[str, Dict] = {}
    for r in resources:
        if r.get('resource_type') == 'aws:ec2:volume':
            enc_volume_by_id[r.get('resource_id', '')] = r

    # Track volumes attached to VMs
    enc_attached_vol_ids: set = set()

    # Build encryption-separated workload summary
    enc_workloads: Dict[str, Dict[str, Any]] = {}  # "Category - Encrypted/Unencrypted" -> stats

    for r in resources:
        rtype = r.get('resource_type', '')
        meta = r.get('metadata', {}) or {}

        if meta.get('is_read_replica'):
            continue
        if 'snapshot' in rtype.lower():
            continue

        # Calculate size_gb (special handling for VMs)
        size_gb = 0
        is_encrypted = False  # Only guest-level encryption or TDE affects dedupe

        # Check if database - get specific engine first
        db_engine = _get_db_engine_group(r)
        if db_engine:
            category = db_engine
            size_gb = r.get('size_gb', 0) or 0
            # TDE (Transparent Data Encryption) DOES affect dedupe/compression
            # Azure SQL has TDE enabled by default since 2017
            if rtype in ['azure:sql:database', 'azure:sql:managedinstance']:
                is_encrypted = meta.get('tde_enabled', True)  # Default True for Azure SQL
            else:
                is_encrypted = meta.get('tde_enabled', False)
        elif rtype == 'aws:ec2:instance':
            # Calculate storage from attached volumes
            attached_vols = meta.get('attached_volumes', [])
            for vol_id in attached_vols:
                if vol_id in enc_volume_by_id:
                    vol = enc_volume_by_id[vol_id]
                    vol_size = vol.get('size_gb', 0) or 0
                    size_gb += vol_size
                    enc_attached_vol_ids.add(vol_id)
                    # AWS EBS encryption is server-side (KMS) - transparent, doesn't affect dedupe
            if _is_kubernetes_node(r):
                category = 'Kubernetes/Containers'
            else:
                category = 'Virtual Machines'
        elif rtype == 'aws:ec2:volume':
            # Skip volumes attached to instances
            vol_id = r.get('resource_id', '')
            if vol_id in enc_attached_vol_ids:
                continue
            if meta.get('attached_instance'):
                continue
            category = 'Block Storage'
            size_gb = r.get('size_gb', 0) or 0
            # AWS EBS encryption is server-side (KMS) - transparent, doesn't affect dedupe
            is_encrypted = False
        else:
            # Determine category
            category = get_workload_category(rtype)
            if category in ['Snapshots', 'Backup Services', 'Other']:
                continue
            size_gb = r.get('size_gb', 0) or 0
            # Server-side encryption (AWS KMS, Azure platform keys, GCP default) doesn't affect dedupe
            # Only guest-level encryption would affect dedupe - not detectable from cloud API
            is_encrypted = False

        # Create key with encryption status
        enc_status = "Encrypted" if is_encrypted else "Unencrypted"
        key = f"{category} - {enc_status}"

        if key not in enc_workloads:
            enc_workloads[key] = {'count': 0, 'size_gb': 0, 'category': category, 'encrypted': is_encrypted}

        enc_workloads[key]['count'] += 1
        enc_workloads[key]['size_gb'] += size_gb

    write_header_row(ws, row, [
        "Workload Type", "Encryption", "Count", "Size (GB)", "Size (TB)",
        "Daily Change Rate (%)", "Est. Daily Change (GB)"
    ])
    row += 1

    # Sort: by category first, then encrypted last (show unencrypted first)
    sorted_enc = sorted(
        enc_workloads.keys(),
        key=lambda k: (
            priority_order.index(enc_workloads[k]['category']) if enc_workloads[k]['category'] in priority_order else 100,
            1 if enc_workloads[k]['encrypted'] else 0
        )
    )

    enc_total_size = 0
    enc_total_change = 0

    for key in sorted_enc:
        data = enc_workloads[key]
        if data['count'] == 0:
            continue

        size_gb = data['size_gb']
        enc_total_size += size_gb

        cr = change_rates.get(data['category'], 2.0)
        daily_change = size_gb * (cr / 100)
        enc_total_change += daily_change

        enc_label = "Yes" if data['encrypted'] else "No"

        write_data_row(ws, row, [
            data['category'],
            enc_label,
            data['count'],
            round(size_gb, 1),
            round(size_gb / 1024, 2),
            cr,
            round(daily_change, 1)
        ])
        row += 1

    # Total
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(enc_total_size, 1)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(enc_total_size / 1024, 2)).font = Font(bold=True)
    ws.cell(row=row, column=7, value=round(enc_total_change, 1)).font = Font(bold=True)
    row += 2

    # Sizing note
    ws.cell(row=row, column=1, value="Note: Use encrypted workloads with reduced dedupe estimates in sizing calculator.")
    row += 3

    # ==========================================================================
    # SECTION 5: Encryption Summary (affects deduplication efficiency)
    # ==========================================================================
    row = write_section_header(ws, row, "Encryption Summary",
                                "(Overall encryption stats for sizing deduplication estimates)")

    # Analyze encryption across all resources
    encrypted_data = {'count': 0, 'size_gb': 0}
    unencrypted_data = {'count': 0, 'size_gb': 0}
    encryption_by_type: Dict[str, Dict[str, Any]] = {}

    for r in resources:
        rtype = r.get('resource_type', '')
        meta = r.get('metadata', {}) or {}
        size_gb = r.get('size_gb', 0) or 0

        # Skip replicas and snapshots
        if meta.get('is_read_replica'):
            continue
        if 'snapshot' in rtype.lower():
            continue

        # Check encryption status based on resource type
        # Server-side encryption (AWS KMS, Azure platform keys, GCP default) is transparent
        # TDE (Transparent Data Encryption) DOES affect dedupe - check tde_enabled
        # Guest-level encryption (BitLocker, LUKS) would affect dedupe but is not detectable
        is_encrypted = False

        # Check for TDE on database resources
        # Azure SQL has TDE enabled by default since 2017
        if rtype in ['azure:sql:database', 'azure:sql:managedinstance']:
            is_encrypted = meta.get('tde_enabled', True)  # Default True for Azure SQL
        elif meta.get('tde_enabled', False):
            is_encrypted = True

        if is_encrypted:
            encrypted_data['count'] += 1
            encrypted_data['size_gb'] += size_gb
        else:
            unencrypted_data['count'] += 1
            unencrypted_data['size_gb'] += size_gb

        # Track by type
        if rtype not in encryption_by_type:
            encryption_by_type[rtype] = {'encrypted': 0, 'encrypted_gb': 0,
                                         'unencrypted': 0, 'unencrypted_gb': 0}
        if is_encrypted:
            encryption_by_type[rtype]['encrypted'] += 1
            encryption_by_type[rtype]['encrypted_gb'] += size_gb
        else:
            encryption_by_type[rtype]['unencrypted'] += 1
            encryption_by_type[rtype]['unencrypted_gb'] += size_gb

    # Summary
    total_analyzed = encrypted_data['size_gb'] + unencrypted_data['size_gb']
    encrypted_pct = (encrypted_data['size_gb'] / total_analyzed * 100) if total_analyzed > 0 else 0

    write_header_row(ws, row, ["Status", "Resource Count", "Size (GB)", "Size (TB)", "% of Total"])
    row += 1

    write_data_row(ws, row, [
        "Encrypted",
        encrypted_data['count'],
        round(encrypted_data['size_gb'], 1),
        round(encrypted_data['size_gb'] / 1024, 2),
        f"{encrypted_pct:.1f}%"
    ])
    row += 1

    write_data_row(ws, row, [
        "Unencrypted",
        unencrypted_data['count'],
        round(unencrypted_data['size_gb'], 1),
        round(unencrypted_data['size_gb'] / 1024, 2),
        f"{100-encrypted_pct:.1f}%"
    ])
    row += 2

    # Note about sizing implications
    ws.cell(row=row, column=1, value="Note: Encrypted data typically achieves 30-50% less deduplication.")
    row += 1
    ws.cell(row=row, column=1, value="Consider this when sizing Cohesity storage capacity.")
    row += 3

    # ==========================================================================
    # SECTION 6: Database Sizing Details (Transaction Logs vs Data)
    # ==========================================================================
    row = write_section_header(ws, row, "Database Sizing Details",
                                "(Transaction logs require separate sizing from data)")

    # Collect database details by engine
    db_types = ['aws:rds:instance', 'aws:rds:cluster', 'azure:sql:database',
                'azure:sql:managedinstance', 'azure:cosmosdb:account', 'gcp:sql:instance']

    db_by_engine: Dict[str, Dict[str, Any]] = {}

    for r in resources:
        rtype = r.get('resource_type', '')
        if rtype not in db_types:
            continue

        meta = r.get('metadata', {}) or {}

        # Skip read replicas
        if meta.get('is_read_replica'):
            continue

        size_gb = r.get('size_gb', 0) or 0

        # Determine engine type
        engine = 'Unknown'
        if rtype in ['aws:rds:instance', 'aws:rds:cluster']:
            engine = meta.get('engine', 'Unknown')
        elif rtype == 'azure:sql:database':
            engine = 'SQL Server (Azure)'
        elif rtype == 'azure:sql:managedinstance':
            engine = 'SQL Server (Azure MI)'
        elif rtype == 'azure:cosmosdb:account':
            engine = 'Cosmos DB'
        elif rtype == 'gcp:sql:instance':
            db_version = meta.get('database_version', '')
            if 'MYSQL' in db_version.upper():
                engine = 'MySQL (GCP)'
            elif 'POSTGRES' in db_version.upper():
                engine = 'PostgreSQL (GCP)'
            elif 'SQLSERVER' in db_version.upper():
                engine = 'SQL Server (GCP)'
            else:
                engine = db_version or 'Unknown (GCP)'

        # Normalize similar engines
        engine_lower = engine.lower()
        if 'mysql' in engine_lower or 'mariadb' in engine_lower or 'aurora-mysql' in engine_lower:
            engine_group = 'MySQL/MariaDB'
        elif 'postgres' in engine_lower or 'aurora-postgresql' in engine_lower:
            engine_group = 'PostgreSQL'
        elif 'sqlserver' in engine_lower or 'sql server' in engine_lower:
            engine_group = 'SQL Server'
        elif 'oracle' in engine_lower:
            engine_group = 'Oracle'
        elif 'cosmos' in engine_lower:
            engine_group = 'NoSQL (Cosmos DB)'
        elif 'docdb' in engine_lower or 'documentdb' in engine_lower:
            engine_group = 'NoSQL (DocumentDB)'
        elif 'neptune' in engine_lower:
            engine_group = 'Graph DB (Neptune)'
        elif 'dynamodb' in engine_lower:
            engine_group = 'NoSQL (DynamoDB)'
        else:
            engine_group = engine

        if engine_group not in db_by_engine:
            db_by_engine[engine_group] = {
                'count': 0,
                'data_size_gb': 0,
                'encrypted_count': 0,
            }

        db_by_engine[engine_group]['count'] += 1
        db_by_engine[engine_group]['data_size_gb'] += size_gb
        # TDE (Transparent Data Encryption) DOES affect dedupe/compression
        # Azure SQL has TDE enabled by default since 2017
        if rtype in ['azure:sql:database', 'azure:sql:managedinstance']:
            if meta.get('tde_enabled', True):  # Default True for Azure SQL
                db_by_engine[engine_group]['encrypted_count'] += 1
        elif meta.get('tde_enabled', False):
            db_by_engine[engine_group]['encrypted_count'] += 1

    if db_by_engine:
        # Typical transaction log generation rates by engine (as % of data size per day)
        # These are industry estimates for OLTP workloads - used as fallback when no actual data
        tlog_rate_estimates = {
            'MySQL/MariaDB': 10.0,         # Binary logs, row-based replication
            'PostgreSQL': 15.0,            # WAL logs, MVCC overhead
            'SQL Server': 12.0,            # Transaction log, full recovery
            'Oracle': 10.0,                # Redo logs
            'NoSQL (Cosmos DB)': 5.0,      # Change feed
            'NoSQL (DocumentDB)': 8.0,     # Oplog (MongoDB compatible)
            'NoSQL (DynamoDB)': 3.0,       # Streams
            'Graph DB (Neptune)': 8.0,     # Journal logs
        }

        # Map engine groups to change rate data keys
        engine_to_cr_keys = {
            'MySQL/MariaDB': ['aws:rds-mysql', 'aws:rds-mariadb', 'aws:rds-aurora-mysql',
                             'gcp:cloudsql-mysql', 'azure:mysql'],
            'PostgreSQL': ['aws:rds-postgres', 'aws:rds-aurora-postgresql',
                          'gcp:cloudsql-postgres', 'azure:postgres'],
            'SQL Server': ['aws:rds-sqlserver', 'gcp:cloudsql-sqlserver',
                          'azure:sql-database', 'azure:sql-managedinstance'],
            'Oracle': ['aws:rds-oracle'],
            'NoSQL (Cosmos DB)': ['azure:cosmosdb'],
            'NoSQL (DocumentDB)': ['aws:documentdb'],
            'NoSQL (DynamoDB)': ['aws:dynamodb'],
            'Graph DB (Neptune)': ['aws:neptune'],
        }

        # Check if we have actual change rate data
        has_actual_data = change_rate_data and change_rate_data.get('has_actual_data', False)
        actual_change_rates = change_rate_data.get('change_rates', {}) if change_rate_data and has_actual_data else {}

        # Function to get actual tlog rate for an engine group
        def get_actual_tlog_gb(engine_group: str, data_size_gb: float) -> Tuple[Optional[float], bool]:
            """Returns (daily_tlog_gb, is_actual) tuple."""
            if not actual_change_rates:
                return None, False

            cr_keys = engine_to_cr_keys.get(engine_group, [])
            total_actual_tlog = 0.0
            found_actual = False

            for key in cr_keys:
                if key in actual_change_rates:
                    cr = actual_change_rates[key]
                    tlog = cr.get('transaction_logs')
                    if tlog and 'daily_generation_gb' in tlog:
                        total_actual_tlog += tlog.get('daily_generation_gb', 0)
                        found_actual = True

            if found_actual:
                return total_actual_tlog, True
            return None, False

        write_header_row(ws, row, [
            "Database Engine", "Count", "Data Size (TB)",
            "Daily Tlog (GB)", "Monthly Tlog (TB)", "Source", "Notes"
        ])
        row += 1

        total_db_size = 0
        total_tlog_daily = 0
        any_actual = False

        for engine_group in sorted(db_by_engine.keys()):
            data = db_by_engine[engine_group]
            data_size_tb = data['data_size_gb'] / 1024
            total_db_size += data['data_size_gb']

            # Try to get actual tlog data, fall back to estimate
            actual_tlog, is_actual = get_actual_tlog_gb(engine_group, data['data_size_gb'])

            if is_actual and actual_tlog is not None:
                daily_tlog_gb = actual_tlog
                source = "Actual"
                any_actual = True
            else:
                # Estimate transaction log generation
                tlog_rate = tlog_rate_estimates.get(engine_group, 8.0)  # Default 8%
                daily_tlog_gb = data['data_size_gb'] * (tlog_rate / 100)
                source = "Estimated"

            monthly_tlog_tb = (daily_tlog_gb * 30) / 1024
            total_tlog_daily += daily_tlog_gb

            # Encryption note
            enc_pct = (data['encrypted_count'] / data['count'] * 100) if data['count'] > 0 else 0
            notes = f"{enc_pct:.0f}% encrypted"
            if enc_pct > 50:
                notes += " (reduced dedupe)"

            write_data_row(ws, row, [
                engine_group,
                data['count'],
                round(data_size_tb, 2),
                round(daily_tlog_gb, 1),
                round(monthly_tlog_tb, 2),
                source,
                notes
            ])
            row += 1

        # Totals
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=sum(d['count'] for d in db_by_engine.values())).font = Font(bold=True)
        ws.cell(row=row, column=3, value=round(total_db_size / 1024, 2)).font = Font(bold=True)
        ws.cell(row=row, column=4, value=round(total_tlog_daily, 1)).font = Font(bold=True)
        ws.cell(row=row, column=5, value=round((total_tlog_daily * 30) / 1024, 2)).font = Font(bold=True)
        row += 2

        # Sizing guidance
        ws.cell(row=row, column=1, value="Database Backup Sizing Notes:")
        row += 1
        ws.cell(row=row, column=1, value="• Data backups: Use incremental forever (changed blocks only)")
        row += 1
        ws.cell(row=row, column=1, value="• Transaction logs: Require 100% capture (full log shipping)")
        row += 1
        if any_actual:
            ws.cell(row=row, column=1, value="• Tlog rates marked 'Actual' are from CloudWatch metrics (7-day average)")
            row += 1
            ws.cell(row=row, column=1, value="• Tlog rates marked 'Estimated' use industry-standard assumptions")
        else:
            ws.cell(row=row, column=1, value="• Tlog rates above are estimates - actual rates depend on workload activity")
            row += 1
            ws.cell(row=row, column=1, value="• For actual tlog sizing, re-run collection without --skip-change-rate flag")
        row += 1
    else:
        ws.cell(row=row, column=1, value="No databases found in inventory")
        row += 1

    row += 2

    # === Detailed Breakdown Section ===
    row = write_section_header(ws, row, "Detailed Resource Breakdown",
                                "(Resource counts by type)")

    # Count by resource type
    type_counts = defaultdict(lambda: {'count': 0, 'size_gb': 0})
    for r in resources:
        rtype = r.get('resource_type', 'unknown')
        type_counts[rtype]['count'] += 1
        type_counts[rtype]['size_gb'] += r.get('size_gb', 0) or 0

    write_header_row(ws, row, ["Resource Type", "Provider", "Count", "Size (GB)"])
    row += 1

    for rtype in sorted(type_counts.keys()):
        data = type_counts[rtype]
        write_data_row(ws, row, [
            rtype,
            get_provider(rtype),
            data['count'],
            round(data['size_gb'], 1)
        ])
        row += 1

    # Set column widths
    set_column_widths(ws, {'A': 35, 'B': 15, 'C': 12, 'D': 15, 'E': 20, 'F': 20})

    # Freeze header
    ws.freeze_panes = 'A4'


def _get_region_group(region: str) -> tuple:
    """
    Group cloud regions by geographic proximity for cluster placement.
    Returns (group_name, preferred_region) tuple.
    """
    # AWS region groupings
    AWS_GROUPS = {
        # US regions
        'us-east-1': ('US East', 'us-east-1'),
        'us-east-2': ('US East', 'us-east-1'),
        'us-west-1': ('US West', 'us-west-2'),
        'us-west-2': ('US West', 'us-west-2'),
        # Europe regions
        'eu-west-1': ('Europe West', 'eu-west-1'),
        'eu-west-2': ('Europe West', 'eu-west-1'),
        'eu-west-3': ('Europe West', 'eu-west-1'),
        'eu-central-1': ('Europe Central', 'eu-central-1'),
        'eu-central-2': ('Europe Central', 'eu-central-1'),
        'eu-north-1': ('Europe North', 'eu-north-1'),
        'eu-south-1': ('Europe South', 'eu-south-1'),
        # Asia Pacific
        'ap-northeast-1': ('Asia Pacific NE', 'ap-northeast-1'),
        'ap-northeast-2': ('Asia Pacific NE', 'ap-northeast-1'),
        'ap-northeast-3': ('Asia Pacific NE', 'ap-northeast-1'),
        'ap-southeast-1': ('Asia Pacific SE', 'ap-southeast-1'),
        'ap-southeast-2': ('Asia Pacific SE', 'ap-southeast-2'),
        'ap-southeast-3': ('Asia Pacific SE', 'ap-southeast-1'),
        'ap-south-1': ('Asia Pacific South', 'ap-south-1'),
        'ap-south-2': ('Asia Pacific South', 'ap-south-1'),
        # Other
        'ca-central-1': ('Canada', 'ca-central-1'),
        'sa-east-1': ('South America', 'sa-east-1'),
        'me-south-1': ('Middle East', 'me-south-1'),
        'af-south-1': ('Africa', 'af-south-1'),
    }

    # Azure region groupings (normalize to lowercase)
    AZURE_GROUPS = {
        'eastus': ('US East', 'eastus'),
        'eastus2': ('US East', 'eastus'),
        'westus': ('US West', 'westus2'),
        'westus2': ('US West', 'westus2'),
        'westus3': ('US West', 'westus2'),
        'centralus': ('US Central', 'centralus'),
        'northcentralus': ('US Central', 'centralus'),
        'southcentralus': ('US Central', 'centralus'),
        'westeurope': ('Europe West', 'westeurope'),
        'northeurope': ('Europe West', 'northeurope'),
        'uksouth': ('UK', 'uksouth'),
        'ukwest': ('UK', 'uksouth'),
    }

    # GCP region groupings
    GCP_GROUPS = {
        'us-east1': ('US East', 'us-east1'),
        'us-east4': ('US East', 'us-east1'),
        'us-east5': ('US East', 'us-east1'),
        'us-west1': ('US West', 'us-west1'),
        'us-west2': ('US West', 'us-west1'),
        'us-west3': ('US West', 'us-west1'),
        'us-west4': ('US West', 'us-west1'),
        'us-central1': ('US Central', 'us-central1'),
        'europe-west1': ('Europe West', 'europe-west1'),
        'europe-west2': ('Europe West', 'europe-west1'),
        'europe-west3': ('Europe West', 'europe-west1'),
        'europe-west4': ('Europe West', 'europe-west1'),
    }

    region_lower = region.lower()

    # Check each mapping
    if region in AWS_GROUPS:
        return AWS_GROUPS[region]
    if region_lower in AZURE_GROUPS:
        return AZURE_GROUPS[region_lower]
    if region_lower in GCP_GROUPS:
        return GCP_GROUPS[region_lower]

    # Default: use region as its own group
    return (region, region)


def _get_db_engine_group(resource: Dict) -> Optional[str]:
    """
    Get the normalized database engine group for a resource.
    Returns None if not a database resource.
    """
    rtype = resource.get('resource_type', '')
    meta = resource.get('metadata', {}) or {}

    # Skip read replicas
    if meta.get('is_read_replica'):
        return None

    # Determine raw engine
    engine = None
    if rtype in ['aws:rds:instance', 'aws:rds:cluster']:
        engine = meta.get('engine', 'Unknown')
    elif rtype == 'azure:sql:database':
        engine = 'SQL Server'
    elif rtype == 'azure:sql:managedinstance':
        engine = 'SQL Server'
    elif rtype == 'azure:cosmosdb:account':
        engine = 'Cosmos DB'
    elif rtype == 'gcp:sql:instance':
        db_version = meta.get('database_version', '')
        if 'MYSQL' in db_version.upper():
            engine = 'MySQL'
        elif 'POSTGRES' in db_version.upper():
            engine = 'PostgreSQL'
        elif 'SQLSERVER' in db_version.upper():
            engine = 'SQL Server'
        else:
            engine = db_version or 'Unknown'
    elif rtype == 'aws:dynamodb:table':
        return 'DB: DynamoDB'
    elif rtype == 'aws:neptune:cluster':
        return 'DB: Neptune'
    elif rtype == 'aws:docdb:cluster':
        return 'DB: DocumentDB'
    elif rtype == 'aws:redshift:cluster':
        return 'DB: Redshift'
    elif rtype == 'azure:synapse:workspace':
        return 'DB: Synapse'
    elif rtype == 'gcp:bigtable:instance':
        return 'DB: BigTable'
    elif rtype == 'gcp:spanner:instance':
        return 'DB: Spanner'
    else:
        return None  # Not a database

    # Normalize to engine groups
    engine_lower = engine.lower()
    if 'mysql' in engine_lower or 'mariadb' in engine_lower or 'aurora-mysql' in engine_lower:
        return 'DB: MySQL/MariaDB'
    elif 'postgres' in engine_lower or 'aurora-postgresql' in engine_lower:
        return 'DB: PostgreSQL'
    elif 'sqlserver' in engine_lower or 'sql server' in engine_lower:
        return 'DB: SQL Server'
    elif 'oracle' in engine_lower:
        return 'DB: Oracle'
    elif 'cosmos' in engine_lower:
        return 'DB: Cosmos DB'
    elif 'docdb' in engine_lower or 'documentdb' in engine_lower:
        return 'DB: DocumentDB'
    elif 'neptune' in engine_lower:
        return 'DB: Neptune'
    elif 'dynamodb' in engine_lower:
        return 'DB: DynamoDB'
    else:
        return f'DB: {engine}'


def generate_regional_distribution(wb: Workbook, resources: List[Dict]) -> None:
    """Generate Regional Distribution tab."""
    ws = wb.create_sheet(title="Regional Distribution")

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Regional Distribution").font = TITLE_FONT
    ws.cell(row=row + 1, column=1, value="(Use for Cohesity cluster placement planning)")
    row += 3

    regions = analyze_regions(resources)

    # === Summary by Region ===
    row = write_section_header(ws, row, "Summary by Region")

    write_header_row(ws, row, ["Region", "Provider(s)", "Resource Count", "Size (GB)", "Size (TB)"])
    row += 1

    # Sort by size descending for better visibility
    sorted_regions = sorted(regions.items(), key=lambda x: x[1]['size_gb'], reverse=True)

    for region, data in sorted_regions:
        providers = ', '.join(sorted(data['providers']))
        write_data_row(ws, row, [
            region,
            providers,
            data['count'],
            round(data['size_gb'], 1),
            round(data['size_gb'] / 1024, 2)
        ])
        row += 1

    row += 2

    # === Cluster Placement Recommendations ===
    row = write_section_header(ws, row, "Cluster Placement Recommendations")

    ws.cell(row=row, column=1, value="Grouped by geographic proximity (nearby regions consolidated):")
    row += 1

    # Group regions by geographic area
    region_groups: Dict[str, Dict] = {}

    for region, data in sorted_regions:
        if data['size_gb'] < 1:  # Skip negligible regions
            continue
        group_name, preferred = _get_region_group(region)
        if group_name not in region_groups:
            region_groups[group_name] = {'regions': [], 'size_gb': 0, 'count': 0, 'preferred': preferred}
        region_groups[group_name]['regions'].append(region)
        region_groups[group_name]['size_gb'] += data['size_gb']
        region_groups[group_name]['count'] += data['count']

    # Sort groups by size
    sorted_groups = sorted(region_groups.items(), key=lambda x: x[1]['size_gb'], reverse=True)

    # Filter to significant groups (>100 GB)
    significant_groups = [(name, data) for name, data in sorted_groups if data['size_gb'] > 100]

    if significant_groups:
        write_header_row(ws, row, ["Priority", "Geographic Area", "Regions", "Size (TB)", "Recommendation"])
        row += 1

        for idx, (group_name, data) in enumerate(significant_groups, 1):
            size_tb = data['size_gb'] / 1024
            regions_str = ', '.join(sorted(data['regions']))
            preferred = data['preferred']

            if size_tb > 50:
                rec = f"Primary cluster in {preferred}"
            elif size_tb > 10:
                rec = f"Dedicated cluster in {preferred}"
            else:
                rec = "Can be protected from primary cluster"

            write_data_row(ws, row, [idx, group_name, regions_str, round(size_tb, 2), rec])
            row += 1
    else:
        ws.cell(row=row, column=1, value="No regions with >100 GB of data")

    # Set column widths
    set_column_widths(ws, {'A': 25, 'B': 20, 'C': 15, 'D': 15, 'E': 45})

    ws.freeze_panes = 'A5'


def generate_protection_analysis(wb: Workbook, resources: List[Dict]) -> None:
    """Generate Protection Analysis tab."""
    ws = wb.create_sheet(title="Protection Analysis")

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Protection Analysis").font = TITLE_FONT
    row += 2

    protection = analyze_protection_status(resources)

    # === Coverage Summary ===
    row = write_section_header(ws, row, "Protection Coverage Summary")

    coverage_data = [
        ("Total Protectable Resources", protection['total_protectable'], ""),
        ("Protected Resources", protection['protected_count'],
         format_percent(protection['coverage_percent'])),
        ("Unprotected Resources", protection['unprotected_count'],
         format_percent(100 - protection['coverage_percent'])),
    ]

    write_header_row(ws, row, ["Metric", "Count", "Percentage"])
    row += 1

    for label, count, pct in coverage_data:
        write_data_row(ws, row, [label, count, pct])
        if "Unprotected" in label:
            ws.cell(row=row, column=1).fill = STATUS_COLORS['unprotected']
        elif "Protected" in label and "Total" not in label:
            ws.cell(row=row, column=1).fill = STATUS_COLORS['protected']
        row += 1

    row += 2

    # === Size-based Coverage ===
    row = write_section_header(ws, row, "Size-based Coverage")

    write_header_row(ws, row, ["Status", "Size (GB)", "Size (TB)", "Percentage"])
    row += 1

    total_size = protection['total_size_gb'] or 1  # Avoid division by zero

    for status, size_key, fill in [
        ("Protected", 'protected_size_gb', STATUS_COLORS['protected']),
        ("Unprotected", 'unprotected_size_gb', STATUS_COLORS['unprotected']),
    ]:
        size_gb = protection[size_key]
        pct = (size_gb / total_size * 100) if total_size else 0
        write_data_row(ws, row, [
            status,
            round(size_gb, 1),
            round(size_gb / 1024, 2),
            format_percent(pct)
        ])
        ws.cell(row=row, column=1).fill = fill
        row += 1

    row += 2

    # === Snapshot Analysis ===
    row = write_section_header(ws, row, "Snapshot Inventory")

    snapshot_analysis = analyze_snapshots(resources)

    ws.cell(row=row, column=1, value="Total Snapshots")
    ws.cell(row=row, column=2, value=snapshot_analysis['total_count'])
    row += 1
    ws.cell(row=row, column=1, value="Total Snapshot Size (GB)")
    ws.cell(row=row, column=2, value=round(snapshot_analysis['total_size_gb'], 1))
    row += 1
    ws.cell(row=row, column=1, value="Total Snapshot Size (TB)")
    ws.cell(row=row, column=2, value=round(snapshot_analysis['total_size_gb'] / 1024, 2))
    row += 2

    if snapshot_analysis['by_type']:
        write_header_row(ws, row, ["Snapshot Type", "Count", "Size (GB)"])
        row += 1

        for stype, data in sorted(snapshot_analysis['by_type'].items()):
            write_data_row(ws, row, [stype, data['count'], round(data['size_gb'], 1)])
            row += 1

    # Set column widths
    set_column_widths(ws, {'A': 35, 'B': 15, 'C': 15, 'D': 15})

    ws.freeze_panes = 'A3'


def generate_snapshot_analysis(wb: Workbook, resources: List[Dict]) -> None:
    """Generate Snapshot Analysis tab showing automated backup patterns."""
    ws = wb.create_sheet(title="Snapshot Analysis")

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Snapshot Automation Analysis").font = TITLE_FONT
    ws.cell(row=row + 1, column=1, value="(Identifies automated backup sources outside AWS Backup)")
    row += 3

    # Get snapshot analysis
    analysis = analyze_snapshot_patterns(resources)

    if not analysis or not analysis.get('has_data'):
        ws.cell(row=row, column=1, value="No snapshot data found")
        return

    total_ebs = analysis.get('total_ebs_snapshots', 0)
    categories = analysis.get('categories', {})

    # === EBS Snapshot Sources Section ===
    if total_ebs > 0:
        row = write_section_header(ws, row, "EBS Snapshot Sources",
                                   f"Total: {total_ebs:,} snapshots")

        write_header_row(ws, row, ["Source", "Count", "% of Total", "Notes"])
        row += 1

        # Combine script categories for display
        script_total = (categories.get('script_daily', 0) + categories.get('script_weekly', 0) +
                       categories.get('script_monthly', 0) + categories.get('script_other', 0))

        sources = [
            ('AWS Backup', categories.get('aws_backup', 0), 'AWS Backup service'),
            ('Data Lifecycle Manager (DLM)', categories.get('dlm', 0), 'AWS DLM policies'),
            ('Script-based (Daily)', categories.get('script_daily', 0), 'Automated scripts - daily schedule'),
            ('Script-based (Weekly)', categories.get('script_weekly', 0), 'Automated scripts - weekly schedule'),
            ('Script-based (Monthly)', categories.get('script_monthly', 0), 'Automated scripts - monthly schedule'),
            ('Script-based (Other)', categories.get('script_other', 0), 'Automated with hourly/backup/snapshot patterns'),
            ('AMI Artifacts', categories.get('ami_artifact', 0), 'Created alongside AMIs'),
            ('Cross-Region/Account Copies', categories.get('cross_copy', 0), 'Copied from other regions/accounts'),
            ('Manual/Custom', categories.get('manual', 0), 'Other or unidentified'),
        ]

        for name, count, notes in sources:
            if count > 0:
                pct = (count / total_ebs * 100) if total_ebs else 0
                write_data_row(ws, row, [name, count, f"{pct:.1f}%", notes])
                row += 1
        row += 1

        # Summary of non-AWS-Backup automation
        non_awsbackup_automated = categories.get('dlm', 0) + script_total
        if non_awsbackup_automated > 0:
            row = write_section_header(ws, row, "Automated Backups Summary", "")
            ws.cell(row=row, column=1, value="Total automated (non-AWS-Backup):")
            ws.cell(row=row, column=2, value=non_awsbackup_automated)
            pct = (non_awsbackup_automated / total_ebs * 100) if total_ebs else 0
            ws.cell(row=row, column=3, value=f"{pct:.1f}%")
            row += 2

        # === DLM Policies Section ===
        dlm_policies = analysis.get('dlm_policies', {})
        if dlm_policies:
            row = write_section_header(ws, row, "DLM Policy Details",
                                       f"({len(dlm_policies)} policies found)")

            write_header_row(ws, row, ["Policy ID", "Snapshot Count", "Total Size (GB)", "Schedule"])
            row += 1

            # Sort by count descending
            sorted_policies = sorted(dlm_policies.items(),
                                    key=lambda x: x[1]['count'], reverse=True)

            for policy_id, info in sorted_policies[:20]:
                write_data_row(ws, row, [
                    policy_id or 'Unknown',
                    info.get('count', 0),
                    round(info.get('size_gb', 0), 1),
                    info.get('schedule_name', 'Unknown')
                ])
                row += 1
            row += 1

        # === Top Automated Patterns Section ===
        desc_patterns = analysis.get('desc_patterns', {})
        if desc_patterns:
            row = write_section_header(ws, row, "Top Automated Backup Patterns",
                                       "(Description patterns indicating scripted backups)")

            write_header_row(ws, row, ["Pattern/Description", "Count", "% of Total"])
            row += 1

            # Sort by count, show top 15
            sorted_patterns = sorted(desc_patterns.items(),
                                    key=lambda x: x[1], reverse=True)

            for pattern, count in sorted_patterns[:15]:
                if count >= 10:  # Only show significant patterns
                    pct = (count / total_ebs * 100) if total_ebs else 0
                    write_data_row(ws, row, [pattern[:60], count, f"{pct:.1f}%"])
                    row += 1
            row += 1

        # === Schedule Analysis ===
        schedule = analysis.get('schedule_analysis', {})
        if schedule and schedule.get('peak_hours'):
            row = write_section_header(ws, row, "Schedule Patterns",
                                       "(Detected from snapshot creation times)")

            # Peak hours
            ws.cell(row=row, column=1, value="Peak Creation Hours (UTC):").font = HEADER_FONT
            row += 1
            for hour, pct in schedule['peak_hours'][:5]:
                ws.cell(row=row, column=1, value=f"  {hour:02d}:00")
                ws.cell(row=row, column=2, value=f"{pct:.1f}%")
                row += 1
            row += 1

            # Day distribution
            if schedule.get('dow_distribution'):
                ws.cell(row=row, column=1, value="Day of Week Distribution:").font = HEADER_FONT
                row += 1
                for day, pct in schedule['dow_distribution']:
                    ws.cell(row=row, column=1, value=f"  {day}")
                    ws.cell(row=row, column=2, value=f"{pct:.1f}%")
                    row += 1
                row += 1

            # Scheduling indicator
            if schedule.get('likely_scheduled'):
                ws.cell(row=row, column=1, value="⚠ High concentration in specific hours suggests automated scheduling")
                row += 2

        # === Retention Analysis ===
        retention = analysis.get('retention_analysis', {})
        if retention:
            row = write_section_header(ws, row, "Retention Analysis",
                                       "(Age distribution of snapshots)")

            write_header_row(ws, row, ["Age Range", "Count", "% of Total"])
            row += 1

            age_ranges = [
                ('< 7 days', retention.get('under_7_days', 0)),
                ('7-14 days', retention.get('7_to_14_days', 0)),
                ('14-30 days', retention.get('14_to_30_days', 0)),
                ('30-90 days', retention.get('30_to_90_days', 0)),
                ('90-365 days', retention.get('90_to_365_days', 0)),
                ('> 1 year', retention.get('over_365_days', 0)),
            ]

            for range_name, count in age_ranges:
                if count > 0:
                    pct = (count / total_ebs * 100) if total_ebs else 0
                    write_data_row(ws, row, [range_name, count, f"{pct:.1f}%"])
                    row += 1
            row += 1

            # Inferred retention policies
            inferred = retention.get('inferred_policies', [])
            if inferred:
                ws.cell(row=row, column=1, value="Inferred Retention Policies:").font = HEADER_FONT
                row += 1
                for policy in inferred:
                    ws.cell(row=row, column=1, value=f"  • {policy}")
                    row += 1
                row += 1

    # === RDS Snapshot Analysis ===
    rds_analysis = analysis.get('rds_analysis', {})
    total_rds = analysis.get('total_rds_snapshots', 0)
    if total_rds > 0:
        row += 1
        row = write_section_header(ws, row, "RDS Snapshot Sources",
                                   f"Total: {total_rds:,} snapshots")

        write_header_row(ws, row, ["Source", "Count", "% of Total"])
        row += 1

        rds_sources = [
            ('AWS Backup', rds_analysis.get('aws_backup', 0)),
            ('Automated (RDS)', rds_analysis.get('automated', 0)),
            ('Manual', rds_analysis.get('manual', 0)),
            ('Cross-Region Copy', rds_analysis.get('cross_region', 0)),
        ]

        for name, count in rds_sources:
            if count > 0:
                pct = (count / total_rds * 100) if total_rds else 0
                write_data_row(ws, row, [name, count, f"{pct:.1f}%"])
                row += 1

    # Set column widths
    set_column_widths(ws, {'A': 50, 'B': 15, 'C': 12, 'D': 50})

    ws.freeze_panes = 'A3'


def generate_unprotected_resources(wb: Workbook, resources: List[Dict]) -> None:
    """Generate Unprotected Resources tab."""
    ws = wb.create_sheet(title="Unprotected Resources")

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Unprotected Resources").font = TITLE_FONT
    ws.cell(row=row + 1, column=1, value="(Prioritized list for protection planning)")
    row += 3

    protection = analyze_protection_status(resources)
    unprotected = protection['unprotected_resources']

    # Sort by size descending (prioritize largest)
    unprotected_sorted = sorted(
        unprotected,
        key=lambda x: x.get('size_gb', 0) or 0,
        reverse=True
    )

    # === Summary ===
    row = write_section_header(ws, row, "Summary")

    ws.cell(row=row, column=1, value="Total Unprotected Resources")
    ws.cell(row=row, column=2, value=len(unprotected))
    row += 1
    ws.cell(row=row, column=1, value="Total Unprotected Size (GB)")
    ws.cell(row=row, column=2, value=round(protection['unprotected_size_gb'], 1))
    row += 1
    ws.cell(row=row, column=1, value="Total Unprotected Size (TB)")
    ws.cell(row=row, column=2, value=round(protection['unprotected_size_gb'] / 1024, 2))
    row += 3

    # === Resource List ===
    row = write_section_header(ws, row, "Unprotected Resource List",
                                "(Sorted by size, largest first)")

    headers = [
        "Priority", "Provider", "Account", "Region", "Resource Type",
        "Resource Name", "Resource ID", "Size (GB)", "Environment"
    ]
    write_header_row(ws, row, headers)
    row += 1

    for idx, r in enumerate(unprotected_sorted[:500], 1):  # Limit to 500 rows
        tags = r.get('tags', {}) or {}
        env = tags.get('Environment', tags.get('environment',
               tags.get('Env', tags.get('env', ''))))

        write_data_row(ws, row, [
            idx,
            get_provider(r.get('resource_type', '')),
            r.get('account_id', ''),
            r.get('region', ''),
            r.get('resource_type', ''),
            r.get('name', ''),
            r.get('resource_id', ''),
            r.get('size_gb', 0) or 0,
            env
        ])

        # Highlight production resources
        if env and env.lower() in ['prod', 'production', 'prd']:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = STATUS_COLORS['unprotected']

        row += 1

    if len(unprotected_sorted) > 500:
        ws.cell(row=row, column=1, value=f"... and {len(unprotected_sorted) - 500} more resources")

    # Set column widths
    set_column_widths(ws, {
        'A': 10, 'B': 12, 'C': 20, 'D': 15, 'E': 30,
        'F': 30, 'G': 40, 'H': 12, 'I': 15
    })

    ws.freeze_panes = 'A7'
    if unprotected_sorted:
        ws.auto_filter.ref = f"A6:I{row - 1}"


def generate_tco_inputs(wb: Workbook, resources: List[Dict],
                        cost_data: Dict) -> None:
    """Generate TCO Inputs tab."""
    ws = wb.create_sheet(title="TCO Inputs")

    row = 1

    # Title
    ws.cell(row=row, column=1, value="TCO Calculator Inputs").font = TITLE_FONT
    ws.cell(row=row + 1, column=1, value="(Use these values for Cohesity TCO analysis)")
    row += 3

    # === Current Costs Section ===
    row = write_section_header(ws, row, "Current Backup/Storage Costs",
                                "(Monthly costs from cloud billing)")

    if cost_data.get('total_cost', 0) > 0:
        ws.cell(row=row, column=1, value="Total Monthly Cost")
        ws.cell(row=row, column=2, value=format_currency(cost_data['total_cost']))
        row += 2

        # By provider
        if cost_data.get('by_provider'):
            write_header_row(ws, row, ["Provider", "Category", "Monthly Cost"])
            row += 1

            for provider, pdata in sorted(cost_data['by_provider'].items()):
                # Provider total
                write_data_row(ws, row, [
                    provider.upper(),
                    "Total",
                    format_currency(pdata.get('total', 0))
                ])
                ws.cell(row=row, column=1).font = Font(bold=True)
                row += 1

                # Category breakdown
                for category, cost in sorted(pdata.get('categories', {}).items()):
                    write_data_row(ws, row, ["", category, format_currency(cost)])
                    row += 1
    else:
        ws.cell(row=row, column=1, value="No cost data available")
        ws.cell(row=row + 1, column=1, value="Run 'python cost_collect.py' to collect backup cost data")
        row += 3

    row += 2

    # === Sizing Inputs for TCO ===
    row = write_section_header(ws, row, "TCO Sizing Inputs")

    categories = categorize_resources(resources)

    # Calculate totals
    total_size = sum(
        data['size_gb'] for cat, data in categories.items()
        if cat not in ['Snapshots', 'Backup Services', 'Other']
    )

    tco_inputs = [
        ("Total Protected Data (TB)", round(total_size / 1024, 2)),
        ("Estimated Daily Change Rate (%)", "2-5% (typical)"),
        ("Retention Period (days)", "30 (adjust as needed)"),
        ("Number of Replicas", "1-2 (disaster recovery)"),
    ]

    for label, value in tco_inputs:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # === Projected Annual Costs ===
    row = write_section_header(ws, row, "Projected Annual Costs",
                                "(Based on current monthly costs)")

    if cost_data.get('total_cost', 0) > 0:
        monthly = cost_data['total_cost']
        tco_data = [
            ("Current Monthly Cost", format_currency(monthly)),
            ("Projected Annual Cost", format_currency(monthly * 12)),
            ("Projected 3-Year Cost", format_currency(monthly * 36)),
            ("Projected 5-Year Cost", format_currency(monthly * 60)),
        ]

        for label, value in tco_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

    # Set column widths
    set_column_widths(ws, {'A': 35, 'B': 25, 'C': 20})


def generate_account_detail(wb: Workbook, resources: List[Dict]) -> None:
    """Generate Account Detail tab."""
    ws = wb.create_sheet(title="Account Detail")

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Account/Subscription Detail").font = TITLE_FONT
    row += 2

    accounts = analyze_accounts(resources)

    # === Account Summary ===
    row = write_section_header(ws, row, "Account Summary")

    write_header_row(ws, row, [
        "Account/Subscription", "Provider", "Regions", "Resources", "Size (GB)", "Size (TB)"
    ])
    row += 1

    # Sort by size descending
    sorted_accounts = sorted(
        accounts.items(),
        key=lambda x: x[1]['size_gb'],
        reverse=True
    )

    for account_id, data in sorted_accounts:
        regions = ', '.join(sorted(data['regions']))
        write_data_row(ws, row, [
            account_id,
            data['provider'],
            regions,
            data['count'],
            round(data['size_gb'], 1),
            round(data['size_gb'] / 1024, 2)
        ])

        # Color by provider
        provider = data['provider']
        if provider in PROVIDER_COLORS:
            ws.cell(row=row, column=2).fill = PROVIDER_COLORS[provider]
            ws.cell(row=row, column=2).font = Font(color="FFFFFF")

        row += 1

    row += 2

    # === Resource Type Breakdown by Account ===
    row = write_section_header(ws, row, "Resource Types by Account")

    write_header_row(ws, row, ["Account", "Resource Type", "Count", "Size (GB)"])
    row += 1

    for account_id, data in sorted_accounts:
        for rtype, count in sorted(data['types'].items()):
            # Use pre-computed size from analyze_accounts
            type_size = data.get('type_sizes', {}).get(rtype, 0)
            write_data_row(ws, row, [
                account_id,
                rtype,
                count,
                round(type_size, 1)
            ])
            row += 1

    # Set column widths
    set_column_widths(ws, {'A': 25, 'B': 35, 'C': 15, 'D': 12, 'E': 15, 'F': 12})

    ws.freeze_panes = 'A4'


def generate_raw_data(wb: Workbook, resources: List[Dict]) -> None:
    """Generate Raw Data tab."""
    ws = wb.create_sheet(title="Raw Data")

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Raw Resource Inventory").font = TITLE_FONT
    ws.cell(row=row + 1, column=1, value=f"Total resources: {len(resources)}")
    row += 3

    # Headers
    headers = [
        "Provider", "Account", "Region", "Resource Type", "Resource Name",
        "Resource ID", "Size (GB)", "Service Family", "Created", "Tags"
    ]
    write_header_row(ws, row, headers)
    row += 1

    # Sort resources by provider, then type
    sorted_resources = sorted(
        resources,
        key=lambda x: (get_provider(x.get('resource_type', '')), x.get('resource_type', ''))
    )

    for r in sorted_resources:
        tags = r.get('tags', {}) or {}
        tags_str = '; '.join(f"{k}={v}" for k, v in tags.items()) if tags else ''

        metadata = r.get('metadata', {}) or {}
        created = metadata.get('created', metadata.get('creation_date',
                   metadata.get('launch_time', '')))

        write_data_row(ws, row, [
            get_provider(r.get('resource_type', '')),
            r.get('account_id', ''),
            r.get('region', ''),
            r.get('resource_type', ''),
            r.get('name', ''),
            r.get('resource_id', ''),
            r.get('size_gb', ''),
            r.get('service_family', ''),
            created,
            tags_str[:200]  # Truncate long tags
        ])

        # Color by provider
        provider = get_provider(r.get('resource_type', ''))
        if provider in PROVIDER_COLORS:
            ws.cell(row=row, column=1).fill = PROVIDER_COLORS[provider]
            ws.cell(row=row, column=1).font = Font(color="FFFFFF")

        row += 1

    # Set column widths
    set_column_widths(ws, {
        'A': 12, 'B': 20, 'C': 15, 'D': 30, 'E': 30,
        'F': 45, 'G': 12, 'H': 15, 'I': 20, 'J': 60
    })

    ws.freeze_panes = 'A4'
    if resources:
        ws.auto_filter.ref = f"A3:J{row - 1}"


# =============================================================================
# MAIN REPORT GENERATION
# =============================================================================

def generate_report(inventory_files: List[str], cost_files: List[str],
                    output_path: str, change_rate_files: Optional[List[str]] = None) -> None:
    """
    Generate comprehensive assessment report.

    Args:
        inventory_files: List of paths to inventory JSON files
        cost_files: List of paths to cost data JSON files
        output_path: Output Excel file path
        change_rate_files: List of paths to change rate JSON files
    """
    print("Loading inventory files...")
    resources, metadata = load_inventory_files(inventory_files)

    if not resources:
        print("Error: No resources found in inventory files")
        sys.exit(1)

    print(f"  Loaded {len(resources)} resources from {len(inventory_files)} files")

    print("Loading cost data...")
    cost_data = load_cost_files(cost_files) if cost_files else {}

    if cost_data.get('total_cost', 0) > 0:
        print(f"  Found ${cost_data['total_cost']:,.2f} in monthly costs")
    else:
        print("  No cost data found")

    # Load change rate data
    change_rate_data: Dict[str, Any] = {}
    if change_rate_files:
        print("Loading change rate data...")
        change_rate_data = load_change_rate_files(change_rate_files)
        if change_rate_data.get('has_actual_data'):
            print(f"  Found actual change rate data for {len(change_rate_data.get('change_rates', {}))} service families")
        else:
            print("  No change rate data found in files")

    print("Generating report...")

    # Create workbook
    wb = Workbook()

    # Generate each tab
    generate_executive_summary(wb, resources, cost_data, metadata)
    generate_sizing_inputs(wb, resources, change_rate_data)
    generate_regional_distribution(wb, resources)
    generate_protection_analysis(wb, resources)
    generate_snapshot_analysis(wb, resources)
    generate_unprotected_resources(wb, resources)
    generate_tco_inputs(wb, resources, cost_data)
    generate_account_detail(wb, resources)
    generate_raw_data(wb, resources)

    # Save workbook
    wb.save(output_path)

    print(f"\nReport generated: {output_path}")
    print("=" * 60)
    print(f"Total Resources: {len(resources)}")
    print(f"Providers: {', '.join(metadata.get('providers', ['Unknown']))}")
    print(f"Orgs/Tenants: {len(metadata.get('orgs', []))}")
    print(f"Accounts/Subscriptions: {len(metadata.get('accounts', []))}")

    # Protection summary
    protection = analyze_protection_status(resources)
    print(f"Protection Coverage: {protection['coverage_percent']:.1f}%")
    print(f"  Protected: {protection['protected_count']}")
    print(f"  Unprotected: {protection['unprotected_count']}")


def find_data_files(directory: str) -> Tuple[List[str], List[str], List[str]]:
    """
    Auto-discover inventory, cost, and change rate files in a directory.

    Returns:
        Tuple of (inventory_files, cost_files, change_rate_files)
    """
    inventory_files = []
    cost_files = []
    change_rate_files = []

    # Find inventory files - support multiple naming patterns
    inv_patterns = [
        '**/cca_inv_*.json',       # Generic pattern
        '**/cca_aws_inv_*.json',   # AWS-specific
        '**/cca_azure_inv_*.json', # Azure-specific
        '**/cca_gcp_inv_*.json',   # GCP-specific
        '**/cca_m365_inv_*.json',  # M365-specific
    ]
    for pattern in inv_patterns:
        inv_pattern = os.path.join(directory, pattern)
        inventory_files.extend(glob.glob(inv_pattern, recursive=True))

    # Remove duplicates while preserving order
    inventory_files = list(dict.fromkeys(inventory_files))

    # Find cost summary files (cca_cost_sum_*.json)
    cost_pattern = os.path.join(directory, '**/cca_cost_sum_*.json')
    cost_files.extend(glob.glob(cost_pattern, recursive=True))

    # Find change rate files (cca_*_change_rates_*.json)
    cr_patterns = [
        '**/cca_change_rates_*.json',
        '**/cca_aws_change_rates_*.json',
        '**/cca_azure_change_rates_*.json',
        '**/cca_gcp_change_rates_*.json',
    ]
    for pattern in cr_patterns:
        cr_pattern = os.path.join(directory, pattern)
        change_rate_files.extend(glob.glob(cr_pattern, recursive=True))

    change_rate_files = list(dict.fromkeys(change_rate_files))

    return inventory_files, cost_files, change_rate_files


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Generate comprehensive assessment report for Cohesity sizing/TCO',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  # Auto-discover files in current directory
  python generate_assessment_report.py

  # Specify input directory
  python generate_assessment_report.py --directory ./output

  # Specify individual files
  python generate_assessment_report.py --inventory cca_inv_*.json --cost cca_cost_sum_*.json

  # Specify output filename
  python generate_assessment_report.py -o my_assessment.xlsx
'''
    )

    parser.add_argument(
        '--directory', '-d',
        help='Directory to search for inventory and cost files (default: current directory)',
        default='.'
    )

    parser.add_argument(
        '--inventory', '-i',
        nargs='+',
        help='Inventory file(s) to include (cca_inv_*.json)'
    )

    parser.add_argument(
        '--cost', '-c',
        nargs='+',
        help='Cost data file(s) to include (cca_cost_sum_*.json)'
    )

    parser.add_argument(
        '--output', '-o',
        help='Output Excel filename (default: assessment_report_TIMESTAMP.xlsx)'
    )

    args = parser.parse_args()

    # Determine input files
    if args.inventory:
        # Expand globs
        inventory_files = []
        for pattern in args.inventory:
            inventory_files.extend(glob.glob(pattern))
        # Still need to discover cost and change rate files
        _, cost_files_auto, change_rate_files = find_data_files(args.directory)
    else:
        # Auto-discover all files
        inventory_files, cost_files_auto, change_rate_files = find_data_files(args.directory)

    if args.cost:
        cost_files = []
        for pattern in args.cost:
            cost_files.extend(glob.glob(pattern))
    else:
        cost_files = cost_files_auto

    if not inventory_files:
        print("Error: No inventory files found")
        print(f"Looking in: {os.path.abspath(args.directory)}")
        print("\nRun collectors first:")
        print("  python collect.py --auto")
        print("  python aws_collect.py -p my-profile")
        print("  python azure_collect.py")
        sys.exit(1)

    print(f"Found {len(inventory_files)} inventory file(s)")
    for f in inventory_files:
        print(f"  - {f}")

    if cost_files:
        print(f"Found {len(cost_files)} cost file(s)")
        for f in cost_files:
            print(f"  - {f}")

    if change_rate_files:
        print(f"Found {len(change_rate_files)} change rate file(s)")
        for f in change_rate_files:
            print(f"  - {f}")

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
        output_path = f"assessment_report_{timestamp}.xlsx"

    # Generate report
    generate_report(inventory_files, cost_files, output_path, change_rate_files)


if __name__ == '__main__':
    main()
