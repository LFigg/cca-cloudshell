#!/usr/bin/env python3
"""
Microsoft 365 Assessment Report Generator

Generates a multi-tab Excel report specifically designed for M365 SaaS workloads.
Provides comprehensive visibility into Exchange, SharePoint, OneDrive, and Teams
for Cohesity DataProtect M365 sizing and planning.

Tabs:
1. Executive Summary - Tenant overview, total users, storage by service
2. Exchange Online - Mailbox types, archive status, quotas, storage distribution
3. SharePoint Online - Site types, storage distribution, deleted sites
4. OneDrive for Business - Account details, storage per user, top consumers
5. Microsoft Teams - Team count, activity metrics, archived teams
6. Sizing Inputs - Cohesity DataProtect M365 sizing calculator inputs
7. Growth Analysis - Storage trends, change rates (if usage history collected)
8. Raw Data - Full resource inventory for reference

Usage:
    python scripts/generate_m365_report.py --directory <output_folder>
    python scripts/generate_m365_report.py --inventory cca_m365_inv_*.json
"""

import argparse
import glob
import json
import os
import sys
from collections import defaultdict
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Add lib directory to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))


# =============================================================================
# CONSTANTS AND STYLING
# =============================================================================

# Excel styling - Microsoft 365 theme colors
HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")  # Microsoft blue
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(italic=True, size=10, color="666666")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# M365 Service colors
SERVICE_COLORS = {
    'Exchange': PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid"),  # Blue
    'SharePoint': PatternFill(start_color="038387", end_color="038387", fill_type="solid"),  # Teal
    'OneDrive': PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid"),  # Blue
    'Teams': PatternFill(start_color="6264A7", end_color="6264A7", fill_type="solid"),  # Purple
    'EntraID': PatternFill(start_color="00A4EF", end_color="00A4EF", fill_type="solid"),  # Light blue
}

# Status colors
STATUS_COLORS = {
    'active': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    'warning': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    'critical': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    'info': PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    'archived': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}


# =============================================================================
# DATA LOADING
# =============================================================================

def load_json_file(filepath: str) -> Optional[Dict[str, Any]]:
    """Load and parse a JSON file."""
    try:
        with open(filepath, 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError) as e:
        print(f"Warning: Could not load {filepath}: {e}")
        return None


def load_inventory_files(paths: List[str]) -> Tuple[List[Dict], Dict[str, Any]]:
    """
    Load and merge M365 inventory files.

    Returns:
        Tuple of (resources, metadata)
    """
    all_resources = []
    metadata = {
        'tenant_id': None,
        'tenant_name': None,
        'run_ids': [],
        'timestamps': [],
        'total_user_count': 0,
    }

    for path in paths:
        data = load_json_file(path)
        if not data:
            continue

        # Handle both list format (resources directly) and envelope format ({"resources": [...]})
        if isinstance(data, list):
            resources = data
        else:
            resources = data.get('resources', [])
        all_resources.extend(resources)

        # Collect metadata (only available in envelope format)
        if isinstance(data, dict):
            if data.get('tenant_id') and not metadata['tenant_id']:
                metadata['tenant_id'] = data['tenant_id']
            if data.get('tenant_name') and not metadata['tenant_name']:
                metadata['tenant_name'] = data['tenant_name']
            if data.get('run_id'):
                metadata['run_ids'].append(data['run_id'])
            if data.get('timestamp'):
                metadata['timestamps'].append(data['timestamp'])

    return all_resources, metadata


def load_summary_data(inventory_paths: List[str]) -> Dict[str, Any]:
    """
    Load M365 summary data from corresponding summary files.
    """
    summary_data = {
        'exchange_mailbox_breakdown': None,
        'sharepoint_site_breakdown': None,
        'teams_activity': None,
        'change_rates': None,
        'total_user_count': 0,
        'total_capacity_gb': 0,
        # Cohesity sizer format summaries
        'exchange_summary': None,
        'sharepoint_summary': None,
        'onedrive_summary': None,
        'growth_rates': None,
        # Tenant and licensing info
        'tenant_info': None,
        'licensing': None,
        'tenant_name': None,
        'primary_domain': None,
    }

    for inv_path in inventory_paths:
        # Try to find corresponding summary file
        sum_path = inv_path.replace('_inv_', '_sum_')

        data = load_json_file(sum_path)
        if not data:
            continue

        if 'exchange_mailbox_breakdown' in data:
            summary_data['exchange_mailbox_breakdown'] = data['exchange_mailbox_breakdown']
        if 'sharepoint_site_breakdown' in data:
            summary_data['sharepoint_site_breakdown'] = data['sharepoint_site_breakdown']
        if 'teams_activity' in data:
            summary_data['teams_activity'] = data['teams_activity']
        if 'change_rates' in data:
            summary_data['change_rates'] = data['change_rates']
        if 'total_user_count' in data:
            summary_data['total_user_count'] = data['total_user_count']
        if 'total_capacity_gb' in data:
            summary_data['total_capacity_gb'] = data['total_capacity_gb']
        # Cohesity sizer format summaries
        if 'exchange_summary' in data:
            summary_data['exchange_summary'] = data['exchange_summary']
        if 'sharepoint_summary' in data:
            summary_data['sharepoint_summary'] = data['sharepoint_summary']
        if 'onedrive_summary' in data:
            summary_data['onedrive_summary'] = data['onedrive_summary']
        if 'growth_rates' in data:
            summary_data['growth_rates'] = data['growth_rates']
        # Tenant and licensing info
        if 'tenant_info' in data:
            summary_data['tenant_info'] = data['tenant_info']
        if 'licensing' in data:
            summary_data['licensing'] = data['licensing']
        if 'tenant_name' in data:
            summary_data['tenant_name'] = data['tenant_name']
        if 'primary_domain' in data:
            summary_data['primary_domain'] = data['primary_domain']

    return summary_data


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def write_header_row(ws, row: int, headers: List[str]) -> None:
    """Write a formatted header row."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')


def write_data_row(ws, row: int, values: List[Any], borders: bool = True) -> None:
    """Write a data row with optional borders."""
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        if borders:
            cell.border = THIN_BORDER


def write_section_header(ws, row: int, title: str, subtitle: Optional[str] = None) -> int:
    """Write a section header and return the next row."""
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    if subtitle:
        ws.cell(row=row + 1, column=1, value=subtitle).font = SUBTITLE_FONT
        return row + 3
    return row + 2


def set_column_widths(ws, widths: Dict[str, int]) -> None:
    """Set column widths from a dictionary."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def format_size(size_gb: float) -> str:
    """Format size with appropriate units."""
    if size_gb >= 1024:
        return f"{size_gb / 1024:.2f} TB"
    return f"{size_gb:.1f} GB"


def format_number(num: int) -> str:
    """Format number with thousand separators."""
    return f"{num:,}"


# =============================================================================
# REPORT GENERATION - EXECUTIVE SUMMARY
# =============================================================================

def generate_executive_summary(wb: Workbook, resources: List[Dict],
                                metadata: Dict, summary_data: Dict) -> None:
    """Generate Executive Summary tab."""
    ws = wb.active
    assert ws is not None, "Workbook must have an active sheet"
    ws.title = "Executive Summary"

    row = 1

    # Title
    ws.cell(row=row, column=1, value="Microsoft 365 Assessment Report").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}").font = SUBTITLE_FONT
    row += 3

    # === Tenant Information ===
    row = write_section_header(ws, row, "Tenant Information")

    # Get tenant name from summary data (new) or metadata (fallback)
    tenant_name = summary_data.get('tenant_name') or metadata.get('tenant_name') or 'N/A'
    primary_domain = summary_data.get('primary_domain') or 'N/A'

    tenant_info = [
        ("Tenant Name", tenant_name),
        ("Primary Domain", primary_domain),
        ("Tenant ID", metadata.get('tenant_id', 'N/A')),
        ("Total Users", format_number(summary_data.get('total_user_count', 0))),
        ("Collection Date", metadata['timestamps'][0][:10] if metadata.get('timestamps') else 'N/A'),
    ]

    for label, value in tenant_info:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 2

    # === Licensing Information ===
    licensing = summary_data.get('licensing')
    if licensing:
        row = write_section_header(ws, row, "Microsoft 365 Licensing Summary")

        # Filter to M365-related SKUs (exclude free/viral/standalone add-ons)
        skus = licensing.get('skus', [])
        # All patterns UPPERCASE for case-insensitive matching
        m365_sku_patterns = (
            # Core M365/O365 suites
            'SPE_',           # Secure Productive Enterprise (M365 E3/E5/F1/F3/F5)
            'M365_',          # M365 F1/F3 variants
            'MICROSOFT_365',  # Microsoft 365 E3/E5/F3/Copilot
            'O365_',          # Office 365 E1/E3/E5
            'ENTERPRISE',     # E1/E3/E5 plans
            'BUSINESS',       # Business Basic/Standard/Premium
            # Core workload standalone
            'EXCHANGE',       # Exchange Online plans
            'SHAREPOINT',     # SharePoint Online plans
            'PROJECT',        # Project Online
            'VISIO',          # Visio
            'TEAMS',          # Teams standalone/premium
            'EMS',            # Enterprise Mobility + Security
            'AAD_PREMIUM',    # Azure AD Premium
            'DEFENDER',       # Defender plans
            'MDATP',          # Microsoft Defender ATP
            'INTUNE',         # Intune
        )
        # Patterns to exclude (free/viral/trial) - UPPERCASE
        exclude_patterns = (
            '_FREE', '_VIRAL', '_TRIAL', 'PREVIEW', 'DEVELOPERPACK',
            'FLOW_FREE', 'POWER_BI_STANDARD', 'STREAM', 'WINDOWS_STORE',
            'FORMS_PRO', 'CCIBOTS', 'SPZA_IW', 'SMB_APPS', 'MADEIRA',
        )

        def is_m365_sku(sku_name: str) -> bool:
            name_upper = sku_name.upper()
            # Exclude free/viral/trial SKUs
            for excl in exclude_patterns:
                if excl in name_upper:
                    return False
            # Include M365-related SKUs
            for pattern in m365_sku_patterns:
                if pattern in name_upper:
                    return True
            return False

        m365_skus = [s for s in skus if is_m365_sku(s.get('name', ''))]

        # Calculate totals for M365 SKUs only
        m365_purchased = sum(s.get('purchased', 0) for s in m365_skus)
        m365_consumed = sum(s.get('consumed', 0) for s in m365_skus)

        ws.cell(row=row, column=1, value="M365 Licenses Purchased")
        ws.cell(row=row, column=2, value=format_number(m365_purchased))
        row += 1
        ws.cell(row=row, column=1, value="M365 Licenses Consumed")
        ws.cell(row=row, column=2, value=format_number(m365_consumed))
        row += 1
        ws.cell(row=row, column=1, value="M365 Licenses Available")
        ws.cell(row=row, column=2, value=format_number(m365_purchased - m365_consumed))
        row += 2

        # M365 SKUs breakdown
        if m365_skus:
            write_header_row(ws, row, ["SKU Name", "Purchased", "Consumed", "Available"])
            row += 1

            # Sort by consumed descending (most active SKUs first)
            sorted_skus = sorted(m365_skus, key=lambda x: x.get('consumed', 0), reverse=True)
            for sku in sorted_skus:
                purchased = sku.get('purchased', 0)
                consumed = sku.get('consumed', 0)
                write_data_row(ws, row, [
                    sku.get('name', 'Unknown'),
                    format_number(purchased),
                    format_number(consumed),
                    format_number(purchased - consumed)
                ])
                row += 1

        row += 2

    # === Storage Summary by Service ===
    row = write_section_header(ws, row, "Storage Summary by Service")

    # Categorize resources
    # Teams storage comes from SharePoint Group/Channel sites, so we need to separate them
    services = {
        'Exchange Online': {'count': 0, 'size_gb': 0, 'type': 'mailbox'},
        'OneDrive for Business': {'count': 0, 'size_gb': 0, 'type': 'onedrive'},
        'SharePoint Online': {'count': 0, 'size_gb': 0, 'type': 'sharepoint'},
        'Microsoft Teams': {'count': 0, 'size_gb': 0, 'type': 'teams'},
    }
    
    # Track Teams-related SharePoint separately
    teams_sp_count = 0
    teams_sp_size = 0

    for r in resources:
        rtype = r.get('resource_type', '')
        size = r.get('size_gb', 0) or 0
        metadata = r.get('metadata', {}) or {}

        if 'mailbox' in rtype:
            services['Exchange Online']['count'] += 1
            services['Exchange Online']['size_gb'] += size
        elif 'onedrive' in rtype:
            services['OneDrive for Business']['count'] += 1
            services['OneDrive for Business']['size_gb'] += size
        elif 'sharepoint' in rtype:
            # Check if this is a Teams-related site
            template = metadata.get('root_web_template', '')
            if template in ('Group', 'Team Channel'):
                # This is Teams storage - count separately
                teams_sp_count += 1
                teams_sp_size += size
            else:
                # Non-Teams SharePoint
                services['SharePoint Online']['count'] += 1
                services['SharePoint Online']['size_gb'] += size
        elif 'teams' in rtype:
            services['Microsoft Teams']['count'] += 1
            # Get Teams storage from metadata (correlated from SharePoint Group sites)
            teams_storage = metadata.get('sharepoint_storage_gb') or metadata.get('estimated_storage_gb', 0)
            services['Microsoft Teams']['size_gb'] += teams_storage

    # If no Teams storage from metadata but we have Teams-related SP sites, use that
    if services['Microsoft Teams']['size_gb'] == 0 and teams_sp_size > 0:
        services['Microsoft Teams']['size_gb'] = teams_sp_size

    # If SharePoint has no resources but we have change rate data, estimate size
    change_rates = summary_data.get('change_rates', {})
    sp_change = change_rates.get('SharePoint', {})
    if services['SharePoint Online']['count'] == 0 and sp_change:
        daily_change_gb = sp_change.get('daily_change_gb', 0)
        daily_change_pct = sp_change.get('daily_change_percent', 0)
        if daily_change_gb > 0 and daily_change_pct > 0:
            estimated_size = daily_change_gb / (daily_change_pct / 100)
            services['SharePoint Online']['size_gb'] = estimated_size
            services['SharePoint Online']['estimated'] = True

    write_header_row(ws, row, ["Service", "Objects", "Storage (GB)", "Storage (TB)"])
    row += 1

    total_size = 0
    total_count = 0
    for service, data in services.items():
        if data['count'] > 0 or data.get('size_gb', 0) > 0:
            total_size += data['size_gb']
            total_count += data['count']
            count_display = format_number(data['count']) if data['count'] > 0 else "Unknown*"
            service_display = service + " (estimated)" if data.get('estimated') else service
            write_data_row(ws, row, [
                service_display,
                count_display,
                f"{data['size_gb']:,.1f}",
                f"{data['size_gb'] / 1024:.2f}"
            ])
            row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=format_number(total_count)).font = Font(bold=True)
    ws.cell(row=row, column=3, value=f"{total_size:,.1f}").font = Font(bold=True)
    ws.cell(row=row, column=4, value=f"{total_size / 1024:.2f}").font = Font(bold=True)
    row += 2

    # Note about Teams/SharePoint separation
    teams_correlation = summary_data.get('teams_sharepoint_correlation', {})
    if teams_correlation:
        ws.cell(row=row, column=1, value="* Note: Teams storage = SharePoint Group + Channel sites. SharePoint = non-Teams sites only.").font = Font(italic=True)
    row += 2

    # === Quick Stats ===
    row = write_section_header(ws, row, "Quick Statistics")

    # Exchange stats
    mailbox_breakdown = summary_data.get('exchange_mailbox_breakdown', {})
    if mailbox_breakdown:
        archive_count = mailbox_breakdown.get('archive_enabled_count', 0)
        deleted_count = mailbox_breakdown.get('soft_deleted_count', 0)

        ws.cell(row=row, column=1, value="Exchange:")
        row += 1
        ws.cell(row=row, column=1, value="  Archive-enabled mailboxes")
        ws.cell(row=row, column=2, value=format_number(archive_count))
        row += 1
        ws.cell(row=row, column=1, value="  Soft-deleted mailboxes")
        ws.cell(row=row, column=2, value=format_number(deleted_count))
        row += 2

    # SharePoint stats
    sp_breakdown = summary_data.get('sharepoint_site_breakdown', {})
    if sp_breakdown:
        deleted_sites = sp_breakdown.get('deleted_count', 0)
        ws.cell(row=row, column=1, value="SharePoint:")
        row += 1
        ws.cell(row=row, column=1, value="  Deleted sites")
        ws.cell(row=row, column=2, value=format_number(deleted_sites))
        row += 2

    # Teams stats
    teams_activity = summary_data.get('teams_activity', {})
    if teams_activity:
        ws.cell(row=row, column=1, value="Teams Activity:")
        row += 1
        ws.cell(row=row, column=1, value="  Active users")
        ws.cell(row=row, column=2, value=format_number(teams_activity.get('active_users', 0)))
        row += 1
        ws.cell(row=row, column=1, value="  Total messages (30 days)")
        total_msgs = teams_activity.get('team_chat_messages', 0) + teams_activity.get('private_chat_messages', 0)
        ws.cell(row=row, column=2, value=format_number(total_msgs))
        row += 1

    set_column_widths(ws, {'A': 35, 'B': 20, 'C': 15, 'D': 15})


# =============================================================================
# REPORT GENERATION - EXCHANGE ONLINE
# =============================================================================

def generate_exchange_tab(wb: Workbook, resources: List[Dict], summary_data: Dict) -> None:
    """Generate Exchange Online tab."""
    ws = wb.create_sheet(title="Exchange Online")

    row = 1
    ws.cell(row=row, column=1, value="Exchange Online Analysis").font = TITLE_FONT
    row += 3

    # Filter mailbox resources
    mailboxes = [r for r in resources if 'mailbox' in r.get('resource_type', '')]

    if not mailboxes:
        ws.cell(row=row, column=1, value="No Exchange mailbox data collected")
        return

    # === Mailbox Type Breakdown ===
    row = write_section_header(ws, row, "Mailbox Type Distribution")

    mailbox_types = defaultdict(lambda: {'count': 0, 'size_gb': 0})
    archive_count = 0
    deleted_count = 0

    for r in mailboxes:
        metadata = r.get('metadata', {}) or {}
        mtype = metadata.get('mailbox_type', 'Unknown')
        mailbox_types[mtype]['count'] += 1
        mailbox_types[mtype]['size_gb'] += r.get('size_gb', 0) or 0

        if metadata.get('has_archive'):
            archive_count += 1
        if metadata.get('is_deleted'):
            deleted_count += 1

    write_header_row(ws, row, ["Mailbox Type", "Count", "Size (GB)", "Size (TB)", "Avg Size (GB)"])
    row += 1

    # Store data start row for chart

    total_size = 0
    for mtype in ['User', 'Shared', 'Room', 'Equipment', 'Group', 'Discovery', 'Unknown']:
        data = mailbox_types.get(mtype, {'count': 0, 'size_gb': 0})
        if data['count'] > 0:
            total_size += data['size_gb']
            avg_size = data['size_gb'] / data['count'] if data['count'] > 0 else 0
            write_data_row(ws, row, [
                mtype,
                data['count'],
                round(data['size_gb'], 1),
                round(data['size_gb'] / 1024, 2),
                round(avg_size, 2)
            ])
            row += 1

    row - 1

    # Total
    avg_total = total_size / len(mailboxes) if mailboxes else 0
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(mailboxes)).font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(total_size, 1)).font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_size / 1024, 2)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(avg_total, 2)).font = Font(bold=True)
    row += 3

    # === Archive and Deleted Status ===
    row = write_section_header(ws, row, "Mailbox Status")

    write_header_row(ws, row, ["Status", "Count", "Percentage"])
    row += 1

    write_data_row(ws, row, ["Archive Enabled", archive_count, f"{archive_count/len(mailboxes)*100:.1f}%"])
    row += 1
    write_data_row(ws, row, ["Soft Deleted", deleted_count, f"{deleted_count/len(mailboxes)*100:.1f}%"])
    row += 1
    write_data_row(ws, row, ["Active (not deleted)", len(mailboxes) - deleted_count,
                             f"{(len(mailboxes)-deleted_count)/len(mailboxes)*100:.1f}%"])
    row += 3

    # === Quota Analysis ===
    row = write_section_header(ws, row, "Quota Utilization")

    quota_ranges = {'0-25%': 0, '25-50%': 0, '50-75%': 0, '75-90%': 0, '90-100%': 0, 'Over quota': 0}

    for r in mailboxes:
        metadata = r.get('metadata', {}) or {}
        usage_pct = metadata.get('quota_usage_percent', 0) or 0

        if usage_pct >= 100:
            quota_ranges['Over quota'] += 1
        elif usage_pct >= 90:
            quota_ranges['90-100%'] += 1
        elif usage_pct >= 75:
            quota_ranges['75-90%'] += 1
        elif usage_pct >= 50:
            quota_ranges['50-75%'] += 1
        elif usage_pct >= 25:
            quota_ranges['25-50%'] += 1
        else:
            quota_ranges['0-25%'] += 1

    write_header_row(ws, row, ["Quota Usage Range", "Mailbox Count", "Percentage"])
    row += 1

    for range_name, count in quota_ranges.items():
        pct = count / len(mailboxes) * 100 if mailboxes else 0
        write_data_row(ws, row, [range_name, count, f"{pct:.1f}%"])

        # Color critical ranges
        if range_name == 'Over quota':
            ws.cell(row=row, column=1).fill = STATUS_COLORS['critical']
        elif range_name == '90-100%':
            ws.cell(row=row, column=1).fill = STATUS_COLORS['warning']
        row += 1
    row += 2

    # === Top 10 Largest Mailboxes ===
    row = write_section_header(ws, row, "Top 10 Largest Mailboxes")

    sorted_mailboxes = sorted(mailboxes, key=lambda x: x.get('size_gb', 0) or 0, reverse=True)[:10]

    write_header_row(ws, row, ["User", "Type", "Size (GB)", "Items", "Quota %"])
    row += 1

    for r in sorted_mailboxes:
        metadata = r.get('metadata', {}) or {}
        write_data_row(ws, row, [
            r.get('name', 'Unknown'),
            metadata.get('mailbox_type', 'Unknown'),
            round(r.get('size_gb', 0) or 0, 2),
            format_number(metadata.get('item_count', 0) or 0),
            f"{metadata.get('quota_usage_percent', 0) or 0:.1f}%"
        ])
        row += 1

    set_column_widths(ws, {'A': 35, 'B': 15, 'C': 15, 'D': 15, 'E': 15})


# =============================================================================
# REPORT GENERATION - SHAREPOINT ONLINE
# =============================================================================

def generate_sharepoint_tab(wb: Workbook, resources: List[Dict], summary_data: Dict) -> None:
    """Generate SharePoint Online tab."""
    ws = wb.create_sheet(title="SharePoint Online")

    row = 1
    ws.cell(row=row, column=1, value="SharePoint Online Analysis").font = TITLE_FONT
    row += 3

    # Filter SharePoint resources
    sites = [r for r in resources if 'sharepoint' in r.get('resource_type', '')]

    if not sites:
        ws.cell(row=row, column=1, value="No SharePoint site data collected")
        return

    # === Site Type Breakdown ===
    row = write_section_header(ws, row, "Site Type Distribution")

    site_types = defaultdict(lambda: {'count': 0, 'size_gb': 0})
    deleted_count = 0

    for r in sites:
        metadata = r.get('metadata', {}) or {}
        # Use root_web_template if available, fall back to site_type
        stype = metadata.get('root_web_template') or metadata.get('site_type') or 'Unknown'
        if not stype or stype == '':
            stype = 'Unknown'
        site_types[stype]['count'] += 1
        site_types[stype]['size_gb'] += r.get('size_gb', 0) or 0

        if metadata.get('is_deleted'):
            deleted_count += 1

    write_header_row(ws, row, ["Site Type", "Count", "Size (GB)", "Size (TB)", "Avg Size (GB)", "Notes"])
    row += 1

    # Site type display names and notes
    site_type_info = {
        'Group': ('M365 Group Sites', 'Includes Teams file storage'),
        'Team Channel': ('Private/Shared Channels', 'Separate from main Team site'),
        'Team Site': ('Classic Team Sites', 'Not associated with Teams'),
        'Communication Site': ('Communication Sites', 'Publishing/intranet sites'),
        'Site Page Publishing': ('Publishing Sites', 'Content publishing'),
        'Personal': ('Personal (OneDrive)', 'User OneDrive sites'),
    }

    total_size = 0
    # Sort by size descending
    for stype, data in sorted(site_types.items(), key=lambda x: -x[1]['size_gb']):
        if data['count'] > 0:
            total_size += data['size_gb']
            avg_size = data['size_gb'] / data['count'] if data['count'] > 0 else 0
            display_name, note = site_type_info.get(stype, (stype, ''))
            write_data_row(ws, row, [
                display_name,
                data['count'],
                round(data['size_gb'], 1),
                round(data['size_gb'] / 1024, 2),
                round(avg_size, 2),
                note
            ])
            row += 1

    # Total
    avg_total = total_size / len(sites) if sites else 0
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(sites)).font = Font(bold=True)
    ws.cell(row=row, column=3, value=round(total_size, 1)).font = Font(bold=True)
    ws.cell(row=row, column=4, value=round(total_size / 1024, 2)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=round(avg_total, 2)).font = Font(bold=True)
    row += 3

    # === Site Status ===
    row = write_section_header(ws, row, "Site Status")

    write_header_row(ws, row, ["Status", "Count", "Percentage"])
    row += 1

    active_count = len(sites) - deleted_count
    write_data_row(ws, row, ["Active Sites", active_count, f"{active_count/len(sites)*100:.1f}%"])
    row += 1
    write_data_row(ws, row, ["Deleted Sites", deleted_count, f"{deleted_count/len(sites)*100:.1f}%"])
    ws.cell(row=row, column=1).fill = STATUS_COLORS['archived']
    row += 3

    # === Storage Distribution ===
    row = write_section_header(ws, row, "Storage Distribution")

    size_ranges = {'< 1 GB': 0, '1-10 GB': 0, '10-100 GB': 0, '100-500 GB': 0, '> 500 GB': 0}

    for r in sites:
        size = r.get('size_gb', 0) or 0
        if size < 1:
            size_ranges['< 1 GB'] += 1
        elif size < 10:
            size_ranges['1-10 GB'] += 1
        elif size < 100:
            size_ranges['10-100 GB'] += 1
        elif size < 500:
            size_ranges['100-500 GB'] += 1
        else:
            size_ranges['> 500 GB'] += 1

    write_header_row(ws, row, ["Size Range", "Site Count", "Percentage"])
    row += 1

    for range_name, count in size_ranges.items():
        pct = count / len(sites) * 100 if sites else 0
        write_data_row(ws, row, [range_name, count, f"{pct:.1f}%"])
        row += 1
    row += 2

    # === Top 10 Largest Sites ===
    row = write_section_header(ws, row, "Top 10 Largest Sites")

    sorted_sites = sorted(sites, key=lambda x: x.get('size_gb', 0) or 0, reverse=True)[:10]

    write_header_row(ws, row, ["Site Name", "Type", "Size (GB)", "Files", "URL"])
    row += 1

    for r in sorted_sites:
        metadata = r.get('metadata', {}) or {}
        write_data_row(ws, row, [
            r.get('name', 'Unknown')[:40],  # Truncate long names
            metadata.get('site_type', 'Unknown'),
            round(r.get('size_gb', 0) or 0, 2),
            format_number(metadata.get('file_count', 0) or 0),
            metadata.get('web_url', '')[:50]  # Truncate URLs
        ])
        row += 1

    set_column_widths(ws, {'A': 40, 'B': 20, 'C': 15, 'D': 15, 'E': 50})


# =============================================================================
# REPORT GENERATION - ONEDRIVE
# =============================================================================

def generate_onedrive_tab(wb: Workbook, resources: List[Dict]) -> None:
    """Generate OneDrive for Business tab."""
    ws = wb.create_sheet(title="OneDrive for Business")

    row = 1
    ws.cell(row=row, column=1, value="OneDrive for Business Analysis").font = TITLE_FONT
    row += 3

    # Filter OneDrive resources
    drives = [r for r in resources if 'onedrive' in r.get('resource_type', '')]

    if not drives:
        ws.cell(row=row, column=1, value="No OneDrive data collected")
        return

    # === Summary Statistics ===
    row = write_section_header(ws, row, "OneDrive Summary")

    total_size = sum(r.get('size_gb', 0) or 0 for r in drives)
    avg_size = total_size / len(drives) if drives else 0

    # Find largest and smallest
    sorted_drives = sorted(drives, key=lambda x: x.get('size_gb', 0) or 0, reverse=True)
    largest = sorted_drives[0] if sorted_drives else None

    active_drives = [d for d in drives if (d.get('size_gb', 0) or 0) > 0.001]  # > 1 MB

    stats = [
        ("Total OneDrive Accounts", format_number(len(drives))),
        ("Active Accounts (>1 MB)", format_number(len(active_drives))),
        ("Total Storage (GB)", f"{total_size:,.1f}"),
        ("Total Storage (TB)", f"{total_size / 1024:.2f}"),
        ("Average per User (GB)", f"{avg_size:.2f}"),
        ("Largest Account", f"{largest.get('name', 'N/A')} ({largest.get('size_gb', 0):.1f} GB)" if largest else "N/A"),
    ]

    write_header_row(ws, row, ["Metric", "Value"])
    row += 1

    for label, value in stats:
        write_data_row(ws, row, [label, value])
        row += 1
    row += 2

    # === Storage Distribution ===
    row = write_section_header(ws, row, "Storage Distribution by User")

    size_ranges = {
        'Empty (0 GB)': 0,
        '< 1 GB': 0,
        '1-10 GB': 0,
        '10-50 GB': 0,
        '50-100 GB': 0,
        '100-200 GB': 0,
        '> 200 GB': 0
    }

    for r in drives:
        size = r.get('size_gb', 0) or 0
        if size < 0.001:
            size_ranges['Empty (0 GB)'] += 1
        elif size < 1:
            size_ranges['< 1 GB'] += 1
        elif size < 10:
            size_ranges['1-10 GB'] += 1
        elif size < 50:
            size_ranges['10-50 GB'] += 1
        elif size < 100:
            size_ranges['50-100 GB'] += 1
        elif size < 200:
            size_ranges['100-200 GB'] += 1
        else:
            size_ranges['> 200 GB'] += 1

    write_header_row(ws, row, ["Storage Range", "User Count", "Percentage", "Cumulative %"])
    row += 1

    cumulative = 0
    for range_name, count in size_ranges.items():
        pct = count / len(drives) * 100 if drives else 0
        cumulative += pct
        write_data_row(ws, row, [range_name, count, f"{pct:.1f}%", f"{cumulative:.1f}%"])
        row += 1
    row += 2

    # === Top 20 OneDrive Users ===
    row = write_section_header(ws, row, "Top 20 OneDrive Users by Storage")

    write_header_row(ws, row, ["User", "Storage (GB)", "% of Total"])
    row += 1

    for r in sorted_drives[:20]:
        size = r.get('size_gb', 0) or 0
        pct = size / total_size * 100 if total_size > 0 else 0
        write_data_row(ws, row, [
            r.get('name', 'Unknown'),
            round(size, 2),
            f"{pct:.2f}%"
        ])
        row += 1

    set_column_widths(ws, {'A': 40, 'B': 20, 'C': 15, 'D': 15})


# =============================================================================
# REPORT GENERATION - MICROSOFT TEAMS
# =============================================================================

def generate_teams_tab(wb: Workbook, resources: List[Dict], summary_data: Dict) -> None:
    """Generate Microsoft Teams tab."""
    ws = wb.create_sheet(title="Microsoft Teams")

    row = 1
    ws.cell(row=row, column=1, value="Microsoft Teams Analysis").font = TITLE_FONT
    row += 3

    # Filter Teams resources
    teams = [r for r in resources if 'teams' in r.get('resource_type', '')]
    teams_activity = summary_data.get('teams_activity', {})

    if not teams and not teams_activity:
        ws.cell(row=row, column=1, value="No Microsoft Teams data collected")
        return

    # === Teams Summary ===
    if teams:
        row = write_section_header(ws, row, "Teams Overview")

        archived = [t for t in teams if (t.get('metadata', {}) or {}).get('is_archived')]
        active = len(teams) - len(archived)

        # Visibility breakdown
        visibility = defaultdict(int)
        for t in teams:
            vis = (t.get('metadata', {}) or {}).get('visibility', 'Unknown')
            visibility[vis] += 1

        write_header_row(ws, row, ["Metric", "Count"])
        row += 1

        write_data_row(ws, row, ["Total Teams", len(teams)])
        row += 1
        write_data_row(ws, row, ["Active Teams", active])
        ws.cell(row=row, column=1).fill = STATUS_COLORS['active']
        row += 1
        write_data_row(ws, row, ["Archived Teams", len(archived)])
        ws.cell(row=row, column=1).fill = STATUS_COLORS['archived']
        row += 3

        # Visibility breakdown
        row = write_section_header(ws, row, "Teams by Visibility")

        write_header_row(ws, row, ["Visibility", "Count", "Percentage"])
        row += 1

        for vis, count in sorted(visibility.items(), key=lambda x: -x[1]):
            pct = count / len(teams) * 100 if teams else 0
            write_data_row(ws, row, [vis, count, f"{pct:.1f}%"])
            row += 1
        row += 2

    # === Teams Activity ===
    if teams_activity:
        row = write_section_header(ws, row, "Teams Activity (Last 30 Days)")

        write_header_row(ws, row, ["Activity Metric", "Count"])
        row += 1

        activity_metrics = [
            ("Active Users", teams_activity.get('active_users', 0)),
            ("Team Chat Messages", teams_activity.get('team_chat_messages', 0)),
            ("Private Chat Messages", teams_activity.get('private_chat_messages', 0)),
            ("Total Messages", teams_activity.get('team_chat_messages', 0) + teams_activity.get('private_chat_messages', 0)),
            ("Calls", teams_activity.get('calls', 0)),
            ("Meetings", teams_activity.get('meetings', 0)),
        ]

        for label, value in activity_metrics:
            write_data_row(ws, row, [label, format_number(value)])
            row += 1
        row += 2

        # Calculate per-user metrics if we have active users
        active_users = teams_activity.get('active_users', 0)
        if active_users > 0:
            row = write_section_header(ws, row, "Per-User Activity (30-Day Average)")

            write_header_row(ws, row, ["Metric", "Avg per User"])
            row += 1

            total_msgs = teams_activity.get('team_chat_messages', 0) + teams_activity.get('private_chat_messages', 0)
            write_data_row(ws, row, ["Messages per User", f"{total_msgs / active_users:.1f}"])
            row += 1
            write_data_row(ws, row, ["Calls per User", f"{teams_activity.get('calls', 0) / active_users:.1f}"])
            row += 1
            write_data_row(ws, row, ["Meetings per User", f"{teams_activity.get('meetings', 0) / active_users:.1f}"])
            row += 1

    set_column_widths(ws, {'A': 35, 'B': 20, 'C': 15})


# =============================================================================
# REPORT GENERATION - SIZING INPUTS
# =============================================================================

def generate_sizing_inputs(wb: Workbook, resources: List[Dict], summary_data: Dict) -> None:
    """Generate Cohesity DataProtect M365 Sizing Inputs tab in sizer format."""
    ws = wb.create_sheet(title="Sizing Inputs")

    row = 1
    ws.cell(row=row, column=1, value="Cohesity DataProtect M365 Sizing Inputs").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Use these values for the Cohesity M365 sizing calculator").font = SUBTITLE_FONT
    row += 3

    # Headers matching Cohesity sizer format
    headers = [
        "Workload", "Type", "Count", "Item Count", "Item Size (GiB)",
        "Recoverable Item Count", "Recoverable Item Size (GiB)",
        "Total Item Count", "Total Item Size (GiB)", "Effective Size (GiB)",
        "Data Growth over last 180 days (GiB)", "Growth Rate over last 180 days (%)"
    ]

    write_header_row(ws, row, headers)
    row += 1

    # ASP overhead factor
    asp_factor = 1.1

    # Get exchange_summary from summary_data
    exchange_summary = summary_data.get('exchange_summary', {})
    sharepoint_summary = summary_data.get('sharepoint_summary', {})
    onedrive_summary = summary_data.get('onedrive_summary', {})
    growth_rates = summary_data.get('growth_rates') or {}
    teams_activity = summary_data.get('teams_activity') or {}

    # Exchange Online rows
    # If we have detailed exchange_summary from collector
    if exchange_summary and 'user_active' in exchange_summary:
        # User Active Mailboxes
        ua = exchange_summary.get('user_active', {})
        write_data_row(ws, row, [
            "", "User Active Mailboxes", ua.get('count', 0),
            ua.get('item_count', 0), round(ua.get('item_size_gib', 0), 3),
            ua.get('recoverable_item_count', 0), round(ua.get('recoverable_item_size_gib', 0), 3),
            ua.get('total_item_count', 0), round(ua.get('total_item_size_gib', 0), 3),
            round(ua.get('total_item_size_gib', 0) * asp_factor, 3),
            "", ""
        ])
        row += 1

        # User Archive Mailboxes
        uar = exchange_summary.get('user_archive_enabled', {})
        write_data_row(ws, row, [
            "", "User Archive Mailboxes", uar.get('count', 0),
            "", "", "", "", "", "", "", "", ""
        ])
        row += 1

        # SoftDeleted Active Mailboxes
        sda = exchange_summary.get('softdeleted_active', {})
        write_data_row(ws, row, [
            "", "SoftDeleted Active Mailboxes", sda.get('count', 0),
            sda.get('item_count', 0), round(sda.get('item_size_gib', 0), 3),
            sda.get('recoverable_item_count', 0), round(sda.get('recoverable_item_size_gib', 0), 3),
            sda.get('total_item_count', 0), round(sda.get('total_item_size_gib', 0), 3),
            "", "", ""
        ])
        row += 1

        # SoftDeleted Archive Mailboxes
        sdar = exchange_summary.get('softdeleted_archive', {})
        write_data_row(ws, row, [
            "", "SoftDeleted Archive Mailboxes", sdar.get('count', 0),
            "", "", "", "", "", "", "", "", ""
        ])
        row += 1

        # Group Active Mailboxes
        ga = exchange_summary.get('group_active', {})
        write_data_row(ws, row, [
            "", "Group Active Mailboxes", ga.get('count', 0),
            ga.get('item_count', 0), round(ga.get('item_size_gib', 0), 3),
            ga.get('recoverable_item_count', 0), round(ga.get('recoverable_item_size_gib', 0), 3),
            ga.get('total_item_count', 0), round(ga.get('total_item_size_gib', 0), 3),
            round(ga.get('total_item_size_gib', 0) * asp_factor, 3),
            "", ""
        ])
        row += 1

        # Group Archive Mailboxes
        gar = exchange_summary.get('group_archive', {})
        write_data_row(ws, row, [
            "", "Group Archive Mailboxes", gar.get('count', 0),
            "", "", "", "", "", "", "", "", ""
        ])
        row += 1

        # PublicFolder Active Mailboxes
        pfa = exchange_summary.get('publicfolder_active', {})
        write_data_row(ws, row, [
            "", "PublicFolder Active Mailboxes", pfa.get('count', 0),
            pfa.get('item_count', 0), round(pfa.get('item_size_gib', 0), 3),
            "", "",
            pfa.get('item_count', 0), round(pfa.get('item_size_gib', 0), 3),
            round(pfa.get('item_size_gib', 0) * asp_factor, 3),
            "", ""
        ])
        row += 1

        # Totals with Default Options Only
        td = exchange_summary.get('totals_default', {})
        ex_growth = growth_rates.get('exchange', {})
        ws.cell(row=row, column=1, value="Exchange Online").font = Font(bold=True)
        write_data_row(ws, row, [
            "Exchange Online", "Total Mailboxes with Default Options Only", td.get('count', 0),
            td.get('item_count', 0), round(td.get('item_size_gib', 0), 3),
            td.get('recoverable_item_count', 0), round(td.get('recoverable_item_size_gib', 0), 3),
            td.get('total_item_count', 0), round(td.get('total_item_size_gib', 0), 3),
            round(td.get('effective_size_asp_gib', 0), 3),
            round(ex_growth.get('growth_gib', 0), 2), ex_growth.get('growth_rate_percent', 0)
        ])
        for col in range(1, 13):
            ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1

        # Totals with All Options
        ta = exchange_summary.get('totals_all', {})
        ws.cell(row=row, column=1, value="Exchange Online").font = Font(bold=True)
        write_data_row(ws, row, [
            "Exchange Online", "Total Mailboxes with All Option Enabled", ta.get('count', 0),
            ta.get('item_count', 0), round(ta.get('item_size_gib', 0), 3),
            ta.get('recoverable_item_count', 0), round(ta.get('recoverable_item_size_gib', 0), 3),
            ta.get('total_item_count', 0), round(ta.get('total_item_size_gib', 0), 3),
            round(ta.get('effective_size_asp_gib', 0), 3),
            round(ex_growth.get('growth_gib', 0), 2), ex_growth.get('growth_rate_percent', 0)
        ])
        for col in range(1, 13):
            ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1
    else:
        # Fallback: calculate from resources if no detailed summary
        exchange_resources = [r for r in resources if 'mailbox' in r.get('resource_type', '')]
        total_exchange_gb = sum(r.get('size_gb', 0) or 0 for r in exchange_resources)
        total_items = sum(r.get('metadata', {}).get('item_count', 0) or 0 for r in exchange_resources)
        total_recoverable = sum(r.get('metadata', {}).get('deleted_item_count', 0) or 0 for r in exchange_resources)
        total_recoverable_gb = sum(r.get('metadata', {}).get('deleted_item_size_gb', 0) or 0 for r in exchange_resources)

        write_data_row(ws, row, [
            "Exchange Online", "Total Mailboxes", len(exchange_resources),
            total_items, round(total_exchange_gb, 3),
            total_recoverable, round(total_recoverable_gb, 3),
            total_items + total_recoverable, round(total_exchange_gb + total_recoverable_gb, 3),
            round((total_exchange_gb + total_recoverable_gb) * asp_factor, 3),
            "", ""
        ])
        for col in range(1, 13):
            ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1

    row += 1  # Blank row

    # SharePoint Online rows
    sp_growth = growth_rates.get('sharepoint', {})
    if sharepoint_summary and 'sharepoint_sites' in sharepoint_summary:
        sps = sharepoint_summary.get('sharepoint_sites', {})
        write_data_row(ws, row, [
            "", "SharePoint Sites", sps.get('count', 0),
            "", "", "", "", "", round(sps.get('storage_gib', 0), 3),
            "", "", ""
        ])
        row += 1

        ts = sharepoint_summary.get('team_sites', {})
        write_data_row(ws, row, [
            "", "Team Sites", ts.get('count', 0),
            "", "", "", "", "", round(ts.get('storage_gib', 0), 3),
            "", "", ""
        ])
        row += 1

        spt = sharepoint_summary.get('total', {})
        write_data_row(ws, row, [
            "SharePoint Online", "Total Sites", spt.get('count', 0),
            "", "", "", "", "", round(spt.get('storage_gib', 0), 3),
            round(spt.get('effective_size_asp_gib', spt.get('storage_gib', 0) * asp_factor), 3),
            round(sp_growth.get('growth_gib', 0), 2), sp_growth.get('growth_rate_percent', 0)
        ])
        for col in range(1, 13):
            ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1
    else:
        # Fallback: try resources first, then estimate from change rates
        sp_resources = [r for r in resources if 'sharepoint' in r.get('resource_type', '')]
        total_sp_gb = sum(r.get('size_gb', 0) or 0 for r in sp_resources)
        sp_count = len(sp_resources)
        sp_note = ""
        growth_180d_gb = ""
        growth_180d_pct = ""
        
        # If no resources but we have change rate data, estimate total size
        change_rates = summary_data.get('change_rates', {})
        sp_change = change_rates.get('SharePoint', {})
        if total_sp_gb == 0 and sp_change:
            daily_change_gb = sp_change.get('daily_change_gb', 0)
            daily_change_pct = sp_change.get('daily_change_percent', 0)
            if daily_change_gb > 0 and daily_change_pct > 0:
                # Estimate: total = daily_change_gb / (daily_change_percent / 100)
                total_sp_gb = daily_change_gb / (daily_change_pct / 100)
                sp_note = " (estimated from change rate)"
                sp_count = sp_change.get('resource_count', 0) or "Unknown"
                # Calculate 180-day growth from daily change
                growth_180d_gb = round(daily_change_gb * 180, 2)
                growth_180d_pct = round(daily_change_pct * 180, 2)
        
        # Use sp_growth if available, else use calculated values
        if sp_growth:
            growth_180d_gb = round(sp_growth.get('growth_gib', 0), 2) or growth_180d_gb
            growth_180d_pct = sp_growth.get('growth_rate_percent', 0) or growth_180d_pct
        
        write_data_row(ws, row, [
            "SharePoint Online", f"Total Sites{sp_note}", sp_count,
            "", "", "", "", "", round(total_sp_gb, 3),
            round(total_sp_gb * asp_factor, 3),
            growth_180d_gb, growth_180d_pct
        ])
        for col in range(1, 13):
            ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1

    row += 1  # Blank row

    # OneDrive for Business
    od_growth = growth_rates.get('onedrive', {})
    if onedrive_summary and 'personal_sites' in onedrive_summary:
        ps = onedrive_summary.get('personal_sites', {})
        write_data_row(ws, row, [
            "OneDrive for Business", "Personal Sites", ps.get('count', 0),
            "", "", "", "", "", round(ps.get('storage_gib', 0), 3),
            round(ps.get('effective_size_asp_gib', ps.get('storage_gib', 0) * asp_factor), 3),
            round(od_growth.get('growth_gib', 0), 2), od_growth.get('growth_rate_percent', 0)
        ])
        for col in range(1, 13):
            ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1
    else:
        # Fallback
        od_resources = [r for r in resources if 'onedrive' in r.get('resource_type', '')]
        total_od_gb = sum(r.get('size_gb', 0) or 0 for r in od_resources)
        write_data_row(ws, row, [
            "OneDrive for Business", "Personal Sites", len(od_resources),
            "", "", "", "", "", round(total_od_gb, 3),
            round(total_od_gb * asp_factor, 3),
            "", ""
        ])
        for col in range(1, 13):
            ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1

    row += 1  # Blank row

    # Teams Chat (metered units)
    if teams_activity:
        user_chats = teams_activity.get('estimated_metered_units_user_chats',
                                        teams_activity.get('private_chat_messages', 0))
        channel_chats = teams_activity.get('estimated_metered_units_channel_conversations',
                                           teams_activity.get('team_chat_messages', 0))
        total_metered = teams_activity.get('total_estimated_metered_units',
                                           user_chats + channel_chats)
        projected_annual = teams_activity.get('projected_annual_metered_units',
                                              total_metered * 12)

        write_data_row(ws, row, [
            "", "Estimated Metered Units for User Chats", user_chats,
            "", "", "", "", "", "", "", "", ""
        ])
        row += 1

        write_data_row(ws, row, [
            "", "Estimated Metered Units for Teams Channel Conversations", channel_chats,
            "", "", "", "", "", "", "", "", ""
        ])
        row += 1

        write_data_row(ws, row, [
            "", "Total Estimated Metered Units (Last 180 Days)", total_metered,
            "", "", "", "", "", "", "", "", ""
        ])
        row += 1

        write_data_row(ws, row, [
            "Teams Chat", "Total Estimated Metered Units (Last 180 Days + Next 1 Year)", projected_annual,
            "", "", "", "", "", "", "", "", ""
        ])
        for col in range(1, 13):
            ws.cell(row=row, column=col).font = Font(bold=True)
        row += 1

    row += 1  # Blank row

    # M365 Tenant Totals
    licensed_users = summary_data.get('total_user_count', 0)
    total_capacity_gb = summary_data.get('total_capacity_gb', 0)

    # Calculate from summaries if available
    total_size_gib = 0
    if exchange_summary and 'totals_all' in exchange_summary:
        total_size_gib += exchange_summary['totals_all'].get('total_item_size_gib', 0)
    if sharepoint_summary and 'total' in sharepoint_summary:
        total_size_gib += sharepoint_summary['total'].get('storage_gib', 0)
    if onedrive_summary and 'personal_sites' in onedrive_summary:
        total_size_gib += onedrive_summary['personal_sites'].get('storage_gib', 0)

    if total_size_gib == 0:
        total_size_gib = total_capacity_gb

    total_asp_gib = total_size_gib * asp_factor

    write_data_row(ws, row, [
        "M365 Tenant", "Licensed Users", licensed_users,
        "", "", "", "", "", "", "", "", ""
    ])
    row += 1

    write_data_row(ws, row, [
        "M365 Tenant", "Total Size", "",
        "", "", "", "", "", round(total_size_gib, 3),
        round(total_asp_gib, 3), "", ""
    ])
    for col in range(1, 13):
        ws.cell(row=row, column=col).font = Font(bold=True)
    row += 3

    # === Notes ===
    row = write_section_header(ws, row, "Notes")

    notes = [
        "• Item Count = number of emails/items in mailbox",
        "• Recoverable Items = items in the Recoverable Items folder (deleted items that can still be recovered)",
        "• ASP (Application Specific Protection) = Total size + ~10% overhead for metadata",
        "• Growth rates are calculated from 180-day storage history when available",
        "• Teams Chat metered units are based on message counts from Teams User Activity report",
        "• Archive mailbox sizes require Exchange PowerShell; only archive-enabled counts shown here",
    ]

    for note in notes:
        ws.cell(row=row, column=1, value=note)
        row += 1

    set_column_widths(ws, {
        'A': 18, 'B': 50, 'C': 12, 'D': 14, 'E': 16,
        'F': 20, 'G': 22, 'H': 16, 'I': 18, 'J': 22,
        'K': 30, 'L': 28
    })


# =============================================================================
# REPORT GENERATION - GROWTH ANALYSIS
# =============================================================================

def generate_growth_analysis(wb: Workbook, resources: List[Dict], summary_data: Dict) -> None:
    """Generate Growth Analysis tab."""
    ws = wb.create_sheet(title="Growth Analysis")

    row = 1
    ws.cell(row=row, column=1, value="Storage Growth Analysis").font = TITLE_FONT
    row += 3

    change_rates = summary_data.get('change_rates', {})

    if not change_rates:
        row = write_section_header(ws, row, "Growth Data")
        ws.cell(row=row, column=1, value="No historical growth data available")
        ws.cell(row=row + 1, column=1, value="Run the M365 collector with usage reports enabled to collect growth metrics")
        ws.cell(row=row + 3, column=1, value="Estimated Change Rates (Industry Typical):").font = SECTION_FONT
        row += 5

        estimates = [
            ("Exchange Online", "2-3%", "Email tends to have consistent daily growth"),
            ("OneDrive for Business", "3-5%", "User files more variable"),
            ("SharePoint Online", "2-4%", "Document libraries vary by organization"),
        ]

        write_header_row(ws, row, ["Service", "Daily Change Rate", "Notes"])
        row += 1

        for service, rate, note in estimates:
            write_data_row(ws, row, [service, rate, note])
            row += 1

        set_column_widths(ws, {'A': 30, 'B': 20, 'C': 50})
        return

    # === Actual Change Rates ===
    row = write_section_header(ws, row, "Measured Change Rates",
                                "(Based on historical usage data)")

    write_header_row(ws, row, ["Service", "Daily Change (GB)", "Daily Change (%)", "Annual Growth (%)", "Sample Period"])
    row += 1

    for service, data in change_rates.items():
        write_data_row(ws, row, [
            service,
            f"{data.get('daily_change_gb', 0):.2f}",
            f"{data.get('daily_change_percent', 0):.2f}%",
            f"{data.get('annual_growth_percent', 0):.1f}%",
            f"{data.get('sample_period_days', 0)} days"
        ])
        row += 1

    set_column_widths(ws, {'A': 30, 'B': 20, 'C': 20, 'D': 20, 'E': 15})


# =============================================================================
# REPORT GENERATION - RAW DATA
# =============================================================================

def generate_raw_data(wb: Workbook, resources: List[Dict]) -> None:
    """Generate Raw Data tab with full resource inventory."""
    ws = wb.create_sheet(title="Raw Data")

    row = 1
    ws.cell(row=row, column=1, value="Full Resource Inventory").font = TITLE_FONT
    row += 2

    if not resources:
        ws.cell(row=row, column=1, value="No resources found")
        return

    # Headers
    write_header_row(ws, row, [
        "Resource Type", "Service", "Name", "Resource ID", "Size (GB)",
        "Type/Status", "Created", "Last Activity"
    ])
    row += 1

    for r in resources:
        metadata = r.get('metadata', {}) or {}

        # Determine type/status based on resource
        rtype = r.get('resource_type', '')
        if 'mailbox' in rtype:
            type_status = metadata.get('mailbox_type', '')
        elif 'sharepoint' in rtype:
            type_status = metadata.get('site_type', '')
        elif 'teams' in rtype:
            type_status = 'Archived' if metadata.get('is_archived') else 'Active'
        else:
            type_status = ''

        write_data_row(ws, row, [
            r.get('resource_type', ''),
            r.get('service_family', ''),
            r.get('name', '')[:40],  # Truncate
            r.get('resource_id', '')[:40],  # Truncate
            round(r.get('size_gb', 0) or 0, 2),
            type_status,
            metadata.get('created_date', metadata.get('created_datetime', ''))[:10] if metadata.get('created_date') or metadata.get('created_datetime') else '',
            metadata.get('last_activity_date', '')[:10] if metadata.get('last_activity_date') else '',
        ])
        row += 1

    # Freeze header row
    ws.freeze_panes = 'A4'

    set_column_widths(ws, {
        'A': 25, 'B': 15, 'C': 35, 'D': 40, 'E': 12,
        'F': 15, 'G': 12, 'H': 15
    })


# =============================================================================
# MAIN REPORT GENERATION
# =============================================================================

def generate_report(inventory_files: List[str], output_path: str) -> None:
    """
    Generate M365 assessment report.

    Args:
        inventory_files: List of paths to M365 inventory JSON files
        output_path: Output Excel file path
    """
    print("Loading M365 inventory files...")
    resources, metadata = load_inventory_files(inventory_files)

    if not resources:
        print("Error: No M365 resources found in inventory files")
        sys.exit(1)

    print(f"  Loaded {len(resources)} resources")

    print("Loading summary data...")
    summary_data = load_summary_data(inventory_files)

    if summary_data.get('teams_activity') or summary_data.get('exchange_mailbox_breakdown'):
        print("  Found enhanced M365 summary data")

    print("Generating M365 report...")

    # Create workbook
    wb = Workbook()

    # Generate each tab
    generate_executive_summary(wb, resources, metadata, summary_data)
    generate_exchange_tab(wb, resources, summary_data)
    generate_sharepoint_tab(wb, resources, summary_data)
    generate_onedrive_tab(wb, resources)
    generate_teams_tab(wb, resources, summary_data)
    generate_sizing_inputs(wb, resources, summary_data)
    generate_growth_analysis(wb, resources, summary_data)
    generate_raw_data(wb, resources)

    # Save workbook
    wb.save(output_path)

    # Print summary
    print(f"\nReport generated: {output_path}")
    print("=" * 60)

    # Quick stats
    mailboxes = len([r for r in resources if 'mailbox' in r.get('resource_type', '')])
    sites = len([r for r in resources if 'sharepoint' in r.get('resource_type', '')])
    drives = len([r for r in resources if 'onedrive' in r.get('resource_type', '')])
    teams = len([r for r in resources if 'teams' in r.get('resource_type', '')])

    total_size = sum(r.get('size_gb', 0) or 0 for r in resources)

    print(f"Tenant: {metadata.get('tenant_name', metadata.get('tenant_id', 'Unknown'))}")
    print(f"Total Users: {format_number(summary_data.get('total_user_count', 0))}")
    print(f"Total Storage: {total_size / 1024:.2f} TB")
    print("\nWorkload Summary:")
    print(f"  Exchange Mailboxes: {format_number(mailboxes)}")
    print(f"  SharePoint Sites: {format_number(sites)}")
    print(f"  OneDrive Accounts: {format_number(drives)}")
    print(f"  Microsoft Teams: {format_number(teams)}")


def find_m365_files(directory: str) -> List[str]:
    """Find M365 inventory files in a directory."""
    patterns = [
        '**/cca_inv_*.json',
        '**/cca_m365_inv_*.json',
    ]

    inventory_files = []
    for pattern in patterns:
        matches = glob.glob(os.path.join(directory, pattern), recursive=True)
        for match in matches:
            # Check if it's actually M365 data
            data = load_json_file(match)
            if data:
                # Handle both list format (array of resources) and dict format
                if isinstance(data, list) and len(data) > 0:
                    # Check first resource for provider
                    if data[0].get('provider') == 'microsoft365':
                        inventory_files.append(match)
                elif isinstance(data, dict) and data.get('provider') == 'microsoft365':
                    inventory_files.append(match)

    return list(set(inventory_files))


def main():
    parser = argparse.ArgumentParser(
        description='Generate Microsoft 365 Assessment Report',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Auto-discover M365 files in current directory
  python scripts/generate_m365_report.py

  # Specify directory containing M365 output
  python scripts/generate_m365_report.py --directory ./m365_output

  # Specify inventory files directly
  python scripts/generate_m365_report.py --inventory cca_m365_inv_*.json

  # Custom output filename
  python scripts/generate_m365_report.py -o my_m365_report.xlsx
        """
    )

    parser.add_argument('--directory', '-d', type=str, default='.',
                        help='Directory containing M365 inventory files (default: current)')
    parser.add_argument('--inventory', '-i', nargs='+',
                        help='M365 inventory file(s) to process')
    parser.add_argument('--output', '-o', type=str, default='m365_assessment_report.xlsx',
                        help='Output Excel file path')

    args = parser.parse_args()

    # Find inventory files
    if args.inventory:
        inventory_files = []
        for pattern in args.inventory:
            inventory_files.extend(glob.glob(pattern))
    else:
        inventory_files = find_m365_files(args.directory)

    if not inventory_files:
        print("Error: No M365 inventory files found")
        print(f"\nSearched in: {os.path.abspath(args.directory)}")
        print("\nRun the M365 collector first:")
        print("  python m365_collect.py")
        sys.exit(1)

    print(f"Found {len(inventory_files)} M365 inventory file(s)")
    for f in inventory_files:
        print(f"  - {f}")

    generate_report(inventory_files, args.output)


if __name__ == "__main__":
    main()
