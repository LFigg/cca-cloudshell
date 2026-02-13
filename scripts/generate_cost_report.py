#!/usr/bin/env python3
"""
CCA CloudShell - Cost Report Generator

Generates an Excel cost report from cost collector output with:
- Executive Summary with KPIs and trends
- Detailed cost breakdown by provider, category, service
- Cost optimization recommendations
- Monthly trend analysis

Usage:
    python3 scripts/generate_cost_report.py --inventory cca_cost_inv.json --summary cca_cost_sum.json
    python3 scripts/generate_cost_report.py --inventory cca_cost_inv.json --summary cca_cost_sum.json --output report.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from openpyxl import Workbook  # type: ignore[import-not-found]

try:
    from openpyxl import Workbook as _Workbook  # type: ignore[import-not-found]
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # type: ignore[import-not-found]
    from openpyxl.utils import get_column_letter  # type: ignore[import-not-found]
    from openpyxl.chart import LineChart, Reference  # type: ignore[import-not-found]
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    _Workbook: Any = None
    Font: Any = None
    PatternFill: Any = None
    Alignment: Any = None
    Border: Any = None
    Side: Any = None
    get_column_letter: Any = None
    LineChart: Any = None
    Reference: Any = None


# =============================================================================
# Data Loading
# =============================================================================

def load_json(filepath: str) -> dict[str, Any]:
    """Load JSON file."""
    with open(filepath, 'r') as f:
        return json.load(f)


# =============================================================================
# Analysis Functions
# =============================================================================

def analyze_costs(inventory: dict[str, Any], summary: dict[str, Any]) -> dict[str, Any]:
    """
    Analyze cost data and generate insights.
    
    Returns dict with:
    - totals: Overall totals
    - by_provider: Costs grouped by provider
    - by_category: Costs grouped by category
    - by_service: Costs grouped by service
    - monthly_trend: Monthly cost trend (if multi-month data)
    - optimization: Cost optimization recommendations
    """
    records = inventory.get('records', [])
    
    analysis: dict[str, Any] = {
        'metadata': {
            'run_id': inventory.get('run_id', 'unknown'),
            'timestamp': inventory.get('timestamp', ''),
            'period_start': inventory.get('period', {}).get('start', ''),
            'period_end': inventory.get('period', {}).get('end', ''),
            'providers': inventory.get('providers', []),
        },
        'totals': {
            'total_cost': inventory.get('total_cost', 0),
            'total_records': inventory.get('total_records', 0),
        },
        'by_provider': defaultdict(lambda: {'cost': 0, 'services': defaultdict(float)}),
        'by_category': defaultdict(lambda: {'cost': 0, 'providers': defaultdict(float)}),
        'by_service': defaultdict(lambda: {'cost': 0, 'category': '', 'provider': ''}),
        'by_account': defaultdict(lambda: {'cost': 0, 'services': defaultdict(float), 'provider': ''}),
        'monthly_trend': defaultdict(lambda: defaultdict(float)),
        'records': records,
    }
    
    # Aggregate from records
    for record in records:
        provider = record.get('provider', 'unknown')
        category = record.get('category', 'unknown')
        service = record.get('service', 'unknown')
        cost = record.get('cost', 0)
        period = record.get('period_start', '')[:7]  # YYYY-MM
        
        # By provider
        analysis['by_provider'][provider]['cost'] += cost
        analysis['by_provider'][provider]['services'][service] += cost
        
        # By category
        analysis['by_category'][category]['cost'] += cost
        analysis['by_category'][category]['providers'][provider] += cost
        
        # By service
        analysis['by_service'][service]['cost'] += cost
        analysis['by_service'][service]['category'] = category
        analysis['by_service'][service]['provider'] = provider
        
        # By account (for Organizations/multi-account data)
        account_id = record.get('account_id', '')
        if account_id:
            analysis['by_account'][account_id]['cost'] += cost
            analysis['by_account'][account_id]['services'][service] += cost
            analysis['by_account'][account_id]['provider'] = provider
        
        # Monthly trend
        if period:
            analysis['monthly_trend'][period][provider] += cost
    
    # Convert defaultdicts to regular dicts
    analysis['by_provider'] = dict(analysis['by_provider'])
    analysis['by_category'] = dict(analysis['by_category'])
    analysis['by_service'] = dict(analysis['by_service'])
    analysis['by_account'] = dict(analysis['by_account'])
    analysis['monthly_trend'] = dict(analysis['monthly_trend'])
    
    # Generate optimization recommendations
    analysis['optimization'] = generate_optimization_recommendations(analysis)
    
    return analysis


def generate_optimization_recommendations(analysis: dict[str, Any]) -> list[dict[str, Any]]:
    """Generate cost optimization recommendations based on analysis."""
    recommendations: list[dict[str, Any]] = []
    
    total_cost = analysis['totals']['total_cost']
    by_category = analysis['by_category']
    by_service = analysis['by_service']
    
    # Check snapshot costs
    snapshot_cost = by_category.get('ec2_snapshot', {}).get('cost', 0)
    snapshot_cost += by_category.get('rds_snapshot', {}).get('cost', 0)
    if snapshot_cost > 0:
        recommendations.append({
            'title': 'Review Snapshot Retention',
            'category': 'Snapshots',
            'current_cost': round(snapshot_cost, 2),
            'potential_savings': round(snapshot_cost * 0.3, 2),
            'savings_percent': 30,
            'description': 'Many organizations retain snapshots longer than necessary. '
                          'Review snapshot retention policies and delete outdated snapshots.',
            'priority': 'High' if snapshot_cost > total_cost * 0.3 else 'Medium',
        })
    
    # Check AWS Backup vault storage
    vault_cost = 0.0
    for service, data in by_service.items():
        if 'vault' in service.lower() or 'backup' in service.lower():
            vault_cost += data['cost']
    
    if vault_cost > 0:
        recommendations.append({
            'title': 'Optimize Backup Vault Storage',
            'category': 'Backup Storage',
            'current_cost': round(vault_cost, 2),
            'potential_savings': round(vault_cost * 0.2, 2),
            'savings_percent': 20,
            'description': 'Consider using lifecycle policies to transition older backups '
                          'to cold storage tiers (e.g., Glacier, Archive).',
            'priority': 'Medium',
        })
    
    # Check for multi-cloud redundancy
    providers = list(analysis['by_provider'].keys())
    if len(providers) > 1:
        provider_costs = {p: analysis['by_provider'][p]['cost'] for p in providers}
        min_provider = min(provider_costs, key=lambda x: provider_costs[x])
        max_provider = max(provider_costs, key=lambda x: provider_costs[x])
        
        if provider_costs[max_provider] > provider_costs[min_provider] * 2:
            recommendations.append({
                'title': 'Balance Multi-Cloud Backup Costs',
                'category': 'Multi-Cloud',
                'current_cost': round(provider_costs[max_provider], 2),
                'potential_savings': round((provider_costs[max_provider] - provider_costs[min_provider]) * 0.25, 2),
                'savings_percent': 25,
                'description': f'{max_provider.upper()} backup costs are significantly higher than '
                              f'{min_provider.upper()}. Consider consolidating or rebalancing workloads.',
                'priority': 'Low',
            })
    
    # General recommendations if no specific ones apply
    if not recommendations and total_cost > 0:
        recommendations.append({
            'title': 'Enable Intelligent Tiering',
            'category': 'General',
            'current_cost': round(total_cost, 2),
            'potential_savings': round(total_cost * 0.15, 2),
            'savings_percent': 15,
            'description': 'Enable automatic storage tiering for backup data to move '
                          'infrequently accessed backups to lower-cost storage classes.',
            'priority': 'Medium',
        })
    
    # Sort by potential savings (descending)
    recommendations.sort(key=lambda x: x['potential_savings'], reverse=True)
    
    return recommendations


# =============================================================================
# Excel Generation
# =============================================================================

def create_executive_summary_sheet(wb: Any, analysis: dict[str, Any]) -> None:
    """Create Executive Summary sheet with KPIs and overview."""
    # Styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    metric_font = Font(name="Calibri", size=24, bold=True, color="1F4E79")
    label_font = Font(name="Calibri", size=10, color="666666")
    
    ws = wb.active
    ws.title = "Executive Summary"
    
    # Title
    ws['A1'] = "Backup & Snapshot Cost Analysis"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    
    # Report metadata
    meta = analysis['metadata']
    ws['A3'] = "Report Period:"
    ws['B3'] = f"{meta['period_start']} to {meta['period_end']}"
    ws['A4'] = "Generated:"
    ws['B4'] = meta['timestamp']
    ws['A5'] = "Providers:"
    ws['B5'] = ", ".join(p.upper() for p in meta['providers'])
    
    for row in range(3, 6):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
    
    # KPI Cards section
    ws['A8'] = "KEY METRICS"
    ws['A8'].font = Font(size=14, bold=True, color="1F4E79")
    
    # Total Cost KPI
    ws['B10'] = "TOTAL COST"
    ws['B10'].font = label_font
    ws['B10'].alignment = Alignment(horizontal='center')
    ws['B11'] = f"${analysis['totals']['total_cost']:,.2f}"
    ws['B11'].font = metric_font
    ws['B11'].alignment = Alignment(horizontal='center')
    
    # Total Records KPI
    ws['D10'] = "RECORDS ANALYZED"
    ws['D10'].font = label_font
    ws['D10'].alignment = Alignment(horizontal='center')
    ws['D11'] = analysis['totals']['total_records']
    ws['D11'].font = metric_font
    ws['D11'].alignment = Alignment(horizontal='center')
    
    # Provider Count KPI
    ws['F10'] = "CLOUD PROVIDERS"
    ws['F10'].font = label_font
    ws['F10'].alignment = Alignment(horizontal='center')
    ws['F11'] = len(meta['providers'])
    ws['F11'].font = metric_font
    ws['F11'].alignment = Alignment(horizontal='center')
    
    # Cost by Provider summary table
    ws['A14'] = "COST BY PROVIDER"
    ws['A14'].font = Font(size=12, bold=True, color="1F4E79")
    
    headers = ['Provider', 'Cost', '% of Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=15, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 16
    total = analysis['totals']['total_cost']
    for provider, data in sorted(analysis['by_provider'].items()):
        cost = data['cost']
        pct = (cost / total * 100) if total > 0 else 0
        
        ws.cell(row=row, column=1, value=provider.upper()).border = thin_border
        cost_cell = ws.cell(row=row, column=2, value=cost)
        cost_cell.number_format = '$#,##0.00'
        cost_cell.border = thin_border
        pct_cell = ws.cell(row=row, column=3, value=pct/100)
        pct_cell.number_format = '0.0%'
        pct_cell.border = thin_border
        row += 1
    
    # Cost by Category summary table
    ws['E14'] = "COST BY CATEGORY"
    ws['E14'].font = Font(size=12, bold=True, color="1F4E79")
    
    for col, header in enumerate(['Category', 'Cost', '% of Total'], 5):
        cell = ws.cell(row=15, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 16
    for category, data in sorted(analysis['by_category'].items()):
        cost = data['cost']
        pct = (cost / total * 100) if total > 0 else 0
        
        ws.cell(row=row, column=5, value=category.replace('_', ' ').title()).border = thin_border
        cost_cell = ws.cell(row=row, column=6, value=cost)
        cost_cell.number_format = '$#,##0.00'
        cost_cell.border = thin_border
        pct_cell = ws.cell(row=row, column=7, value=pct/100)
        pct_cell.number_format = '0.0%'
        pct_cell.border = thin_border
        row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12


def create_provider_detail_sheet(wb: Any, analysis: dict[str, Any]) -> None:
    """Create detailed cost breakdown by provider."""
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    ws = wb.create_sheet("Cost by Provider")
    
    ws['A1'] = "Cost Analysis by Cloud Provider"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    row = 3
    for provider, data in sorted(analysis['by_provider'].items()):
        ws.cell(row=row, column=1, value=f"{provider.upper()} - ${data['cost']:,.2f}")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True, color="2E75B6")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        headers = ['Service', 'Cost', '% of Provider']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = subheader_fill
            cell.border = thin_border
        row += 1
        
        provider_total = data['cost']
        for service, cost in sorted(data['services'].items(), key=lambda x: x[1], reverse=True):
            pct = (cost / provider_total * 100) if provider_total > 0 else 0
            
            ws.cell(row=row, column=1, value=service).border = thin_border
            cost_cell = ws.cell(row=row, column=2, value=cost)
            cost_cell.number_format = '$#,##0.00'
            cost_cell.border = thin_border
            pct_cell = ws.cell(row=row, column=3, value=pct/100)
            pct_cell.number_format = '0.0%'
            pct_cell.border = thin_border
            row += 1
        
        row += 2
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def create_category_detail_sheet(wb: Any, analysis: dict[str, Any]) -> None:
    """Create detailed cost breakdown by category."""
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    ws = wb.create_sheet("Cost by Category")
    
    ws['A1'] = "Cost Analysis by Backup Category"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    row = 3
    for category, data in sorted(analysis['by_category'].items()):
        display_name = category.replace('_', ' ').title()
        ws.cell(row=row, column=1, value=f"{display_name} - ${data['cost']:,.2f}")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True, color="2E75B6")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        headers = ['Provider', 'Cost', '% of Category']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = subheader_fill
            cell.border = thin_border
        row += 1
        
        category_total = data['cost']
        for provider, cost in sorted(data['providers'].items(), key=lambda x: x[1], reverse=True):
            pct = (cost / category_total * 100) if category_total > 0 else 0
            
            ws.cell(row=row, column=1, value=provider.upper()).border = thin_border
            cost_cell = ws.cell(row=row, column=2, value=cost)
            cost_cell.number_format = '$#,##0.00'
            cost_cell.border = thin_border
            pct_cell = ws.cell(row=row, column=3, value=pct/100)
            pct_cell.number_format = '0.0%'
            pct_cell.border = thin_border
            row += 1
        
        row += 2
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


def create_service_detail_sheet(wb: Any, analysis: dict[str, Any]) -> None:
    """Create detailed cost breakdown by service."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    ws = wb.create_sheet("Cost by Service")
    
    ws['A1'] = "Detailed Service Cost Breakdown"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    headers = ['Service', 'Provider', 'Category', 'Cost', '% of Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 4
    total = analysis['totals']['total_cost']
    services_sorted = sorted(analysis['by_service'].items(), key=lambda x: x[1]['cost'], reverse=True)
    
    for service, data in services_sorted:
        cost = data['cost']
        pct = (cost / total * 100) if total > 0 else 0
        
        ws.cell(row=row, column=1, value=service).border = thin_border
        ws.cell(row=row, column=2, value=data['provider'].upper()).border = thin_border
        ws.cell(row=row, column=3, value=data['category'].replace('_', ' ').title()).border = thin_border
        
        cost_cell = ws.cell(row=row, column=4, value=cost)
        cost_cell.number_format = '$#,##0.00'
        cost_cell.border = thin_border
        
        pct_cell = ws.cell(row=row, column=5, value=pct/100)
        pct_cell.number_format = '0.0%'
        pct_cell.border = thin_border
        
        if pct > 20:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = warning_fill
        
        row += 1
    
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12


def create_account_detail_sheet(wb: Any, analysis: dict[str, Any]) -> None:
    """Create detailed cost breakdown by account (for Organizations/multi-account)."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    ws = wb.create_sheet("Cost by Account")
    
    ws['A1'] = "Cost Analysis by Account"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    # Summary table at top - accounts sorted by cost descending
    ws['A3'] = "Account Summary (Top Accounts by Spend)"
    ws['A3'].font = Font(size=12, bold=True, color="1F4E79")
    
    headers = ['Account ID', 'Provider', 'Cost', '% of Total', 'Top Service']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    row = 5
    total = analysis['totals']['total_cost']
    accounts_sorted = sorted(analysis['by_account'].items(), key=lambda x: x[1]['cost'], reverse=True)
    
    # Show top 50 accounts in summary
    for account_id, data in accounts_sorted[:50]:
        cost = data['cost']
        pct = (cost / total * 100) if total > 0 else 0
        
        # Find top service for this account
        top_service = ''
        if data['services']:
            top_service = max(data['services'].items(), key=lambda x: x[1])[0]
        
        ws.cell(row=row, column=1, value=account_id).border = thin_border
        ws.cell(row=row, column=2, value=data['provider'].upper()).border = thin_border
        
        cost_cell = ws.cell(row=row, column=3, value=cost)
        cost_cell.number_format = '$#,##0.00'
        cost_cell.border = thin_border
        
        pct_cell = ws.cell(row=row, column=4, value=pct/100)
        pct_cell.number_format = '0.0%'
        pct_cell.border = thin_border
        
        ws.cell(row=row, column=5, value=top_service).border = thin_border
        
        # Highlight high-cost accounts (>10% of total)
        if pct > 10:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = warning_fill
        
        row += 1
    
    # If more than 50 accounts, show count
    if len(accounts_sorted) > 50:
        ws.cell(row=row + 1, column=1, value=f"... and {len(accounts_sorted) - 50} more accounts")
        ws.cell(row=row + 1, column=1).font = Font(italic=True, color="666666")
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 35


def create_monthly_trend_sheet(wb: Any, analysis: dict[str, Any]) -> None:
    """Create monthly trend analysis sheet."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    ws = wb.create_sheet("Monthly Trends")
    
    ws['A1'] = "Monthly Cost Trends"
    ws['A1'].font = title_font
    ws.merge_cells('A1:F1')
    
    monthly_data = analysis['monthly_trend']
    
    if not monthly_data:
        ws['A3'] = "No monthly trend data available (single month analysis)"
        return
    
    all_providers: set[str] = set()
    for month_data in monthly_data.values():
        all_providers.update(month_data.keys())
    all_providers_list = sorted(all_providers)
    months = sorted(monthly_data.keys())
    
    ws.cell(row=3, column=1, value="Month")
    ws.cell(row=3, column=1).font = header_font
    ws.cell(row=3, column=1).fill = header_fill
    ws.cell(row=3, column=1).border = thin_border
    
    for col, provider in enumerate(all_providers_list, 2):
        cell = ws.cell(row=3, column=col, value=provider.upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    total_col = len(all_providers_list) + 2
    ws.cell(row=3, column=total_col, value="Total")
    ws.cell(row=3, column=total_col).font = header_font
    ws.cell(row=3, column=total_col).fill = header_fill
    ws.cell(row=3, column=total_col).border = thin_border
    
    row = 4
    for month in months:
        ws.cell(row=row, column=1, value=month).border = thin_border
        
        month_total = 0.0
        for col, provider in enumerate(all_providers_list, 2):
            cost = monthly_data[month].get(provider, 0)
            month_total += cost
            cost_cell = ws.cell(row=row, column=col, value=cost)
            cost_cell.number_format = '$#,##0.00'
            cost_cell.border = thin_border
        
        total_cell = ws.cell(row=row, column=total_col, value=month_total)
        total_cell.number_format = '$#,##0.00'
        total_cell.font = Font(bold=True)
        total_cell.border = thin_border
        row += 1
    
    if len(months) > 1:
        chart = LineChart()
        chart.title = "Monthly Cost Trend"
        chart.style = 10
        chart.x_axis.title = "Month"
        chart.y_axis.title = "Cost ($)"
        chart.y_axis.numFmt = '$#,##0'
        
        data = Reference(ws, min_col=2, min_row=3, max_col=total_col, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=4, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 12
        chart.width = 18
        
        ws.add_chart(chart, f"A{row + 2}")
    
    ws.column_dimensions['A'].width = 12
    for col in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14


def create_optimization_sheet(wb: Any, analysis: dict[str, Any]) -> None:
    """Create cost optimization recommendations sheet."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    danger_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    ws = wb.create_sheet("Cost Optimization")
    
    ws['A1'] = "Cost Optimization Recommendations"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    
    recommendations = analysis['optimization']
    total_savings = sum(r['potential_savings'] for r in recommendations)
    
    ws['A3'] = "Total Potential Savings:"
    ws['A3'].font = Font(bold=True)
    ws['B3'] = total_savings
    ws['B3'].font = Font(size=16, bold=True, color="2E7D32")
    ws['B3'].number_format = '$#,##0.00'
    
    ws['D3'] = f"Recommendations: {len(recommendations)}"
    ws['D3'].font = Font(bold=True)
    
    headers = ['Priority', 'Recommendation', 'Category', 'Current Cost', 'Potential Savings', 'Est. Reduction', 'Description']
    row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    row = 7
    for rec in recommendations:
        priority_cell = ws.cell(row=row, column=1, value=rec['priority'])
        priority_cell.alignment = Alignment(horizontal='center')
        priority_cell.border = thin_border
        if rec['priority'] == 'High':
            priority_cell.fill = danger_fill
        elif rec['priority'] == 'Medium':
            priority_cell.fill = warning_fill
        else:
            priority_cell.fill = success_fill
        
        ws.cell(row=row, column=2, value=rec['title']).border = thin_border
        ws.cell(row=row, column=3, value=rec['category']).border = thin_border
        
        cost_cell = ws.cell(row=row, column=4, value=rec['current_cost'])
        cost_cell.number_format = '$#,##0.00'
        cost_cell.border = thin_border
        
        savings_cell = ws.cell(row=row, column=5, value=rec['potential_savings'])
        savings_cell.number_format = '$#,##0.00'
        savings_cell.font = Font(color="2E7D32", bold=True)
        savings_cell.border = thin_border
        
        pct_cell = ws.cell(row=row, column=6, value=rec['savings_percent']/100)
        pct_cell.number_format = '0%'
        pct_cell.border = thin_border
        
        desc_cell = ws.cell(row=row, column=7, value=rec['description'])
        desc_cell.alignment = Alignment(wrap_text=True)
        desc_cell.border = thin_border
        
        ws.row_dimensions[row].height = 45
        row += 1
    
    if not recommendations:
        ws.cell(row=row, column=1, value="No specific optimization recommendations identified.")
        ws.merge_cells(f'A{row}:G{row}')
    
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 50


def create_raw_data_sheet(wb: Any, analysis: dict[str, Any]) -> None:
    """Create raw data sheet with all records."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    title_font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    ws = wb.create_sheet("Raw Data")
    
    ws['A1'] = "Raw Cost Records"
    ws['A1'].font = title_font
    ws.merge_cells('A1:J1')
    
    records = analysis['records']
    if not records:
        ws['A3'] = "No records available"
        return
    
    all_keys: set[str] = set()
    for record in records:
        all_keys.update(record.keys())
    
    ordered_keys = ['provider', 'account_id', 'service', 'category', 'cost', 'currency',
                    'period_start', 'period_end', 'usage_quantity', 'usage_unit']
    for key in sorted(all_keys):
        if key not in ordered_keys and key != 'metadata':
            ordered_keys.append(key)
    
    for col, key in enumerate(ordered_keys, 1):
        cell = ws.cell(row=3, column=col, value=key.replace('_', ' ').title())
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 4
    for record in records:
        for col, key in enumerate(ordered_keys, 1):
            value = record.get(key, '')
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            
            if key == 'cost' and isinstance(value, (int, float)):
                cell.number_format = '$#,##0.00'
        row += 1
    
    ws.auto_filter.ref = f"A3:{get_column_letter(len(ordered_keys))}{row-1}"
    
    col_widths = {
        'provider': 10, 'account_id': 15, 'service': 35, 'category': 18,
        'cost': 12, 'currency': 10, 'period_start': 12, 'period_end': 12,
        'usage_quantity': 15, 'usage_unit': 12
    }
    for col, key in enumerate(ordered_keys, 1):
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(key, 15)


def generate_excel_report(inventory_path: str, summary_path: str, output_path: str) -> None:
    """
    Generate Excel cost report from inventory and summary JSON files.
    
    Args:
        inventory_path: Path to cca_cost_inv_*.json
        summary_path: Path to cca_cost_sum_*.json
        output_path: Output Excel file path
    """
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)
    
    print(f"Loading inventory: {inventory_path}")
    inventory = load_json(inventory_path)
    
    print(f"Loading summary: {summary_path}")
    summary = load_json(summary_path)
    
    print("Analyzing cost data...")
    analysis = analyze_costs(inventory, summary)
    
    print("Generating Excel report...")
    wb = _Workbook()
    
    create_executive_summary_sheet(wb, analysis)
    create_provider_detail_sheet(wb, analysis)
    create_category_detail_sheet(wb, analysis)
    create_service_detail_sheet(wb, analysis)
    
    # Only add account sheet if multi-account data exists
    if len(analysis['by_account']) > 1:
        create_account_detail_sheet(wb, analysis)
    
    create_monthly_trend_sheet(wb, analysis)
    create_optimization_sheet(wb, analysis)
    create_raw_data_sheet(wb, analysis)
    
    wb.save(output_path)
    print(f"\nReport saved: {output_path}")
    print(f"  - Total Cost: ${analysis['totals']['total_cost']:,.2f}")
    print(f"  - Records: {analysis['totals']['total_records']}")
    print(f"  - Optimization Recommendations: {len(analysis['optimization'])}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description='Generate Excel cost report from CCA cost collector output',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 scripts/generate_cost_report.py --inventory output/cca_cost_inv_120000.json --summary output/cca_cost_sum_120000.json
  python3 scripts/generate_cost_report.py --inventory output/cca_cost_inv_120000.json --summary output/cca_cost_sum_120000.json --output my_report.xlsx
"""
    )
    
    parser.add_argument('--inventory', '-i', required=True,
                       help='Path to cost inventory JSON file (cca_cost_inv_*.json)')
    parser.add_argument('--summary', '-s', required=True,
                       help='Path to cost summary JSON file (cca_cost_sum_*.json)')
    parser.add_argument('--output', '-o', default='cost_report.xlsx',
                       help='Output Excel file path (default: cost_report.xlsx)')
    
    args = parser.parse_args()
    
    generate_excel_report(args.inventory, args.summary, args.output)


if __name__ == '__main__':
    main()
