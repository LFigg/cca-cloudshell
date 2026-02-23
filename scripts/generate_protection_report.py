#!/usr/bin/env python3
"""
Generate a Protection Report Excel file from inventory.json

Creates an Excel workbook with:
- Summary tab: High-level statistics, counts, and sizes
- Protection Report tab: Detailed Instance → Volume → Snapshot hierarchy
- Backup Plans tab: All backup plans with rule details
- Backup Selections tab: Resources assigned to backup plans (if any)
"""
import json
import os
import re
import sys
from collections import defaultdict
from typing import Dict, List, Any, Optional
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore[import-untyped]
from openpyxl.utils.dataframe import dataframe_to_rows  # type: ignore[import-untyped]


class InventoryValidationError(Exception):
    """Raised when inventory data fails validation."""
    pass


def validate_inventory(data: Any, filepath: str = "inventory") -> None:
    """Validate inventory data has required structure.
    
    Args:
        data: The loaded inventory data to validate
        filepath: Path to file for error messages
        
    Raises:
        InventoryValidationError: If data is invalid with helpful message
    """
    if data is None:
        raise InventoryValidationError(f"Inventory file is empty or null: {filepath}")
    
    if not isinstance(data, dict):
        raise InventoryValidationError(
            f"Invalid inventory format: expected JSON object, got {type(data).__name__}. "
            f"File: {filepath}"
        )
    
    if 'resources' not in data:
        available_keys = list(data.keys())[:5]  # Show first 5 keys
        raise InventoryValidationError(
            f"Missing 'resources' key in inventory. "
            f"Available keys: {available_keys}. File: {filepath}"
        )
    
    if not isinstance(data['resources'], list):
        raise InventoryValidationError(
            f"Invalid 'resources' field: expected list, got {type(data['resources']).__name__}. "
            f"File: {filepath}"
        )
    
    # Validate first resource has expected structure (if any resources exist)
    if data['resources']:
        first = data['resources'][0]
        if not isinstance(first, dict):
            raise InventoryValidationError(
                f"Invalid resource format: expected objects, got {type(first).__name__}. "
                f"File: {filepath}"
            )
        required_keys = {'resource_id', 'resource_type'}
        missing = required_keys - set(first.keys())
        if missing:
            raise InventoryValidationError(
                f"Invalid resource format: missing required keys {missing}. "
                f"File: {filepath}"
            )


def load_inventory(filepath: str) -> Dict[str, Any]:
    """Load inventory JSON file."""
    with open(filepath) as f:
        return json.load(f)


def stringify_field(value: Any, max_items: int = 10) -> str:
    """Convert a field value to a string suitable for Excel cells.
    
    Handles lists (from merged multi-account data) by joining them,
    with truncation for very long lists.
    """
    if value is None:
        return 'N/A'
    if isinstance(value, list):
        if len(value) == 0:
            return 'N/A'
        if len(value) == 1:
            return str(value[0])
        if len(value) <= max_items:
            return ', '.join(str(v) for v in value)
        return f"{len(value)} accounts (merged)"
    return str(value)


def build_resource_index(resources: List[Dict]) -> Dict[str, Dict]:
    """Build index of resources by resource_id for quick lookup."""
    return {r['resource_id']: r for r in resources}


def build_volume_index(resources: List[Dict]) -> Dict[str, Dict]:
    """Build index of EBS volumes by volume ID for cross-account lookup.
    
    This enables finding a volume's source even if the snapshot was copied
    to a different account. Volume IDs are globally unique within AWS.
    
    Returns:
        Dict mapping volume_id (e.g., 'vol-xxx') to the volume resource
    """
    index = {}
    for r in resources:
        if r['resource_type'] == 'aws:ec2:volume':
            vol_id = r['resource_id']
            # Handle both ARN and raw volume ID formats
            if vol_id.startswith('arn:'):
                # Extract vol-xxx from ARN
                parts = vol_id.split('/')
                if len(parts) > 1:
                    vol_id = parts[-1]
            index[vol_id] = r
    return index


def get_ec2_instances(resources: List[Dict]) -> List[Dict]:
    """Get EC2 instances."""
    return [r for r in resources if r['resource_type'] == 'aws:ec2:instance']


def get_ebs_volumes(resources: List[Dict]) -> List[Dict]:
    """Get EBS volumes."""
    return [r for r in resources if r['resource_type'] == 'aws:ec2:volume']


def get_rds_instances(resources: List[Dict]) -> List[Dict]:
    """Get RDS instances and clusters."""
    return [r for r in resources if r['resource_type'] in ['aws:rds:instance', 'aws:rds:cluster']]


def _format_rds_snapshot_description(snap: Dict) -> str:
    """Format RDS snapshot description with clear cost indication.
    
    RDS automated backups are INCLUDED with the service (up to retention period storage).
    Manual and AWS Backup snapshots are ADDITIONAL COST.
    """
    snapshot_type = snap.get('metadata', {}).get('snapshot_type', 'unknown')
    
    if snapshot_type == 'automated':
        return "Automated (Included with RDS)"
    elif snapshot_type == 'manual':
        return "Manual (Additional Cost)"
    elif snapshot_type == 'awsbackup':
        return "AWS Backup (Additional Cost)"
    else:
        return f"{snapshot_type} snapshot"


def get_other_primary_resources(resources: List[Dict]) -> List[Dict]:
    """Get other primary resources (S3, EFS, DynamoDB, etc.)."""
    other_types = [
        'aws:s3:bucket',
        'aws:efs:filesystem',
        'aws:dynamodb:table',
        'azure:vm',
        'azure:disk',
        'azure:sql:database',
        'azure:storage:account',
    ]
    return [r for r in resources if r['resource_type'] in other_types]


def get_snapshots(resources: List[Dict]) -> List[Dict]:
    """Get all snapshot resources."""
    snapshot_types = [
        'aws:ec2:snapshot',
        'aws:rds:snapshot',
        'aws:rds:cluster-snapshot',
        'azure:snapshot',
    ]
    return [r for r in resources if r['resource_type'] in snapshot_types]


def get_backup_plans(resources: List[Dict]) -> List[Dict]:
    """Get backup plans/policies."""
    plan_types = [
        'aws:backup:plan',
        'aws:backup:vault',
        'azure:backup:policy',
        'azure:recoveryservices:vault',
    ]
    return [r for r in resources if r['resource_type'] in plan_types]


def get_backup_selections(resources: List[Dict]) -> List[Dict]:
    """Get backup selections (resource-to-plan assignments)."""
    return [r for r in resources if r['resource_type'] == 'aws:backup:selection']


def get_protected_resources(resources: List[Dict]) -> List[Dict]:
    """Get resources that have actual recovery points in AWS Backup."""
    return [r for r in resources if r['resource_type'] == 'aws:backup:protected-resource']


def get_cloud_provider(resource_type: str) -> str:
    """Extract cloud provider from resource type prefix."""
    if resource_type.startswith('aws:'):
        return 'AWS'
    elif resource_type.startswith('azure:'):
        return 'Azure'
    elif resource_type.startswith('gcp:'):
        return 'GCP'
    elif resource_type.startswith('m365:'):
        return 'Microsoft 365'
    return 'Unknown'


def get_resource_category(resource_type: str) -> str:
    """Categorize resource types into logical groups."""
    # Compute resources
    compute_types = ['aws:ec2:instance', 'azure:vm', 'gcp:compute:instance']
    if resource_type in compute_types:
        return 'Compute'
    
    # Storage/Disk resources
    disk_types = ['aws:ec2:volume', 'azure:disk', 'gcp:compute:disk']
    if resource_type in disk_types:
        return 'Storage/Disk'
    
    # Snapshot resources
    snapshot_types = ['aws:ec2:snapshot', 'aws:rds:snapshot', 'aws:rds:cluster-snapshot', 
                      'azure:snapshot', 'gcp:compute:snapshot']
    if resource_type in snapshot_types:
        return 'Snapshots'
    
    # Database resources
    db_types = ['aws:rds:instance', 'aws:rds:cluster', 'azure:sql:database', 
                'azure:cosmosdb:account', 'gcp:sql:instance']
    if resource_type in db_types:
        return 'Database'
    
    # Object storage
    object_storage = ['aws:s3:bucket', 'azure:storage:account', 'gcp:storage:bucket']
    if resource_type in object_storage:
        return 'Object Storage'
    
    # File storage (EFS, FSx, Azure Files, etc.)
    file_storage = ['aws:efs:filesystem', 'aws:fsx:filesystem', 'azure:files:share', 
                    'gcp:filestore:instance']
    if resource_type in file_storage:
        return 'File Storage'
    
    # Backup/Recovery resources
    backup_types = ['aws:backup:plan', 'aws:backup:vault', 'aws:backup:selection',
                    'aws:backup:protected-resource', 'azure:recovery:vault', 
                    'azure:backup:policy']
    if resource_type in backup_types:
        return 'Backup/Recovery'
    
    # Container/Kubernetes
    container_types = ['azure:aks:cluster', 'gcp:container:cluster', 'aws:eks:cluster']
    if resource_type in container_types:
        return 'Container/K8s'
    
    # Serverless/Functions
    function_types = ['azure:function:app', 'aws:lambda:function', 'gcp:functions:function']
    if resource_type in function_types:
        return 'Serverless'
    
    # M365 resources
    if resource_type.startswith('m365:'):
        if 'mailbox' in resource_type:
            return 'M365 Mail'
        elif 'onedrive' in resource_type:
            return 'M365 OneDrive'
        elif 'sharepoint' in resource_type:
            return 'M365 SharePoint'
        elif 'teams' in resource_type:
            return 'M365 Teams'
        return 'M365 Other'
    
    return 'Other'


def count_resources_by_provider_and_type(resources: List[Dict]) -> Dict[str, Dict[str, int]]:
    """Count resources grouped by cloud provider and resource type."""
    counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in resources:
        resource_type = r.get('resource_type', 'unknown')
        provider = get_cloud_provider(resource_type)
        counts[provider][resource_type] += 1
    return {k: dict(v) for k, v in counts.items()}


def count_resources_by_category(resources: List[Dict]) -> Dict[str, Dict[str, Any]]:
    """Count resources and sizes by category."""
    counts: Dict[str, Dict[str, Any]] = defaultdict(lambda: {'count': 0, 'size_gb': 0})
    for r in resources:
        resource_type = r.get('resource_type', 'unknown')
        category = get_resource_category(resource_type)
        counts[category]['count'] += 1
        counts[category]['size_gb'] += r.get('size_gb', 0) or 0
    return dict(counts)


def extract_account_from_arn(arn: str) -> Optional[str]:
    """Extract AWS account ID from an ARN."""
    # ARN format: arn:aws:service:region:account:resource
    if not arn or not arn.startswith('arn:'):
        return None
    parts = arn.split(':')
    if len(parts) >= 5:
        return parts[4] if parts[4] else None
    return None


def build_backup_selection_index(selections: List[Dict]) -> tuple:
    """
    Build index of resource ARN patterns to backup plan names.
    Returns: (direct_index, wildcard_selections)
    - direct_index: dict of resource_arn -> [backup_plan_names]
    - wildcard_selections: list of (account, region, plan_name) for * selections
    """
    resource_to_plans: Dict[str, List[str]] = defaultdict(list)
    wildcard_selections: List[tuple] = []
    
    for selection in selections:
        plan_name = selection.get('metadata', {}).get('backup_plan_name', '')
        region = selection.get('region', '')
        
        # Extract account from IAM role ARN (arn:aws:iam::ACCOUNT:role/...)
        iam_role_arn = selection.get('metadata', {}).get('iam_role_arn', '')
        selection_account = extract_account_from_arn(iam_role_arn)
        
        # Direct resource ARNs
        resource_arns = selection.get('metadata', {}).get('resources', []) or []
        for arn in resource_arns:
            if arn == '*':
                # Wildcard selection - only applies to this account/region
                if selection_account:
                    wildcard_selections.append((selection_account, region, plan_name))
            else:
                resource_to_plans[arn].append(plan_name)
        
        # Tag-based selections (store the selection for later matching)
        list_of_tags = selection.get('metadata', {}).get('list_of_tags', [])
        if list_of_tags:
            # Store selection info for tag-based matching
            selection_key = f"__tag_selection__{selection['resource_id']}"
            resource_to_plans[selection_key] = [plan_name]
    
    return dict(resource_to_plans), wildcard_selections


def build_protected_resources_set(protected_resources: List[Dict]) -> set:
    """Build set of resource ARNs that have recovery points."""
    return {r.get('metadata', {}).get('resource_arn', '') for r in protected_resources}


def build_resource_id_to_account_map(protected_resources: List[Dict]) -> Dict[str, str]:
    """
    Build a mapping from short resource IDs to account IDs using protected_resources.
    Protected resources have full ARNs which contain the account ID.
    """
    id_to_account: Dict[str, str] = {}
    for r in protected_resources:
        arn = r.get('metadata', {}).get('resource_arn', '')
        if arn:
            account = extract_account_from_arn(arn)
            if account:
                # Extract the short ID from the ARN (e.g., i-xxx, vol-xxx)
                # ARN format: arn:aws:service:region:account:resource-type/resource-id
                parts = arn.split('/')
                if len(parts) >= 2:
                    short_id = parts[-1]
                    id_to_account[short_id] = account
                # Also store the full ARN mapping
                id_to_account[arn] = account
    return id_to_account


def get_resource_account(resource: Dict, id_to_account: Optional[Dict[str, str]] = None) -> str:
    """Get account ID for a resource.
    
    Uses account_id field directly from resource (set by collector).
    Falls back to ARN extraction or mapping lookup if needed.
    """
    # First check if resource has account_id field (standard field name)
    account = resource.get('account_id')
    if account:
        return account
    
    # Try to extract from resource_id if it's an ARN
    resource_id = resource.get('resource_id', '')
    account = extract_account_from_arn(resource_id)
    if account:
        return account
    
    # Look up in the ID-to-account mapping (fallback)
    if id_to_account and resource_id in id_to_account:
        return id_to_account[resource_id]
    
    return ''


def get_backup_plan_for_resource(
    resource: Dict,
    selection_index: Dict[str, List[str]],
    protected_set: set,
    backup_plans: List[Dict],
    wildcard_selections: Optional[List[tuple]] = None
) -> tuple:
    """
    Determine backup plan and protection source for a resource.
    
    Returns: (backup_plan_name, protection_source)
    protection_source: 'backup_selection' | 'recovery_point' | 'inferred' | None
    """
    resource_arn = resource.get('resource_id', '')
    resource_type = resource.get('resource_type', '')
    
    # Check if directly in a backup selection
    if resource_arn in selection_index:
        return (', '.join(selection_index[resource_arn]), 'backup_selection')
    
    # Check for wildcard matches - must match account (wildcards are account-scoped)
    if wildcard_selections:
        resource_account = extract_account_from_arn(resource_arn)
        resource_region = None
        # Extract region from ARN (arn:aws:service:REGION:account:resource)
        if resource_arn and resource_arn.startswith('arn:'):
            parts = resource_arn.split(':')
            if len(parts) >= 4:
                resource_region = parts[3]
        
        if resource_account:
            for sel_account, sel_region, plan_name in wildcard_selections:
                # Wildcard only matches resources in the SAME account
                if sel_account == resource_account:
                    # Optionally also check region if both are specified
                    if sel_region and resource_region and sel_region != resource_region:
                        continue
                    return (plan_name, 'backup_selection')
    
    # Check for ARN pattern wildcards (e.g., arn:aws:ec2:*:*:volume/*)
    for pattern, plans in selection_index.items():
        if pattern.startswith('__tag_selection__'):
            continue
        if '*' in pattern and pattern != '*':
            # Simple wildcard matching for ARN patterns
            import re
            regex = pattern.replace('*', '.*')
            if re.match(regex, resource_arn):
                return (', '.join(plans), 'backup_selection')
    
    # Check if resource has recovery points (is protected)
    if resource_arn in protected_set:
        return ('(has recovery points)', 'recovery_point')
    
    return (None, None)


def infer_backup_plan(snapshot: Dict, backup_plans: List[Dict]) -> Optional[str]:
    """Try to infer which backup plan created a snapshot based on tags/metadata."""
    tags = snapshot.get('tags', {}) or {}
    metadata = snapshot.get('metadata', {}) or {}
    description = metadata.get('description', '') or ''
    
    # Check multiple tag variations for backup plan info
    backup_type = tags.get('BackupType', '') or tags.get('backup', '') or ''
    
    for plan in backup_plans:
        plan_name = plan.get('name', '').lower()
        plan_rules = plan.get('metadata', {}).get('rule_names', [])
        
        # Direct match: backup tag matches plan name exactly
        if backup_type and backup_type.lower() == plan_name:
            return plan['name']
        
        # Partial match: backup tag contained in plan name or vice versa
        if backup_type:
            if backup_type.lower() in plan_name:
                return plan['name']
            if plan_name in backup_type.lower():
                return plan['name']
            for rule in plan_rules:
                if backup_type.lower() in rule.lower():
                    return plan['name']
        
        if 'daily' in description.lower() and 'daily' in plan_name:
            return plan['name']
        if 'weekly' in description.lower() and 'weekly' in plan_name:
            return plan['name']
    
    # Return the backup tag value if it looks like a plan name (even if not found in inventory)
    if backup_type:
        return f"{backup_type} (inferred)"
    
    # Check if this is an AWS Backup-created snapshot without a specific plan tag
    # These have description "This snapshot is created by the AWS Backup service."
    # and/or the aws:backup:source-resource tag
    if 'aws backup' in description.lower():
        return "AWS Backup"
    if 'aws:backup:source-resource' in tags:
        return "AWS Backup"
    
    return None


def is_backup_created_snapshot(snapshot: Dict) -> bool:
    """Check if a snapshot was created by AWS Backup service.
    
    AWS Backup-created snapshots have:
    - Description containing 'AWS Backup' or 'aws:backup'
    - Tags like 'aws:backup:source-resource'
    
    Excludes:
    - AMI snapshots (Created by CreateImage...)
    - Manual snapshots
    - Other automated snapshots not from AWS Backup
    """
    metadata = snapshot.get('metadata', {}) or {}
    description = (metadata.get('description') or '').lower()
    tags = snapshot.get('tags', {}) or {}
    
    # Check for AWS Backup markers
    if 'aws backup' in description:
        return True
    if 'aws:backup:source-resource' in tags:
        return True
    
    # Check for backup tag that matches known patterns
    backup_tag = tags.get('BackupType', '') or tags.get('backup', '')
    if backup_tag and 'aws backup' in description:
        return True
    
    return False


def get_protection_status(
    has_backup_plan: bool,
    snapshots: List[Dict],
    suffix: str = ''
) -> str:
    """Determine protection status based on backup plan and snapshot analysis.
    
    Args:
        has_backup_plan: Whether resource is in a backup selection/plan
        snapshots: List of snapshots for this resource
        suffix: Optional suffix like '(orphan)' to append
    
    Returns:
        Protection status string
    """
    suffix_str = f' {suffix}' if suffix else ''
    
    if has_backup_plan:
        return f'Protected{suffix_str}'
    
    if snapshots:
        # Check if any snapshot was created by AWS Backup
        backup_snapshots = [s for s in snapshots if is_backup_created_snapshot(s)]
        if backup_snapshots:
            return f'Protected{suffix_str}'
        else:
            return f'Has Snapshots (No Policy){suffix_str}'
    
    return f'Unprotected{suffix_str}'


def format_tags(tags: Dict) -> str:
    """Format tags as semicolon-separated string."""
    return '; '.join(f"{k}={v}" for k, v in tags.items()) if tags else ''


def is_replica_snapshot(snapshot: Dict) -> bool:
    """Check if a snapshot is a replica/copy from another region or source.
    
    Detection methods:
    - EBS snapshots: Description containing 'Copied from' or 'copied for'
    - AWS Backup recovery points: has parent_recovery_point_arn or source_backup_vault_arn
    """
    metadata = snapshot.get('metadata', {}) or {}
    description = (metadata.get('description') or '').lower()
    
    # EBS snapshot copy detection
    if 'copied' in description:
        return True
    
    # AWS Backup recovery point copy detection
    if metadata.get('is_replica'):
        return True
    if metadata.get('parent_recovery_point_arn'):
        return True
    if metadata.get('source_backup_vault_arn'):
        return True
    
    return False


def is_ami_snapshot(snapshot: Dict) -> bool:
    """Check if a snapshot was created as part of AMI generation.
    
    AMI-created snapshots are artifacts of machine image creation, not backups.
    They should not be counted as protection coverage.
    
    Detection patterns:
    - Description: "Created by CreateImage(i-xxxxx) for ami-xxxxx"
    - Description: "Auto-created snapshot for AMI"
    - Description: "Copied for DestinationAmi ami-xxxxx"
    - Tags: aws:backup:source-resource pointing to an AMI
    """
    metadata = snapshot.get('metadata', {}) or {}
    description = (metadata.get('description') or '').lower()
    tags = snapshot.get('tags', {}) or {}
    
    # Check description patterns
    if 'createimage' in description:
        return True
    if 'auto-created snapshot for ami' in description:
        return True
    if 'for ami-' in description or 'for ami ' in description:
        return True
    if 'destinationami' in description:
        return True
    
    # Check if source resource tag points to an AMI
    source_resource = tags.get('aws:backup:source-resource', '')
    if ':image/' in source_resource or source_resource.startswith('ami-'):
        return True
    
    return False


def get_eks_cluster(tags: Dict) -> str:
    """Detect EKS cluster membership from instance tags.
    
    EKS nodes are tagged with:
    - kubernetes.io/cluster/<cluster-name> = owned/shared
    - eks:cluster-name = <cluster-name>
    - aws:eks:cluster-name = <cluster-name>
    """
    if not tags:
        return ''
    
    # Check for explicit eks cluster name tags
    for key in ['eks:cluster-name', 'aws:eks:cluster-name']:
        if key in tags:
            return tags[key]
    
    # Check for kubernetes.io/cluster/* tags
    for key, value in tags.items():
        if key.startswith('kubernetes.io/cluster/'):
            # Extract cluster name from tag key
            cluster_name = key.replace('kubernetes.io/cluster/', '')
            if value in ('owned', 'shared'):
                return cluster_name
    
    return ''


def generate_report(inventory_path: str, output_path: str, preloaded_data: Optional[Dict[str, Any]] = None):
    """Generate the protection report CSV with full hierarchy.
    
    Args:
        inventory_path: Path to inventory JSON file
        output_path: Path to output Excel file
        preloaded_data: Optional pre-loaded inventory data to avoid double file load
        
    Raises:
        InventoryValidationError: If inventory data is invalid or corrupted
    """
    data = preloaded_data if preloaded_data else load_inventory(inventory_path)
    
    # Validate inventory structure before processing
    validate_inventory(data, inventory_path)
    
    resources = data['resources']
    
    # Build indexes
    resource_index = build_resource_index(resources)
    volume_index = build_volume_index(resources)  # Cross-account volume lookup
    
    # Categorize resources
    ec2_instances = get_ec2_instances(resources)
    ebs_volumes = get_ebs_volumes(resources)
    rds_instances = get_rds_instances(resources)
    other_resources = get_other_primary_resources(resources)
    snapshots = get_snapshots(resources)
    backup_plans = get_backup_plans(resources)
    
    # Get backup selections and protected resources for definitive matching
    backup_selections = get_backup_selections(resources)
    protected_resources = get_protected_resources(resources)
    
    # Build indexes for fast lookup
    selection_index, wildcard_selections = build_backup_selection_index(backup_selections)
    protected_set = build_protected_resources_set(protected_resources)
    id_to_account = build_resource_id_to_account_map(protected_resources)
    
    # Separate AMI snapshots from user/backup snapshots
    # AMI snapshots are artifacts of machine image creation, not backups
    user_snapshots = [s for s in snapshots if not is_ami_snapshot(s)]
    ami_snapshots = [s for s in snapshots if is_ami_snapshot(s)]
    
    # Build volume-to-instance mapping
    volume_to_instance = {}
    for vol in ebs_volumes:
        parent = vol.get('parent_resource_id')
        if parent and parent in resource_index:
            volume_to_instance[vol['resource_id']] = resource_index[parent]
    
    # Build snapshot-to-volume mapping
    snapshot_to_volume = {}
    for snap in user_snapshots:
        if snap['resource_type'] == 'aws:ec2:snapshot':
            parent = snap.get('parent_resource_id')
            if parent and parent in resource_index:
                snapshot_to_volume[snap['resource_id']] = resource_index[parent]
    
    rows = []
    processed_volumes = set()
    processed_snapshots = set()
    
    # Process EC2 Instances with their volumes and snapshots
    for instance in ec2_instances:
        instance_id = instance['resource_id']
        
        # Find volumes attached to this instance
        attached_volumes = [v for v in ebs_volumes if v.get('parent_resource_id') == instance_id]
        
        if attached_volumes:
            for volume in attached_volumes:
                processed_volumes.add(volume['resource_id'])
                volume_id = volume['resource_id']
                
                # Find snapshots for this volume
                volume_snapshots = [s for s in user_snapshots 
                                   if s.get('parent_resource_id') == volume_id]
                
                # Check for definitive backup plan assignment (volume or instance level)
                vol_plan, vol_source = get_backup_plan_for_resource(
                    volume, selection_index, protected_set, backup_plans, wildcard_selections
                )
                inst_plan, inst_source = get_backup_plan_for_resource(
                    instance, selection_index, protected_set, backup_plans, wildcard_selections
                )
                
                if volume_snapshots:
                    # Determine protection status once for all snapshots of this volume
                    has_backup_plan = bool(vol_plan or inst_plan)
                    protection_status = get_protection_status(has_backup_plan, volume_snapshots)
                    
                    for snap_idx, snap in enumerate(volume_snapshots):
                        # For backup_plan: prefer definitive, then snapshot inference
                        backup_plan = vol_plan or inst_plan or infer_backup_plan(snap, backup_plans) or ''
                        protection_source = vol_source or inst_source or ('inferred' if backup_plan else '')
                        processed_snapshots.add(snap['resource_id'])
                        
                        rows.append({
                            'service_family': 'EC2',
                            'account': get_resource_account(instance, id_to_account),
                            'instance_name': instance['name'],
                            'instance_id': instance_id,
                            'instance_type': instance.get('metadata', {}).get('instance_type', ''),
                            'eks_cluster': get_eks_cluster(instance.get('tags', {})),
                            'instance_region': instance['region'],
                            'instance_tags': format_tags(instance.get('tags', {})),
                            'volume_name': volume['name'],
                            'volume_id': volume_id,
                            'volume_size_gb': volume['size_gb'] if snap_idx == 0 else '',
                            'snapshot_name': snap['name'],
                            'snapshot_id': snap['resource_id'],
                            'snapshot_size_gb': snap['size_gb'],
                            'snapshot_created': snap.get('metadata', {}).get('start_time', ''),
                            'snapshot_description': snap.get('metadata', {}).get('description', ''),
                            'is_replica': 'Yes' if is_replica_snapshot(snap) else '',
                            'backup_plan': backup_plan,
                            'protection_source': protection_source,
                            'protection_status': protection_status
                        })
                else:
                    # Volume has no snapshots - but might still be in a backup plan
                    backup_plan = vol_plan or inst_plan or ''
                    protection_source = vol_source or inst_source or ''
                    has_backup_plan = bool(backup_plan)
                    
                    rows.append({
                        'service_family': 'EC2',
                        'account': get_resource_account(instance, id_to_account),
                        'instance_name': instance['name'],
                        'instance_id': instance_id,
                        'instance_type': instance.get('metadata', {}).get('instance_type', ''),
                        'eks_cluster': get_eks_cluster(instance.get('tags', {})),
                        'instance_region': instance['region'],
                        'instance_tags': format_tags(instance.get('tags', {})),
                        'volume_name': volume['name'],
                        'volume_id': volume_id,
                        'volume_size_gb': volume['size_gb'],
                        'snapshot_name': '',
                        'snapshot_id': '',
                        'snapshot_size_gb': '',
                        'snapshot_created': '',
                        'snapshot_description': '',
                        'is_replica': '',
                        'backup_plan': backup_plan,
                        'protection_source': protection_source,
                        'protection_status': 'In Backup Plan' if has_backup_plan else 'Unprotected'
                    })
        else:
            # Instance has no attached volumes - check if instance itself is in backup
            inst_plan, inst_source = get_backup_plan_for_resource(
                instance, selection_index, protected_set, backup_plans, wildcard_selections
            )
            rows.append({
                'service_family': 'EC2',
                'account': get_resource_account(instance, id_to_account),
                'instance_name': instance['name'],
                'instance_id': instance_id,
                'instance_type': instance.get('metadata', {}).get('instance_type', ''),
                'eks_cluster': get_eks_cluster(instance.get('tags', {})),
                'instance_region': instance['region'],
                'instance_tags': format_tags(instance.get('tags', {})),
                'volume_name': '(no volumes)',
                'volume_id': '',
                'volume_size_gb': '',
                'snapshot_name': '',
                'snapshot_id': '',
                'snapshot_size_gb': '',
                'snapshot_created': '',
                'snapshot_description': '',
                'is_replica': '',
                'backup_plan': inst_plan or '',
                'protection_source': inst_source or '',
                'protection_status': 'In Backup Plan' if inst_plan else 'No Storage'
            })
    
    # Process orphan volumes (not attached to any instance)
    orphan_volumes = [v for v in ebs_volumes if v['resource_id'] not in processed_volumes]
    for volume in orphan_volumes:
        volume_id = volume['resource_id']
        volume_snapshots = [s for s in user_snapshots if s.get('parent_resource_id') == volume_id]
        
        # Check for definitive backup plan assignment
        vol_plan, vol_source = get_backup_plan_for_resource(
            volume, selection_index, protected_set, backup_plans, wildcard_selections
        )
        
        if volume_snapshots:
            # Determine protection status once for all snapshots of this volume
            has_backup_plan = bool(vol_plan)
            protection_status = get_protection_status(has_backup_plan, volume_snapshots, suffix='(orphan)')
            
            for snap_idx, snap in enumerate(volume_snapshots):
                backup_plan = vol_plan or infer_backup_plan(snap, backup_plans) or ''
                protection_source = vol_source or ('inferred' if backup_plan else '')
                processed_snapshots.add(snap['resource_id'])
                
                rows.append({
                    'service_family': 'EBS',
                    'account': get_resource_account(volume, id_to_account),
                    'instance_name': '(orphan volume)',
                    'instance_id': '',
                    'instance_type': '',
                    'eks_cluster': '',
                    'instance_region': volume['region'],
                    'instance_tags': '',
                    'volume_name': volume['name'],
                    'volume_id': volume_id,
                    'volume_size_gb': volume['size_gb'] if snap_idx == 0 else '',
                    'snapshot_name': snap['name'],
                    'snapshot_id': snap['resource_id'],
                    'snapshot_size_gb': snap['size_gb'],
                    'snapshot_created': snap.get('metadata', {}).get('start_time', ''),
                    'snapshot_description': snap.get('metadata', {}).get('description', ''),
                    'is_replica': 'Yes' if is_replica_snapshot(snap) else '',
                    'backup_plan': backup_plan,
                    'protection_source': protection_source,
                    'protection_status': protection_status
                })
        else:
            # Orphan volume with no snapshots - check if in backup plan
            has_backup_plan = bool(vol_plan)
            rows.append({
                'service_family': 'EBS',
                'account': get_resource_account(volume, id_to_account),
                'instance_name': '(orphan volume)',
                'instance_id': '',
                'instance_type': '',
                'eks_cluster': '',
                'instance_region': volume['region'],
                'instance_tags': '',
                'volume_name': volume['name'],
                'volume_id': volume_id,
                'volume_size_gb': volume['size_gb'],
                'snapshot_name': '',
                'snapshot_id': '',
                'snapshot_size_gb': '',
                'snapshot_created': '',
                'snapshot_description': '',
                'is_replica': '',
                'backup_plan': vol_plan or '',
                'protection_source': vol_source or '',
                'protection_status': 'In Backup Plan (orphan)' if has_backup_plan else 'Unprotected (orphan)'
            })
    
    # Process RDS instances
    rds_snapshots = [s for s in user_snapshots if s['resource_type'] in ['aws:rds:snapshot', 'aws:rds:cluster-snapshot']]
    for rds in rds_instances:
        rds_id = rds['resource_id']
        db_identifier = rds['name']
        
        # Check for definitive backup plan assignment
        rds_plan, rds_source = get_backup_plan_for_resource(
            rds, selection_index, protected_set, backup_plans, wildcard_selections
        )
        
        # Find snapshots for this RDS
        rds_snaps = [s for s in rds_snapshots if s.get('parent_resource_id') == db_identifier]
        
        if rds_snaps:
            # Determine protection status once for all snapshots of this RDS
            has_backup_plan = bool(rds_plan)
            protection_status = get_protection_status(has_backup_plan, rds_snaps)
            
            for snap_idx, snap in enumerate(rds_snaps):
                backup_plan = rds_plan or infer_backup_plan(snap, backup_plans) or ''
                protection_source = rds_source or ('inferred' if backup_plan else '')
                processed_snapshots.add(snap['resource_id'])
                
                rows.append({
                    'service_family': 'RDS',
                    'account': get_resource_account(rds, id_to_account),
                    'instance_name': rds['name'],
                    'instance_id': rds_id,
                    'instance_type': rds.get('metadata', {}).get('instance_class', ''),
                    'eks_cluster': '',
                    'instance_region': rds['region'],
                    'instance_tags': format_tags(rds.get('tags', {})),
                    'volume_name': f"({rds['resource_type'].split(':')[-1]} storage)",
                    'volume_id': '',
                    'volume_size_gb': rds['size_gb'] if snap_idx == 0 else '',
                    'snapshot_name': snap['name'],
                    'snapshot_id': snap['resource_id'],
                    'snapshot_size_gb': snap['size_gb'],
                    'snapshot_created': snap.get('metadata', {}).get('snapshot_create_time', ''),
                    'snapshot_description': _format_rds_snapshot_description(snap),
                    'is_replica': 'Yes' if is_replica_snapshot(snap) else '',
                    'backup_plan': backup_plan,
                    'protection_source': protection_source,
                    'protection_status': protection_status
                })
        else:
            # RDS with no snapshots - check if in backup plan
            has_backup_plan = bool(rds_plan)
            rows.append({
                'service_family': 'RDS',
                'account': get_resource_account(rds, id_to_account),
                'instance_name': rds['name'],
                'instance_id': rds_id,
                'instance_type': rds.get('metadata', {}).get('instance_class', ''),
                'eks_cluster': '',
                'instance_region': rds['region'],
                'instance_tags': format_tags(rds.get('tags', {})),
                'volume_name': f"({rds['resource_type'].split(':')[-1]} storage)",
                'volume_id': '',
                'volume_size_gb': rds['size_gb'],
                'snapshot_name': '',
                'snapshot_id': '',
                'snapshot_size_gb': '',
                'snapshot_created': '',
                'snapshot_description': '',
                'is_replica': '',
                'backup_plan': rds_plan or '',
                'protection_source': rds_source or '',
                'protection_status': 'In Backup Plan' if has_backup_plan else 'Unprotected'
            })
    
    # Process orphan snapshots (no matching volume in same account - may be cross-account copies)
    ebs_snapshots = [s for s in user_snapshots if s['resource_type'] == 'aws:ec2:snapshot']
    orphan_snapshots = [s for s in ebs_snapshots if s['resource_id'] not in processed_snapshots]
    
    for snap in orphan_snapshots:
        snap_plan = infer_backup_plan(snap, backup_plans) or ''
        protection_source = 'inferred' if snap_plan else ''
        
        # Check if this snapshot is itself a known backup
        has_backup = is_backup_created_snapshot(snap)
        if has_backup and not snap_plan:
            snap_plan = 'AWS Backup'
            protection_source = 'inferred'
        
        # Try cross-account volume lookup
        parent_vol_id = snap.get('parent_resource_id', '')
        source_volume = volume_index.get(parent_vol_id)
        snap_account = snap.get('account_id', '')
        
        is_cross_account = False
        source_account = ''
        volume_name = f"(source: {parent_vol_id or 'unknown'})"
        
        if source_volume:
            source_account = source_volume.get('account_id', '')
            if source_account and source_account != snap_account:
                is_cross_account = True
                volume_name = source_volume.get('name', parent_vol_id)
        
        # Determine protection status
        if is_cross_account:
            protection_status = f"Cross-Account Copy (from {source_account[:12]}...)" if (snap_plan or has_backup) else f"Cross-Account Copy (from {source_account[:12]}...)"
        else:
            protection_status = 'Protected (orphan snapshot)' if (snap_plan or has_backup) else 'Orphan Snapshot'
        
        instance_name = '(cross-account copy)' if is_cross_account else '(orphan snapshot)'
        
        rows.append({
            'service_family': 'EBS',
            'account': snap_account,
            'instance_name': instance_name,
            'instance_id': '',
            'instance_type': '',
            'eks_cluster': '',
            'instance_region': snap['region'],
            'instance_tags': '',
            'volume_name': volume_name,
            'volume_id': parent_vol_id,
            'volume_size_gb': source_volume.get('size_gb', '') if source_volume else '',
            'snapshot_name': snap['name'],
            'snapshot_id': snap['resource_id'],
            'snapshot_size_gb': snap['size_gb'],
            'snapshot_created': snap.get('metadata', {}).get('start_time', ''),
            'snapshot_description': snap.get('metadata', {}).get('description', ''),
            'is_replica': 'Yes' if is_replica_snapshot(snap) else '',
            'backup_plan': snap_plan,
            'protection_source': f"cross-account from {source_account}" if is_cross_account else protection_source,
            'protection_status': protection_status
        })
    
    # Process AMI snapshots separately - these are artifacts, not backups
    ebs_ami_snapshots = [s for s in ami_snapshots if s['resource_type'] == 'aws:ec2:snapshot']
    
    for snap in ebs_ami_snapshots:
        rows.append({
            'service_family': 'EBS',
            'account': snap.get('account_id', ''),
            'instance_name': '(AMI artifact)',
            'instance_id': '',
            'instance_type': '',
            'eks_cluster': '',
            'instance_region': snap['region'],
            'instance_tags': '',
            'volume_name': f"(source: {snap.get('parent_resource_id', 'unknown')})",
            'volume_id': snap.get('parent_resource_id', ''),
            'volume_size_gb': '',
            'snapshot_name': snap['name'],
            'snapshot_id': snap['resource_id'],
            'snapshot_size_gb': snap['size_gb'],
            'snapshot_created': snap.get('metadata', {}).get('start_time', ''),
            'snapshot_description': snap.get('metadata', {}).get('description', ''),
            'is_replica': 'Yes' if is_replica_snapshot(snap) else '',
            'backup_plan': '',
            'protection_source': '',
            'protection_status': 'AMI Artifact'
        })
    
    # Calculate summary statistics
    protected = len([r for r in rows if 'Protected' in r['protection_status'] or 'In Backup Plan' in r['protection_status']])
    unprotected = len([r for r in rows if 'Unprotected' in r['protection_status']])
    in_backup_plan = len([r for r in rows if 'In Backup Plan' in r['protection_status']])
    ami_artifacts = len([r for r in rows if r['protection_status'] == 'AMI Artifact'])
    rows_with_snapshots = len([r for r in rows if r['snapshot_id']])
    
    unique_instances = len(set(r['instance_id'] for r in rows if r['instance_id']))
    unique_volumes = len(set(r['volume_id'] for r in rows if r['volume_id']))
    unique_snapshots = len(set(r['snapshot_id'] for r in rows if r['snapshot_id']))
    
    # Calculate sizes
    total_volume_size = sum(v['size_gb'] for v in ebs_volumes)
    orphan_volume_size = sum(v['size_gb'] for v in orphan_volumes)
    attached_volume_size = total_volume_size - orphan_volume_size
    total_snapshot_size = sum(s['size_gb'] for s in user_snapshots if s['resource_type'] == 'aws:ec2:snapshot')
    orphan_snapshot_size = sum(s['size_gb'] for s in orphan_snapshots)
    linked_snapshot_size = total_snapshot_size - orphan_snapshot_size
    total_rds_size = sum(r['size_gb'] for r in rds_instances)
    total_rds_snapshot_size = sum(s['size_gb'] for s in user_snapshots if s['resource_type'] in ['aws:rds:snapshot', 'aws:rds:cluster-snapshot'])
    ami_snapshot_size = sum(s['size_gb'] for s in ami_snapshots if s['resource_type'] == 'aws:ec2:snapshot')
    
    # Break down RDS snapshots by type (automated=included, manual/awsbackup=additional cost)
    rds_snaps_all = [s for s in user_snapshots if s['resource_type'] in ['aws:rds:snapshot', 'aws:rds:cluster-snapshot']]
    rds_snaps_automated = [s for s in rds_snaps_all if s.get('metadata', {}).get('snapshot_type') == 'automated']
    rds_snaps_manual = [s for s in rds_snaps_all if s.get('metadata', {}).get('snapshot_type') == 'manual']
    rds_snaps_awsbackup = [s for s in rds_snaps_all if s.get('metadata', {}).get('snapshot_type') == 'awsbackup']
    rds_automated_size = sum(s['size_gb'] for s in rds_snaps_automated)
    rds_manual_size = sum(s['size_gb'] for s in rds_snaps_manual)
    rds_awsbackup_size = sum(s['size_gb'] for s in rds_snaps_awsbackup)
    
    # Calculate protected vs unprotected sizes
    protected_volume_ids = set(r['volume_id'] for r in rows if r['volume_id'] and 'Protected' in r['protection_status'])
    unprotected_volume_ids = set(r['volume_id'] for r in rows if r['volume_id'] and 'Unprotected' in r['protection_status'])
    protected_volume_size = sum(v['size_gb'] for v in ebs_volumes if v['resource_id'] in protected_volume_ids)
    unprotected_volume_size = sum(v['size_gb'] for v in ebs_volumes if v['resource_id'] in unprotected_volume_ids)
    
    # Create Excel workbook
    wb = Workbook()
    
    # --- Summary Sheet ---
    ws_summary = wb.active
    assert ws_summary is not None, "Workbook must have an active sheet"
    ws_summary.title = "Summary"
    
    # Styles
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws_summary['A1'] = "Protection Report Summary"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:D1')
    
    # Account info
    ws_summary['A3'] = "Account ID:"
    ws_summary['B3'] = stringify_field(data.get('account_id', 'N/A'))
    ws_summary['A4'] = "Provider:"
    ws_summary['B4'] = stringify_field(data.get('provider', 'N/A'))
    ws_summary['A5'] = "Report Generated:"
    ws_summary['B5'] = stringify_field(data.get('timestamp', 'N/A'))
    
    # Resource counts section
    ws_summary['A7'] = "Resource Inventory"
    ws_summary['A7'].font = section_font
    
    summary_data = [
        ["Resource Type", "Count", "Size (GB)", "Notes"],
        ["EC2 Instances", len(ec2_instances), "", ""],
        ["EBS Volumes (Total)", len(ebs_volumes), round(total_volume_size, 1), ""],
        ["  - Attached Volumes", len(ebs_volumes) - len(orphan_volumes), round(attached_volume_size, 1), ""],
        ["  - Orphan Volumes", len(orphan_volumes), round(orphan_volume_size, 1), "Detached from instances"],
        ["EBS Snapshots (Total)", len([s for s in user_snapshots if s['resource_type'] == 'aws:ec2:snapshot']), round(total_snapshot_size, 1), "Excludes AMI artifacts"],
        ["  - Linked Snapshots", len(ebs_snapshots) - len(orphan_snapshots), round(linked_snapshot_size, 1), "Have matching volume"],
        ["  - Orphan Snapshots", len(orphan_snapshots), round(orphan_snapshot_size, 1), "No matching volume (cross-region copies, etc.)"],
        ["AMI Snapshots", len(ebs_ami_snapshots), round(ami_snapshot_size, 1), "Created by AMI generation, not backups"],
        ["RDS Databases", len(rds_instances), round(total_rds_size, 1), "Allocated storage"],
        ["RDS Snapshots (Total)", len(rds_snaps_all), round(total_rds_snapshot_size, 1), ""],
        ["  - Automated (Built-in)", len(rds_snaps_automated), round(rds_automated_size, 1), "INCLUDED with RDS - no extra charge"],
        ["  - Manual", len(rds_snaps_manual), round(rds_manual_size, 1), "Additional storage cost"],
        ["  - AWS Backup", len(rds_snaps_awsbackup), round(rds_awsbackup_size, 1), "Additional storage cost"],
        ["Backup Plans", len([r for r in backup_plans if r['resource_type'] == 'aws:backup:plan']), "", ""],
        ["Backup Selections", len(backup_selections), "", ""],
        ["Protected Resources", len(protected_resources), "", "Resources with recovery points"],
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=8):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 8:  # Header row
                cell.fill = header_fill
                cell.font = header_font_white

    
    # Protection status section
    ws_summary['A21'] = "Protection Status by Volume"
    ws_summary['A21'].font = section_font
    
    status_data = [
        ["Status", "Volume Count", "Size (GB)", "Percentage"],
        ["Protected (with snapshots)", len(protected_volume_ids), round(protected_volume_size, 1), round(protected_volume_size/total_volume_size*100, 1) if total_volume_size else 0],
        ["In Backup Plan (no snapshots)", len([r for r in rows if r['volume_id'] and 'In Backup Plan' in r['protection_status']]), "", ""],
        ["Unprotected", len(unprotected_volume_ids), round(unprotected_volume_size, 1), round(unprotected_volume_size/total_volume_size*100, 1) if total_volume_size else 0],
        ["Total", len(ebs_volumes), round(total_volume_size, 1), 100],
    ]
    
    for row_idx, row_data in enumerate(status_data, start=22):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 22:  # Header row
                cell.fill = header_fill
                cell.font = header_font_white
            elif row_idx == 26:  # Total row
                cell.font = Font(bold=True)
    
    # Adjust column widths for summary
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 30
    
    # --- Multi-Cloud Resource Overview Section ---
    provider_counts = count_resources_by_provider_and_type(resources)
    category_counts = count_resources_by_category(resources)
    
    current_row = 29  # Start after AWS-specific sections
    
    ws_summary.cell(row=current_row, column=1, value="Multi-Cloud Resource Overview").font = section_font
    current_row += 1
    ws_summary.cell(row=current_row, column=1, value="(Summary of all resources by cloud provider)")
    current_row += 2
    
    # Resources by Category
    ws_summary.cell(row=current_row, column=1, value="Resources by Category").font = section_font
    current_row += 1
    
    category_headers = ["Category", "Count", "Size (GB)"]
    for col_idx, header in enumerate(category_headers, start=1):
        cell = ws_summary.cell(row=current_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
    current_row += 1
    
    for category in sorted(category_counts.keys()):
        data = category_counts[category]
        cell_cat = ws_summary.cell(row=current_row, column=1, value=category)
        cell_cnt = ws_summary.cell(row=current_row, column=2, value=data['count'])
        cell_size = ws_summary.cell(row=current_row, column=3, value=round(data['size_gb'], 1) if data['size_gb'] else "")
        for cell in [cell_cat, cell_cnt, cell_size]:
            cell.border = thin_border
        current_row += 1
    
    current_row += 1
    
    # Resources by Provider
    ws_summary.cell(row=current_row, column=1, value="Resources by Cloud Provider").font = section_font
    current_row += 1
    
    provider_headers = ["Provider", "Resource Type", "Count"]
    for col_idx, header in enumerate(provider_headers, start=1):
        cell = ws_summary.cell(row=current_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
    current_row += 1
    
    for provider in sorted(provider_counts.keys()):
        type_counts = provider_counts[provider]
        for resource_type in sorted(type_counts.keys()):
            cell_prov = ws_summary.cell(row=current_row, column=1, value=provider)
            cell_type = ws_summary.cell(row=current_row, column=2, value=resource_type)
            cell_cnt = ws_summary.cell(row=current_row, column=3, value=type_counts[resource_type])
            for cell in [cell_prov, cell_type, cell_cnt]:
                cell.border = thin_border
            current_row += 1
    
    # --- Protection Report Sheet ---
    ws_report = wb.create_sheet(title="Protection Report")
    
    fieldnames = [
        'service_family', 'account', 'instance_name', 'instance_id', 'instance_type', 'eks_cluster', 'instance_region', 'instance_tags',
        'volume_name', 'volume_id', 'volume_size_gb',
        'snapshot_name', 'snapshot_id', 'snapshot_size_gb',
        'snapshot_created', 'snapshot_description', 'is_replica', 'backup_plan', 'protection_source', 'protection_status'
    ]
    
    # Pretty header mapping
    header_display = {
        'service_family': 'Service',
        'account': 'Account',
        'instance_name': 'Instance Name',
        'instance_id': 'Instance ID',
        'instance_type': 'Instance Type',
        'eks_cluster': 'EKS Cluster',
        'instance_region': 'Region',
        'instance_tags': 'Tags',
        'volume_name': 'Volume Name',
        'volume_id': 'Volume ID',
        'volume_size_gb': 'Volume Size (GB)',
        'snapshot_name': 'Snapshot Name',
        'snapshot_id': 'Snapshot ID',
        'snapshot_size_gb': 'Snapshot Size (GB)',
        'snapshot_created': 'Snapshot Created',
        'snapshot_description': 'Snapshot Description',
        'is_replica': 'Replica',
        'backup_plan': 'Backup Plan',
        'protection_source': 'Protection Source',
        'protection_status': 'Protection Status'
    }
    
    # Header row
    for col_idx, header in enumerate(fieldnames, start=1):
        cell = ws_report.cell(row=1, column=col_idx, value=header_display.get(header, header))
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
    
    # Data rows
    status_colors = {
        'Protected': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'In Backup Plan': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'Unprotected': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'No Storage': PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"),
    }
    
    for row_idx, row_data in enumerate(rows, start=2):
        status = row_data.get('protection_status', '')
        
        # Determine row color based on status
        row_fill = None
        for status_key, fill in status_colors.items():
            if status_key in status:
                row_fill = fill
                break
        
        for col_idx, field in enumerate(fieldnames, start=1):
            cell = ws_report.cell(row=row_idx, column=col_idx, value=row_data.get(field, ''))
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
    
    # Adjust column widths for report
    column_widths = {
        'A': 10, 'B': 20, 'C': 22, 'D': 15, 'E': 18, 'F': 12, 'G': 40,
        'H': 18, 'I': 25, 'J': 15,
        'K': 18, 'L': 45, 'M': 15,
        'N': 22, 'O': 30, 'P': 25, 'Q': 15, 'R': 18
    }
    for col, width in column_widths.items():
        ws_report.column_dimensions[col].width = width
    
    # Freeze header row and enable filters
    ws_report.freeze_panes = 'A2'
    ws_report.auto_filter.ref = f"A1:R{len(rows) + 1}"
    
    # --- Backup Plans Sheet ---
    ws_plans = wb.create_sheet(title="Backup Plans")
    
    # Get only backup plans (not vaults)
    plans_only = [r for r in backup_plans if r['resource_type'] == 'aws:backup:plan']
    
    plan_fieldnames = [
        'Plan Name', 'Plan ID', 'Region', 'Version ID', 'Creation Date', 
        'Last Execution', 'Number of Rules', 'Rule Names', 'Rule Details'
    ]
    
    # Header row
    for col_idx, header in enumerate(plan_fieldnames, start=1):
        cell = ws_plans.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
    
    # Data rows
    for row_idx, plan in enumerate(plans_only, start=2):
        metadata = plan.get('metadata', {})
        rules = metadata.get('rules', [])
        
        # Format rule details as readable string
        rule_details_str = ""
        for rule in rules:
            details = []
            if rule.get('schedule'):
                details.append(f"Schedule: {rule['schedule']}")
            if rule.get('target_vault'):
                details.append(f"Vault: {rule['target_vault']}")
            if rule.get('delete_after_days'):
                details.append(f"Retention: {rule['delete_after_days']} days")
            if rule.get('move_to_cold_after_days'):
                details.append(f"Cold storage: {rule['move_to_cold_after_days']} days")
            rule_details_str += f"{rule.get('rule_name', 'unnamed')}: {', '.join(details)}\n"
        
        plan_data = [
            plan.get('name', ''),
            metadata.get('backup_plan_id', ''),
            plan.get('region', ''),
            metadata.get('version_id', ''),
            metadata.get('creation_date', ''),
            metadata.get('last_execution_date', '') or 'Never',
            metadata.get('number_of_rules', 0),
            ', '.join(metadata.get('rule_names', [])),
            rule_details_str.strip()
        ]
        
        for col_idx, value in enumerate(plan_data, start=1):
            cell = ws_plans.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx == 9:  # Rule Details column - allow text wrap
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths for plans
    plan_column_widths = {
        'A': 25, 'B': 40, 'C': 12, 'D': 50, 'E': 25,
        'F': 18, 'G': 15, 'H': 30, 'I': 60
    }
    for col, width in plan_column_widths.items():
        ws_plans.column_dimensions[col].width = width
    
    # Freeze header row and enable filters
    ws_plans.freeze_panes = 'A2'
    ws_plans.auto_filter.ref = f"A1:I{len(plans_only) + 1}"
    
    # --- Backup Selections Sheet (if any exist) ---
    if backup_selections:
        ws_selections = wb.create_sheet(title="Backup Selections")
        
        selection_fieldnames = [
            'Selection Name', 'Backup Plan', 'Plan ID', 'Region', 
            'IAM Role', 'Resources', 'Tags Conditions', 'Not Resources'
        ]
        
        # Header row
        for col_idx, header in enumerate(selection_fieldnames, start=1):
            cell = ws_selections.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font_white
            cell.border = thin_border
        
        # Data rows
        for row_idx, selection in enumerate(backup_selections, start=2):
            metadata = selection.get('metadata', {})
            
            selection_data = [
                selection.get('name', ''),
                metadata.get('backup_plan_name', ''),
                metadata.get('backup_plan_id', ''),
                selection.get('region', ''),
                metadata.get('iam_role_arn', ''),
                '\n'.join(metadata.get('resources', [])),
                str(metadata.get('list_of_tags', [])) if metadata.get('list_of_tags') else '',
                '\n'.join(metadata.get('not_resources', []))
            ]
            
            for col_idx, value in enumerate(selection_data, start=1):
                cell = ws_selections.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx in [6, 7, 8]:  # Multi-line columns
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column widths
        selection_column_widths = {
            'A': 25, 'B': 25, 'C': 40, 'D': 12,
            'E': 50, 'F': 60, 'G': 40, 'H': 40
        }
        for col, width in selection_column_widths.items():
            ws_selections.column_dimensions[col].width = width
        
        ws_selections.freeze_panes = 'A2'
        ws_selections.auto_filter.ref = f"A1:H{len(backup_selections) + 1}"
    
    # --- All Resources Sheet (Multi-Cloud Overview) ---
    ws_all = wb.create_sheet(title="All Resources")
    
    all_resources_headers = [
        'Cloud Provider', 'Resource Type', 'Category', 'Resource ID', 
        'Name', 'Region', 'Size (GB)', 'Created/Modified', 'Tags'
    ]
    
    # Header row
    for col_idx, header in enumerate(all_resources_headers, start=1):
        cell = ws_all.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font_white
        cell.border = thin_border
    
    # Provider colors for visual distinction
    provider_colors = {
        'AWS': PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid"),  # AWS Orange
        'Azure': PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid"),  # Azure Blue
        'GCP': PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid"),  # Google Blue
        'Microsoft 365': PatternFill(start_color="D83B01", end_color="D83B01", fill_type="solid"),  # M365 Orange-Red
    }
    provider_text_color = Font(color="FFFFFF")
    
    # Data rows - sorted by provider, then resource type
    sorted_resources = sorted(resources, key=lambda r: (get_cloud_provider(r.get('resource_type', '')), r.get('resource_type', ''), r.get('name', '')))
    
    for row_idx, resource in enumerate(sorted_resources, start=2):
        resource_type = resource.get('resource_type', '')
        provider = get_cloud_provider(resource_type)
        category = get_resource_category(resource_type)
        
        # Get created/modified date from metadata
        metadata = resource.get('metadata', {})
        created = metadata.get('created_time', '') or metadata.get('start_time', '') or metadata.get('creation_time', '')
        
        row_data = [
            provider,
            resource_type,
            category,
            resource.get('resource_id', ''),
            resource.get('name', ''),
            resource.get('region', ''),
            resource.get('size_gb', ''),
            created,
            format_tags(resource.get('tags', {}))
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Color the provider column
            if col_idx == 1 and provider in provider_colors:
                cell.fill = provider_colors[provider]
                cell.font = provider_text_color
            
            # Wrap text for tags column
            if col_idx == 9:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Adjust column widths
    all_resources_column_widths = {
        'A': 15, 'B': 30, 'C': 18, 'D': 60,
        'E': 40, 'F': 15, 'G': 12, 'H': 25, 'I': 50
    }
    for col, width in all_resources_column_widths.items():
        ws_all.column_dimensions[col].width = width
    
    ws_all.freeze_panes = 'A2'
    if sorted_resources:
        ws_all.auto_filter.ref = f"A1:I{len(sorted_resources) + 1}"
    
    # Save workbook
    wb.save(output_path)
    
    # Print summary to console - include multi-cloud summary
    print(f"Protection Report Generated: {output_path}")
    print(f"=" * 60)
    
    # Multi-cloud summary first
    total_resources = len(resources)
    print(f"TOTAL RESOURCES: {total_resources}")
    print(f"-" * 60)
    print("Resources by Cloud Provider:")
    for provider in sorted(provider_counts.keys()):
        provider_total = sum(provider_counts[provider].values())
        print(f"  {provider}: {provider_total}")
    print(f"-" * 60)
    print("Resources by Category:")
    for category in sorted(category_counts.keys()):
        cat_data = category_counts[category]
        size_str = f" ({cat_data['size_gb']:,.1f} GB)" if cat_data['size_gb'] else ""
        print(f"  {category}: {cat_data['count']}{size_str}")
    print(f"-" * 60)
    
    # AWS-specific details (if AWS resources exist)
    if ec2_instances or ebs_volumes or rds_instances:
        print("AWS Resource Details:")
        print(f"  EC2 Instances: {len(ec2_instances)}")
        print(f"  EBS Volumes: {len(ebs_volumes)} ({total_volume_size:,.1f} GB)")
        print(f"    - Attached: {len(ebs_volumes) - len(orphan_volumes)} ({attached_volume_size:,.1f} GB)")
        print(f"    - Orphan: {len(orphan_volumes)} ({orphan_volume_size:,.1f} GB)")
        print(f"  EBS Snapshots: {len(ebs_snapshots)} ({total_snapshot_size:,.1f} GB)")
        print(f"    - Linked: {len(ebs_snapshots) - len(orphan_snapshots)} ({linked_snapshot_size:,.1f} GB)")
        print(f"    - Orphan: {len(orphan_snapshots)} ({orphan_snapshot_size:,.1f} GB)")
        print(f"  RDS Databases: {len(rds_instances)} ({total_rds_size:,.1f} GB)")
        print(f"  RDS Snapshots: {len([s for s in user_snapshots if s['resource_type'] in ['aws:rds:snapshot', 'aws:rds:cluster-snapshot']])} ({total_rds_snapshot_size:,.1f} GB)")
        print(f"  Backup Plans: {len(plans_only)}")
        print(f"  Backup Selections: {len(backup_selections)}")
        print(f"  Protected Resources: {len(protected_resources)}")
        if total_volume_size:
            print(f"  Protected Volume Size: {protected_volume_size:,.1f} GB ({protected_volume_size/total_volume_size*100:.1f}%)")
            print(f"  Unprotected Volume Size: {unprotected_volume_size:,.1f} GB ({unprotected_volume_size/total_volume_size*100:.1f}%)")


def generate_output_filename(inventory_path: str, data: Dict[str, Any]) -> str:
    """Generate a unique output filename based on inventory data.
    
    Priority for identifier:
    1. org_name field (if set by user/collector)
    2. First account_id (shortened to last 4 digits for readability)
    3. Fallback to 'unknown'
    
    Also includes run_id timestamp for uniqueness.
    
    Args:
        inventory_path: Path to the inventory file
        data: Loaded inventory data
    
    Returns:
        Output path like 'protection_report_acme-corp_162849.xlsx'
    """
    # Get identifier - prefer org_name, fallback to account_id
    identifier = data.get('org_name', '')
    
    if not identifier:
        account_id = data.get('account_id', '')
        if isinstance(account_id, list):
            account_id = account_id[0] if account_id else ''
        # Use last 4 digits of account ID for brevity
        identifier = account_id[-4:] if account_id else 'unknown'
    
    # Sanitize identifier for filename (remove special chars)
    identifier = re.sub(r'[^\w\-]', '_', str(identifier))
    
    # Extract timestamp from run_id (format: YYYYMMDD-HHMMSS-uuid)
    run_id = data.get('run_id', '')
    timestamp = ''
    if run_id:
        # Extract the HHMMSS part
        parts = run_id.split('-')
        if len(parts) >= 2:
            timestamp = parts[1]  # HHMMSS
    
    # Build filename
    if timestamp:
        filename = f"protection_report_{identifier}_{timestamp}.xlsx"
    else:
        filename = f"protection_report_{identifier}.xlsx"
    
    # Put output in same directory as input
    input_dir = os.path.dirname(inventory_path)
    return os.path.join(input_dir, filename) if input_dir else filename


if __name__ == '__main__':
    inventory_path = sys.argv[1] if len(sys.argv) > 1 else 'tests/sample_output/cca_inv_162849.json'
    
    # Load inventory once for both filename generation and report
    with open(inventory_path) as f:
        inventory_data = json.load(f)
    
    # If output path not provided, auto-generate based on inventory data
    if len(sys.argv) > 2:
        output_path = sys.argv[2]
    else:
        output_path = generate_output_filename(inventory_path, inventory_data)
    
    generate_report(inventory_path, output_path, preloaded_data=inventory_data)
