"""
Tests for scripts/generate_cost_report.py helper functions.

Covers:
- load_json function
- analyze_costs function
- generate_optimization_recommendations function
- Excel generation (integration tests)
"""
import json
import os
import sys
import tempfile

import pytest

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'scripts'))

from scripts.generate_cost_report import (
    analyze_costs,
    generate_optimization_recommendations,
    load_json,
)

# =============================================================================
# Fixtures
# =============================================================================

@pytest.fixture
def sample_cost_inventory():
    """Sample cost inventory data for testing."""
    return {
        "run_id": "test_cost_001",
        "timestamp": "2026-02-11T12:00:00Z",
        "period": {
            "start": "2026-01-01",
            "end": "2026-01-31"
        },
        "providers": ["aws", "azure"],
        "total_cost": 5000.00,
        "total_records": 6,
        "records": [
            {
                "provider": "aws",
                "category": "ec2_snapshot",
                "service": "Amazon EC2 EBS - Snapshot",
                "cost": 1500.00,
                "period_start": "2026-01-01",
                "region": "us-east-1",
                "usage_quantity": 15000.0,
                "usage_unit": "GB-Month"
            },
            {
                "provider": "aws",
                "category": "rds_snapshot",
                "service": "Amazon RDS - Snapshot",
                "cost": 500.00,
                "period_start": "2026-01-01",
                "region": "us-east-1",
                "usage_quantity": 5000.0,
                "usage_unit": "GB-Month"
            },
            {
                "provider": "aws",
                "category": "backup_storage",
                "service": "AWS Backup Storage",
                "cost": 800.00,
                "period_start": "2026-01-01",
                "region": "us-east-1",
                "usage_quantity": 16000.0,
                "usage_unit": "GB-Month"
            },
            {
                "provider": "azure",
                "category": "disk_snapshot",
                "service": "Azure Managed Disk Snapshots",
                "cost": 1200.00,
                "period_start": "2026-01-01",
                "region": "eastus",
                "usage_quantity": 12000.0,
                "usage_unit": "GB-Month"
            },
            {
                "provider": "azure",
                "category": "backup_storage",
                "service": "Azure Backup - Recovery Vault Storage",
                "cost": 700.00,
                "period_start": "2026-01-01",
                "region": "eastus",
                "usage_quantity": 14000.0,
                "usage_unit": "GB-Month"
            },
            {
                "provider": "azure",
                "category": "blob_storage",
                "service": "Azure Blob Storage - Hot",
                "cost": 300.00,
                "period_start": "2026-01-01",
                "region": "eastus",
                "usage_quantity": 13000.0,
                "usage_unit": "GB-Month"
            }
        ]
    }


@pytest.fixture
def sample_cost_summary():
    """Sample cost summary data for testing."""
    return {
        "run_id": "test_cost_001",
        "timestamp": "2026-02-11T12:00:00Z",
        "period": {
            "start": "2026-01-01",
            "end": "2026-01-31"
        },
        "summary": {
            "total_cost": 5000.00,
            "currency": "USD",
            "providers": {
                "aws": {
                    "total_cost": 2800.00,
                    "categories": {
                        "ec2_snapshot": 1500.00,
                        "rds_snapshot": 500.00,
                        "backup_storage": 800.00
                    }
                },
                "azure": {
                    "total_cost": 2200.00,
                    "categories": {
                        "disk_snapshot": 1200.00,
                        "backup_storage": 700.00,
                        "blob_storage": 300.00
                    }
                }
            },
            "month_over_month": {
                "previous_month_total": 4500.00,
                "current_month_total": 5000.00,
                "change_amount": 500.00,
                "change_percent": 11.1
            }
        }
    }


@pytest.fixture
def multi_month_cost_inventory():
    """Sample cost inventory with multiple months of data."""
    return {
        "run_id": "test_cost_multimonth",
        "timestamp": "2026-02-11T12:00:00Z",
        "period": {
            "start": "2025-11-01",
            "end": "2026-01-31"
        },
        "providers": ["aws"],
        "total_cost": 4500.00,
        "total_records": 3,
        "records": [
            {
                "provider": "aws",
                "category": "ec2_snapshot",
                "service": "Amazon EC2 EBS - Snapshot",
                "cost": 1400.00,
                "period_start": "2025-11-01",
                "region": "us-east-1"
            },
            {
                "provider": "aws",
                "category": "ec2_snapshot",
                "service": "Amazon EC2 EBS - Snapshot",
                "cost": 1500.00,
                "period_start": "2025-12-01",
                "region": "us-east-1"
            },
            {
                "provider": "aws",
                "category": "ec2_snapshot",
                "service": "Amazon EC2 EBS - Snapshot",
                "cost": 1600.00,
                "period_start": "2026-01-01",
                "region": "us-east-1"
            }
        ]
    }


@pytest.fixture
def temp_cost_inventory_file(sample_cost_inventory):
    """Create a temporary cost inventory file for testing."""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
        json.dump(sample_cost_inventory, f)
        temp_path = f.name
    yield temp_path
    os.unlink(temp_path)


@pytest.fixture
def temp_cost_summary_file(sample_cost_summary):
    """Create a temporary cost summary file for testing."""
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
        json.dump(sample_cost_summary, f)
        temp_path = f.name
    yield temp_path
    os.unlink(temp_path)


# =============================================================================
# Tests for load_json
# =============================================================================

class TestLoadJson:
    """Tests for load_json function."""

    def test_load_json_valid_file(self, temp_cost_inventory_file):
        """Test loading a valid JSON file."""
        result = load_json(temp_cost_inventory_file)
        assert result['run_id'] == 'test_cost_001'
        assert result['total_cost'] == 5000.00
        assert len(result['records']) == 6

    def test_load_json_file_not_found(self):
        """Test FileNotFoundError for missing file."""
        with pytest.raises(FileNotFoundError):
            load_json('/nonexistent/path/file.json')

    def test_load_json_invalid_json(self):
        """Test error handling for invalid JSON content."""
        with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
            f.write('not valid json {{{')
            temp_path = f.name

        try:
            with pytest.raises(json.JSONDecodeError):
                load_json(temp_path)
        finally:
            os.unlink(temp_path)


# =============================================================================
# Tests for analyze_costs
# =============================================================================

class TestAnalyzeCosts:
    """Tests for analyze_costs function."""

    def test_analyze_costs_metadata(self, sample_cost_inventory, sample_cost_summary):
        """Test that metadata is correctly extracted."""
        result = analyze_costs(sample_cost_inventory, sample_cost_summary)

        assert result['metadata']['run_id'] == 'test_cost_001'
        assert result['metadata']['period_start'] == '2026-01-01'
        assert result['metadata']['period_end'] == '2026-01-31'
        assert 'aws' in result['metadata']['providers']
        assert 'azure' in result['metadata']['providers']

    def test_analyze_costs_totals(self, sample_cost_inventory, sample_cost_summary):
        """Test that totals are correctly calculated."""
        result = analyze_costs(sample_cost_inventory, sample_cost_summary)

        assert result['totals']['total_cost'] == 5000.00
        assert result['totals']['total_records'] == 6

    def test_analyze_costs_by_provider(self, sample_cost_inventory, sample_cost_summary):
        """Test costs grouped by provider."""
        result = analyze_costs(sample_cost_inventory, sample_cost_summary)

        assert 'aws' in result['by_provider']
        assert 'azure' in result['by_provider']

        # AWS: 1500 + 500 + 800 = 2800
        assert result['by_provider']['aws']['cost'] == 2800.00

        # Azure: 1200 + 700 + 300 = 2200
        assert result['by_provider']['azure']['cost'] == 2200.00

    def test_analyze_costs_by_category(self, sample_cost_inventory, sample_cost_summary):
        """Test costs grouped by category."""
        result = analyze_costs(sample_cost_inventory, sample_cost_summary)

        assert 'ec2_snapshot' in result['by_category']
        assert 'rds_snapshot' in result['by_category']
        assert 'backup_storage' in result['by_category']
        assert 'disk_snapshot' in result['by_category']
        assert 'blob_storage' in result['by_category']

        assert result['by_category']['ec2_snapshot']['cost'] == 1500.00
        assert result['by_category']['backup_storage']['cost'] == 1500.00  # 800 + 700

    def test_analyze_costs_by_service(self, sample_cost_inventory, sample_cost_summary):
        """Test costs grouped by service."""
        result = analyze_costs(sample_cost_inventory, sample_cost_summary)

        assert 'Amazon EC2 EBS - Snapshot' in result['by_service']
        assert 'Amazon RDS - Snapshot' in result['by_service']
        assert 'Azure Managed Disk Snapshots' in result['by_service']

        assert result['by_service']['Amazon EC2 EBS - Snapshot']['cost'] == 1500.00
        assert result['by_service']['Amazon EC2 EBS - Snapshot']['provider'] == 'aws'

    def test_analyze_costs_monthly_trend(self, multi_month_cost_inventory, sample_cost_summary):
        """Test monthly trend calculation."""
        result = analyze_costs(multi_month_cost_inventory, sample_cost_summary)

        assert '2025-11' in result['monthly_trend']
        assert '2025-12' in result['monthly_trend']
        assert '2026-01' in result['monthly_trend']

        assert result['monthly_trend']['2025-11']['aws'] == 1400.00
        assert result['monthly_trend']['2025-12']['aws'] == 1500.00
        assert result['monthly_trend']['2026-01']['aws'] == 1600.00

    def test_analyze_costs_empty_records(self, sample_cost_summary):
        """Test handling of empty records."""
        empty_inventory = {
            'run_id': 'empty',
            'timestamp': '2026-01-01T00:00:00Z',
            'period': {'start': '2026-01-01', 'end': '2026-01-31'},
            'providers': [],
            'total_cost': 0,
            'total_records': 0,
            'records': []
        }

        result = analyze_costs(empty_inventory, sample_cost_summary)

        assert result['totals']['total_cost'] == 0
        assert result['totals']['total_records'] == 0
        assert len(result['by_provider']) == 0
        assert len(result['by_category']) == 0

    def test_analyze_costs_has_optimization_recommendations(self, sample_cost_inventory, sample_cost_summary):
        """Test that optimization recommendations are generated."""
        result = analyze_costs(sample_cost_inventory, sample_cost_summary)

        assert 'optimization' in result
        assert isinstance(result['optimization'], list)


# =============================================================================
# Tests for generate_optimization_recommendations
# =============================================================================

class TestGenerateOptimizationRecommendations:
    """Tests for generate_optimization_recommendations function."""

    def test_snapshot_retention_recommendation(self):
        """Test recommendation for high snapshot costs."""
        analysis = {
            'totals': {'total_cost': 10000.00},
            'by_category': {
                'ec2_snapshot': {'cost': 4000.00, 'providers': {'aws': 4000.00}},
                'rds_snapshot': {'cost': 1000.00, 'providers': {'aws': 1000.00}}
            },
            'by_service': {},
            'by_provider': {'aws': {'cost': 5000.00, 'services': {}}}
        }

        recommendations = generate_optimization_recommendations(analysis)

        # Should have snapshot retention recommendation
        snapshot_rec = next((r for r in recommendations if 'Snapshot' in r['title']), None)
        assert snapshot_rec is not None
        assert snapshot_rec['current_cost'] == 5000.00  # Combined snapshot costs
        assert snapshot_rec['potential_savings'] == 1500.00  # 30% savings

    def test_backup_vault_optimization_recommendation(self):
        """Test recommendation for backup vault storage."""
        analysis = {
            'totals': {'total_cost': 5000.00},
            'by_category': {
                'backup_storage': {'cost': 2000.00, 'providers': {'aws': 2000.00}}
            },
            'by_service': {
                'AWS Backup Storage': {'cost': 1500.00, 'category': 'backup_storage', 'provider': 'aws'},
                'AWS Backup Vault Storage': {'cost': 500.00, 'category': 'backup_storage', 'provider': 'aws'}
            },
            'by_provider': {'aws': {'cost': 5000.00, 'services': {}}}
        }

        recommendations = generate_optimization_recommendations(analysis)

        # Should have backup vault recommendation
        vault_rec = next((r for r in recommendations if 'Vault' in r['title']), None)
        assert vault_rec is not None
        assert vault_rec['savings_percent'] == 20

    def test_multicloud_balance_recommendation(self):
        """Test recommendation for imbalanced multi-cloud costs."""
        analysis = {
            'totals': {'total_cost': 10000.00},
            'by_category': {},
            'by_service': {},
            'by_provider': {
                'aws': {'cost': 8000.00, 'services': {}},
                'azure': {'cost': 2000.00, 'services': {}}
            }
        }

        recommendations = generate_optimization_recommendations(analysis)

        # Should have multi-cloud balance recommendation
        multicloud_rec = next((r for r in recommendations if 'Multi-Cloud' in r['title']), None)
        assert multicloud_rec is not None
        assert 'AWS' in multicloud_rec['description']

    def test_general_recommendation_when_no_specific_apply(self):
        """Test that general recommendation is given when no specific ones apply."""
        analysis = {
            'totals': {'total_cost': 1000.00},
            'by_category': {
                'other': {'cost': 1000.00, 'providers': {'aws': 1000.00}}
            },
            'by_service': {
                'Some Service': {'cost': 1000.00, 'category': 'other', 'provider': 'aws'}
            },
            'by_provider': {'aws': {'cost': 1000.00, 'services': {}}}
        }

        recommendations = generate_optimization_recommendations(analysis)

        # Should have at least one recommendation
        assert len(recommendations) >= 1
        # Should have the general tiering recommendation
        tiering_rec = next((r for r in recommendations if 'Tiering' in r['title']), None)
        assert tiering_rec is not None

    def test_high_priority_for_large_snapshot_costs(self):
        """Test that high priority is assigned for large snapshot costs."""
        analysis = {
            'totals': {'total_cost': 10000.00},
            'by_category': {
                'ec2_snapshot': {'cost': 5000.00, 'providers': {'aws': 5000.00}}
            },
            'by_service': {},
            'by_provider': {'aws': {'cost': 10000.00, 'services': {}}}
        }

        recommendations = generate_optimization_recommendations(analysis)

        snapshot_rec = next((r for r in recommendations if 'Snapshot' in r['title']), None)
        assert snapshot_rec is not None
        assert snapshot_rec['priority'] == 'High'

    def test_recommendations_have_required_fields(self, sample_cost_inventory, sample_cost_summary):
        """Test that all recommendations have required fields."""
        analysis = analyze_costs(sample_cost_inventory, sample_cost_summary)
        recommendations = analysis['optimization']

        required_fields = ['title', 'category', 'current_cost', 'potential_savings',
                          'savings_percent', 'description', 'priority']

        for rec in recommendations:
            for field in required_fields:
                assert field in rec, f"Missing required field: {field}"


# =============================================================================
# Integration Tests
# =============================================================================

class TestCostReportIntegration:
    """Integration tests for the cost report generator."""

    def test_analyze_combined_multicloud_data(self):
        """Test analysis with combined multi-cloud mock data file."""
        sample_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'cca_cost_inv_sample.json'
        )
        summary_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'cca_cost_sum_sample.json'
        )

        if os.path.exists(sample_path) and os.path.exists(summary_path):
            inventory = load_json(sample_path)
            summary = load_json(summary_path)

            analysis = analyze_costs(inventory, summary)

            # Verify multi-cloud structure
            assert analysis['totals']['total_cost'] > 0
            assert len(analysis['by_provider']) >= 2  # Should have multiple providers
            assert len(analysis['by_category']) >= 1
            assert len(analysis['records']) > 0

    def test_analyze_aws_fixture_data(self):
        """Test analysis with AWS-only fixture data (typical customer workflow)."""
        sample_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'aws',
            'cca_cost_inv_sample.json'
        )
        summary_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'aws',
            'cca_cost_sum_sample.json'
        )

        if os.path.exists(sample_path) and os.path.exists(summary_path):
            inventory = load_json(sample_path)
            summary = load_json(summary_path)

            analysis = analyze_costs(inventory, summary)

            # Verify AWS-specific data
            assert 'aws' in analysis['by_provider']
            assert analysis['by_provider']['aws']['cost'] > 0

            # Verify monthly trend (should have 3 months)
            assert len(analysis['monthly_trend']) == 3

    def test_analyze_azure_fixture_data(self):
        """Test analysis with Azure-only fixture data (typical customer workflow)."""
        sample_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'azure',
            'cca_cost_inv_sample.json'
        )
        summary_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'azure',
            'cca_cost_sum_sample.json'
        )

        if os.path.exists(sample_path) and os.path.exists(summary_path):
            inventory = load_json(sample_path)
            summary = load_json(summary_path)

            analysis = analyze_costs(inventory, summary)

            # Verify Azure-specific data
            assert 'azure' in analysis['by_provider']
            assert len(analysis['by_provider']) == 1  # Only Azure
            assert analysis['by_provider']['azure']['cost'] > 0

            # Verify monthly trend (should have 3 months)
            assert len(analysis['monthly_trend']) == 3

    def test_analyze_gcp_fixture_data(self):
        """Test analysis with GCP-only fixture data (typical customer workflow)."""
        sample_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'gcp',
            'cca_cost_inv_sample.json'
        )
        summary_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'gcp',
            'cca_cost_sum_sample.json'
        )

        if os.path.exists(sample_path) and os.path.exists(summary_path):
            inventory = load_json(sample_path)
            summary = load_json(summary_path)

            analysis = analyze_costs(inventory, summary)

            # Verify GCP-specific data
            assert 'gcp' in analysis['by_provider']
            assert len(analysis['by_provider']) == 1  # Only GCP
            assert analysis['by_provider']['gcp']['cost'] > 0

            # Verify monthly trend (should have 3 months)
            assert len(analysis['monthly_trend']) == 3


# =============================================================================
# Excel Generation Tests (if openpyxl available)
# =============================================================================

class TestExcelGeneration:
    """Tests for Excel report generation."""

    @pytest.fixture
    def temp_output_path(self):
        """Create a temporary output path for Excel files."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
        yield temp_path
        if os.path.exists(temp_path):
            os.unlink(temp_path)

    def test_generate_report_creates_file(self, temp_cost_inventory_file, temp_cost_summary_file, temp_output_path):
        """Test that generate_excel_report creates an Excel file."""
        try:
            from scripts.generate_cost_report import OPENPYXL_AVAILABLE, generate_excel_report
        except ImportError:
            pytest.skip("generate_excel_report not available")

        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        generate_excel_report(temp_cost_inventory_file, temp_cost_summary_file, temp_output_path)

        assert os.path.exists(temp_output_path)
        assert os.path.getsize(temp_output_path) > 0

    def test_generate_report_has_expected_sheets(self, temp_cost_inventory_file, temp_cost_summary_file, temp_output_path):
        """Test that the generated report has expected worksheets."""
        try:
            from openpyxl import load_workbook

            from scripts.generate_cost_report import OPENPYXL_AVAILABLE, generate_excel_report
        except ImportError:
            pytest.skip("Required modules not available")

        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        generate_excel_report(temp_cost_inventory_file, temp_cost_summary_file, temp_output_path)

        wb = load_workbook(temp_output_path)
        sheet_names = wb.sheetnames

        expected_sheets = [
            'Executive Summary',
            'Cost by Provider',
            'Cost by Category',
            'Cost by Service'
        ]

        for expected in expected_sheets:
            assert expected in sheet_names, f"Missing sheet: {expected}"

    def test_generate_report_with_sample_data(self, temp_output_path):
        """Test report generation with combined multi-cloud sample data files."""
        try:
            from scripts.generate_cost_report import OPENPYXL_AVAILABLE, generate_excel_report
        except ImportError:
            pytest.skip("generate_excel_report not available")

        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        inv_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'cca_cost_inv_sample.json'
        )
        sum_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'cca_cost_sum_sample.json'
        )

        if not os.path.exists(inv_path) or not os.path.exists(sum_path):
            pytest.skip("Sample data files not found")

        generate_excel_report(inv_path, sum_path, temp_output_path)

        assert os.path.exists(temp_output_path)
        assert os.path.getsize(temp_output_path) > 1000  # Should be a reasonable size

    def test_generate_aws_only_report(self, temp_output_path):
        """Test report generation with AWS-only data (typical customer workflow)."""
        try:
            from scripts.generate_cost_report import OPENPYXL_AVAILABLE, generate_excel_report
        except ImportError:
            pytest.skip("generate_excel_report not available")

        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        inv_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'aws',
            'cca_cost_inv_sample.json'
        )
        sum_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'aws',
            'cca_cost_sum_sample.json'
        )

        if not os.path.exists(inv_path) or not os.path.exists(sum_path):
            pytest.skip("AWS fixture files not found")

        generate_excel_report(inv_path, sum_path, temp_output_path)

        assert os.path.exists(temp_output_path)
        assert os.path.getsize(temp_output_path) > 1000

    def test_generate_azure_only_report(self, temp_output_path):
        """Test report generation with Azure-only data (typical customer workflow)."""
        try:
            from scripts.generate_cost_report import OPENPYXL_AVAILABLE, generate_excel_report
        except ImportError:
            pytest.skip("generate_excel_report not available")

        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        inv_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'azure',
            'cca_cost_inv_sample.json'
        )
        sum_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'azure',
            'cca_cost_sum_sample.json'
        )

        if not os.path.exists(inv_path) or not os.path.exists(sum_path):
            pytest.skip("Azure fixture files not found")

        generate_excel_report(inv_path, sum_path, temp_output_path)

        assert os.path.exists(temp_output_path)
        assert os.path.getsize(temp_output_path) > 1000

    def test_generate_gcp_only_report(self, temp_output_path):
        """Test report generation with GCP-only data (typical customer workflow)."""
        try:
            from scripts.generate_cost_report import OPENPYXL_AVAILABLE, generate_excel_report
        except ImportError:
            pytest.skip("generate_excel_report not available")

        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not installed")

        inv_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'gcp',
            'cca_cost_inv_sample.json'
        )
        sum_path = os.path.join(
            os.path.dirname(__file__),
            'fixtures',
            'gcp',
            'cca_cost_sum_sample.json'
        )

        if not os.path.exists(inv_path) or not os.path.exists(sum_path):
            pytest.skip("GCP fixture files not found")

        generate_excel_report(inv_path, sum_path, temp_output_path)

        assert os.path.exists(temp_output_path)
        assert os.path.getsize(temp_output_path) > 1000
